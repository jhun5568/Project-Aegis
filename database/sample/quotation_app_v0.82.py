# 고급 모델 검색 시스템 및 기타 함수들은 기존과 동일하게 유지
import streamlit as st

def _read_sheet_robust(xlsx_path, primary, alternates=None):
    import pandas as _pd
    alternates = alternates or []
    try:
        return _pd.read_excel(xlsx_path, sheet_name=primary)
    except Exception:
        pass
    for alt in alternates:
        try:
            return _pd.read_excel(xlsx_path, sheet_name=alt)
        except Exception:
            continue
    xl = _pd.ExcelFile(xlsx_path)
    strip_map = {n.strip(): n for n in xl.sheet_names}
    lower_map = {n.strip().lower(): n for n in xl.sheet_names}
    p = primary.strip()
    if p in strip_map:
        return _pd.read_excel(xlsx_path, sheet_name=strip_map[p])
    pl = p.lower()
    if pl in lower_map:
        return _pd.read_excel(xlsx_path, sheet_name=lower_map[pl])
    for n in xl.sheet_names:
        if n.strip().lower().startswith(pl):
            return _pd.read_excel(xlsx_path, sheet_name=n)
    raise FileNotFoundError(f"Worksheet named '{primary}' (or alternates) not found. Available: {xl.sheet_names}")

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
import io
import re
from difflib import SequenceMatcher

def _to_float_safe(val, default=0.0):
    try:
        # Handle strings with commas or spaces
        if isinstance(val, str):
            val = val.replace(',', '').strip()
        return float(val)
    except Exception:
        return float(default)

# 페이지 설정
st.set_page_config(
    page_title="업무자동화 시스템",
    page_icon="🎨",
    layout="wide",
    initial_sidebar_state="expanded"
)

class DoohoQuotationSystem:
    """업무자동화 시스템"""
    
    def __init__(self):
        self.database_file = 'material_database.xlsx'
        self.load_data()
    
    @st.cache_data
    def load_data(_self):
        """데이터 로딩 (캐시 적용)"""
        try:
            data = {}
            data['models'] = _read_sheet_robust(_self.database_file, 'Models', ['Models '])
            data['pricing'] = _read_sheet_robust(_self.database_file, 'Pricing', ['Pricing '])
            data['bom'] = _read_sheet_robust(_self.database_file, 'BOM', ['BOM '])
            data['inventory'] = _read_sheet_robust(_self.database_file, 'Inventory', ['Inventory '])
            data['main_materials'] = _read_sheet_robust(_self.database_file, 'main_Materials', ['main_Materials '])
            data['sub_materials'] = _read_sheet_robust(_self.database_file, 'sub_Materials', ['sub_Materials '])
            return data
        except Exception as e:
            st.error(f"데이터베이스 로딩 오류: {e}")
            return None
    
    def search_model_price(self, model_name):
        """모델 단가 검색"""
        data = self.load_data()
        if not data:
            return None
            
        pricing_df = data['pricing']
        
        model_clean = str(model_name).strip()
        exact_match = pricing_df[pricing_df['모델명'].str.strip() == model_clean]
        if not exact_match.empty:
            return exact_match.iloc[0]
        
        return None
    
    def generate_quotation(self, site_info, items, contract_type="관급"):
        """견적서 생성"""
        quotation_items = []
        total_supply_price = 0
        
        for item in items:
            price_info = self.search_model_price(item['model_name'])
            
            if price_info is None:
                st.warning(f"'{item['model_name']}' 모델의 단가를 찾을 수 없습니다.")
                continue
            
            unit_price = _to_float_safe(price_info.get('단가', 0), 0)
            supply_amount = _to_float_safe(item.get('quantity', 0), 0) * _to_float_safe(unit_price, 0)
            
            quotation_items.append({
                'model_name': item['model_name'],
                'specification': price_info['규격'],
                'unit': price_info['단위'],
                'quantity': item['quantity'],
                'unit_price': unit_price,
                'supply_amount': supply_amount,
                'notes': item.get('notes', ''),
                '식별번호': price_info.get('식별번호', '')
            })
            
            total_supply_price += supply_amount
        
        vat_amount = total_supply_price * 0.1
        total_amount = total_supply_price + vat_amount
        
        return {
            'site_info': site_info,
            'contract_type': contract_type,
            'items': quotation_items,
            'total_supply_price': total_supply_price,
            'vat_amount': vat_amount,
            'total_amount': total_amount,
            'created_date': datetime.now(),
            'company': '회사명'
        }
    
    def generate_purchase_items_from_quotation(self, quotation_data):
        """견적서 데이터를 기반으로 발주 항목 생성 (카테고리 기반)"""
        data = self.load_data()
        purchase_items = []

        # === BOM 확인 코드 ===
        st.write("📋 선택된 모델들의 BOM 확인:")
        for item in quotation_data['items']:
            model_info = data['models'][data['models']['model_name'] == item['model_name']]
            if not model_info.empty:
                model_id = model_info.iloc[0]['model_id']
                model_bom = data['bom'][data['bom']['model_id'] == model_id]
                st.write(f"- {item['model_name']} (ID: {model_id}): BOM 행수 {len(model_bom)}")
                if model_bom.empty:
                    st.error(f"  ⚠️ {item['model_name']}의 BOM 데이터가 없습니다!")
        # === 확인 코드 끝 ===

        for item in quotation_data['items']:
            model_info = data['models'][data['models']['model_name'] == item['model_name']]
            
            if not model_info.empty:
                model_id = model_info.iloc[0]['model_id']
                model_bom = data['bom'][data['bom']['model_id'] == model_id]
                
                for _, bom_item in model_bom.iterrows():
                    # 기본 필요 수량 계산
                    required_quantity = item['quantity'] * float(bom_item['quantity'])
                    
                    # 카테고리가 PIPE를 포함하는 경우 개수로 변환
                    if 'PIPE' in str(bom_item['category']).upper():
                        required_quantity = self._calculate_pipe_count(
                            required_quantity, 
                            bom_item['standard'], 
                            data
                        )
                        unit = 'EA'
                    else:
                        unit = bom_item['unit']
                    
                    # 기존 발주 항목에서 같은 자재가 있는지 확인
                    existing_item = None
                    for purchase_item in purchase_items:
                        if (purchase_item['material_name'] == bom_item['material_name'] and
                            purchase_item['standard'] == bom_item['standard']):
                            existing_item = purchase_item
                            break
                    
                    if existing_item:
                        existing_item['quantity'] += required_quantity
                    else:
                        purchase_items.append({
                            'material_name': bom_item['material_name'],
                            'standard': bom_item['standard'],
                            'unit': unit,
                            'quantity': required_quantity,
                            'category': bom_item['category'],
                            'model_reference': item['model_name']
                        })
        
        return purchase_items

    def create_material_execution_report(self, quotation_data, delivery_date=None):
            """자재발실행내역서 자동생성 - 템플릿 경로 수정"""
            try:
                # 템플릿 파일 경로 수정 - 여러 위치에서 찾기
                template_paths = [
                    '../templates/자재 및 실행내역서템플릿_v2.0_20250919.xlsx',  # database 폴더에서 실행시
                    'templates/자재 및 실행내역서템플릿_v2.0_20250919.xlsx',     # 루트에서 실행시
                    '자재 및 실행내역서템플릿_v2.0_20250919.xlsx',               # 같은 폴더에 있을 때
                    '../자재 및 실행내역서템플릿_v2.0_20250919.xlsx'             # 상위 폴더에 있을 때
                ]
                
                template_path = None
                for path in template_paths:
                    if os.path.exists(path):
                        template_path = path
                        break
                
                if template_path is None:
                    st.error("템플릿 파일을 찾을 수 없습니다. 다음 위치 중 한 곳에 '자재 및 실행내역서템플릿_v2.0_20250919.xlsx' 파일이 있는지 확인해주세요:")
                    for path in template_paths:
                        st.write(f"• {os.path.abspath(path)}")
                    return None, []
                
                st.info(f"템플릿 파일 발견: {template_path}")
                
                workbook = load_workbook(template_path)
                material_sheet = workbook['자재내역서']
                
                
                # 기본 정보 입력 - 템플릿 위치에 맞게 수정
                site_name = quotation_data['site_info']['site_name']
                total_quantity = sum(item['quantity'] for item in quotation_data['items'])
                
                # 현장명 입력 (A3 다음 셀에)
                material_sheet['B3'] = site_name  # 현장명
                
                # 수량 입력 (E3 다음 셀에)  
                material_sheet['F3'] = total_quantity  # 수량
                
                # 납품기한 입력 (A5 다음 셀에)
                if delivery_date:
                    material_sheet['B5'] = delivery_date.strftime('%Y년 %m월 %d일')
                else:
                    material_sheet['B5'] = (datetime.now() + pd.Timedelta(days=7)).strftime('%Y년 %m월 %d일')
                
                # BOM 기반 자재 데이터 생성
                data = self.load_data()
                material_items = self._generate_material_items_with_pricing(quotation_data, data)
                
                # 자재내역서 데이터 입력 (9행부터 시작)
                start_row = 9
                
                for idx, material in enumerate(material_items):
                    row = start_row + idx
                    
                    # 모델 헤더인 경우 특별 처리
                    if material.get('is_header', False):
                        material_sheet[f'A{row}'] = idx + 1  # No.
                        material_sheet[f'B{row}'] = material['model_name']  # 모델명
                        material_sheet[f'C{row}'] = ''  # 규격 빈칸
                        material_sheet[f'D{row}'] = ''  # 단위 빈칸
                        material_sheet[f'E{row}'] = ''  # 수량 빈칸
                        material_sheet[f'F{row}'] = ''  # 단가 빈칸
                        material_sheet[f'G{row}'] = ''  # 공급가 빈칸
                        material_sheet[f'H{row}'] = ''  # 비고 빈칸
                        material_sheet[f'I{row}'] = ''  # 하차지 빈칸
                        material_sheet[f'J{row}'] = ''  # 발주일 빈칸
                        material_sheet[f'K{row}'] = ''  # 발주처 빈칸
                    else:
                        # 일반 자재 행 처리
                        material_sheet[f'A{row}'] = idx + 1  # No.
                        material_sheet[f'B{row}'] = material['material_name']  # 품목
                        material_sheet[f'C{row}'] = material['standard']  # 규격
                        material_sheet[f'D{row}'] = material['unit']  # 단위
                        material_sheet[f'E{row}'] = material['quantity']  # 수량
                        unit_price = material.get('unit_price', 0)
                        material_sheet[f'F{row}'] = _to_float_safe(unit_price, 0)  # 단가
                        material_sheet[f'G{row}'] = _to_float_safe(material.get('quantity', 0), 0) * _to_float_safe(unit_price, 0)  # 공급가
                        material_sheet[f'H{row}'] = material.get('notes', '')  # 비고
                        material_sheet[f'I{row}'] = '공장'  # 하차지
                        material_sheet[f'J{row}'] = datetime.now().strftime('%Y-%m-%d')  # 발주일
                        material_sheet[f'K{row}'] = '공급업체명'  # 발주처
                
                # 메모리에 저장
                excel_buffer = io.BytesIO()
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
                
                return excel_buffer, material_items
                
            except Exception as e:
                st.error(f"자재 및 실행내역서 생성 오류: {e}")
                return None, []

    def _generate_material_items_with_pricing(self, quotation_data, data):
        """BOM 데이터에 단가 정보를 결합한 자재 목록 생성 - 비고란 추가"""
        material_items_by_model = {}  # 모델별로 그룹화
        
        for item in quotation_data['items']:        
            model_name = item['model_name']
            model_info = data['models'][data['models']['model_name'] == model_name]
            
            if not model_info.empty:
                model_id = model_info.iloc[0]['model_id']
                model_bom = data['bom'][data['bom']['model_id'] == model_id]
                
                # 모델별 자재 리스트 초기화
                if model_name not in material_items_by_model:
                    material_items_by_model[model_name] = []
                
                for _, bom_item in model_bom.iterrows():
                    category = str(bom_item['category'])
                    material_name = bom_item['material_name'] 
                    bom_standard = bom_item['standard']
                    
                    # 기본 필요 수량 계산
                    required_quantity = item['quantity'] * float(bom_item['quantity'])
                    original_meter_quantity = required_quantity  # 미터 수량 보관
                    
                    # PIPE인 경우 처리 - 자재내역서는 미터, 비고는 개수 정보
                    if 'PIPE' in category.upper():
                        pipe_count = self._calculate_pipe_count(
                            required_quantity, 
                            bom_standard, 
                            data
                        )
                        # 자재내역서에는 미터 단위로 표시
                        final_quantity = original_meter_quantity
                        unit = 'M'
                        ea_quantity_for_notes = pipe_count
                    else:
                        final_quantity = required_quantity
                        unit = bom_item['unit']
                        ea_quantity_for_notes = required_quantity

                    # 강화된 자재 정보 찾기 (자재명 포함)
                    material_info = self._find_material_info_by_category(
                        category, bom_standard, data, material_name
                    )
                    
                    if material_info and _to_float_safe(material_info.get('단가', 0), 0) > 0:
                        actual_standard = material_info['규격']  # BOM 원본 규격 사용
                        unit_price = _to_float_safe(material_info.get('단가', 0), 0)
                    else:
                        actual_standard = bom_standard
                        unit_price = 0
                    
                    # 🔧 새로 추가: 비고란 생성
                    notes = self._generate_notes_for_material(
                        category, material_name, ea_quantity_for_notes, 
                        original_meter_quantity, data
                    )
                    
                    # 모델 내에서 중복 자재 체크
                    existing_item = None
                    for existing in material_items_by_model[model_name]:
                        if (existing['material_name'] == material_name and 
                            existing['standard'] == actual_standard):
                            existing_item = existing
                            break
                    
                    if existing_item:
                        existing_item['quantity'] += required_quantity
                        # 수량이 변경되었으므로 비고도 다시 계산
                        existing_item['notes'] = self._generate_notes_for_material(
                            category, material_name, existing_item['quantity'], 
                            existing_item['quantity'], data  # 이미 합산된 수량 사용
                        )
                    else:
                        material_items_by_model[model_name].append({
                            'material_name': material_name,
                            'standard': actual_standard,  # 🔧 수정: 파이프 길이 정보 제거
                            'unit': unit,
                            'quantity': final_quantity,
                            'category': category,
                            'unit_price': unit_price,
                            'model_name': model_name,
                            'notes': notes  # 🔧 새로 추가
                        })
        
        # 모델별 그룹화된 데이터를 하나의 리스트로 통합
        final_material_items = []
        
        for model_name, model_materials in material_items_by_model.items():
            # 모델 헤더 추가
            final_material_items.append({
                'material_name': f"=== 모델: {model_name} ===",
                'standard': '',
                'unit': '',
                'quantity': 0,
                'category': 'MODEL_HEADER',
                'unit_price': 0,
                'model_name': model_name,
                'notes': '',
                'is_header': True
            })
            
            # 해당 모델의 자재들 추가
            final_material_items.extend(model_materials)

        
        return final_material_items

    def _generate_notes_for_material(self, category, material_name, ea_quantity, meter_quantity, data):
        """자재별 비고 생성 - 파이프 소모량 계산"""
        # PIPE 카테고리인 경우만 파이프 소모량 표기
        if 'PIPE' in category.upper():
            # main_Materials에서 카테고리별로 매칭
            main_materials = data.get('main_materials', pd.DataFrame())
            
            if not main_materials.empty:
                # 카테고리에 따라 품목 매칭
                category_mapping = {
                    'HGI PIPE': 'HGI PIPE',
                    'STS PIPE': 'STS PIPE', 
                    'STL PIPE': 'STL PIPE'
                }
                
                target_item = category_mapping.get(category)
                if target_item:
                    # 해당 카테고리의 파이프 찾기
                    pipe_match = main_materials[
                        main_materials['품목'].astype(str).str.contains(target_item, na=False, case=False)
                    ]
                    
                    if not pipe_match.empty:
                        # 표준 파이프 길이 가져오기 (기본 6m)
                        pipe_length = 6.0
                        if '파이프길이(m)' in pipe_match.columns:
                            try:
                                # 첫 번째 매칭된 항목의 길이 사용
                                length_value = pipe_match.iloc[0]['파이프길이(m)']
                                if pd.notna(length_value) and _to_float_safe(length_value, 0) > 0:
                                    pipe_length = float(length_value)
                            except:
                                pipe_length = 6.0
                        
                        # 파이프 소모량: 15.5m (6m*3본)
                        notes = f"파이프 소모량: {pipe_length:.0f}m*{ea_quantity:.0f}본"
                        return notes
        
        return ""  # 파이프가 아니거나 매칭되지 않으면 빈 문자열
    # 3. 필요한 보조 함수들 (기존 함수들과 함께 추가)

    def _compare_specs_order_agnostic(self, bom_spec, main_spec):
        """순서 무관 규격 비교"""
        if not bom_spec or not main_spec:
            return False
        
        if self._compare_complete_specs(bom_spec, main_spec):
            return True
        
        return self._compare_with_reversed_dimensions(bom_spec, main_spec)

    def _compare_with_reversed_dimensions(self, bom_spec, main_spec):
        """치수 순서를 바꿔서 비교"""
        import re
        
        bom_match = re.match(r'(\d+)\*(\d+)\*(.+)', bom_spec)
        main_match = re.match(r'(\d+)\*(\d+)\*(.+)', main_spec)
        
        if bom_match and main_match:
            bom_dim1, bom_dim2, bom_thickness = bom_match.groups()
            main_dim1, main_dim2, main_thickness = main_match.groups()
            
            if bom_thickness.strip() == main_thickness.strip():
                if ((bom_dim1 == main_dim2 and bom_dim2 == main_dim1) or
                    (bom_dim1 == main_dim1 and bom_dim2 == main_dim2)):
                    return True
        
        return False

    def _create_empty_result(self):
        """빈칸 결과 생성"""
        return {
            '완전규격': '',
            '단가': '',
            '품목': '',
            '규격': ''
        }

    def _create_material_result_from_main(self, material_row, category):
        """main_Materials 결과 생성"""
        main_spec = str(material_row['규격']).strip()
        pipe_length = _to_float_safe(material_row.get('파이프길이(m)', 6.0), 6.0)
        unit_price = _to_float_safe(material_row.get('단가', 0), 0)
        
        # 파이프인 경우 m당 단가로 변환
        if any(pipe_word in category.upper() for pipe_word in ['PIPE', '파이프']):
            unit_price = unit_price / pipe_length if pipe_length > 0 else unit_price     

        if any(pipe_word in category.upper() for pipe_word in ['PIPE', '파이프']):
            full_specification = f"{main_spec}×{pipe_length}m"
        else:
            full_specification = main_spec
        
        return {
            '완전규격': full_specification,
            '단가': unit_price,
            '품목': material_row['품목'],
            '규격': material_row['규격']
        }

    def _create_material_result_from_sub(self, material_row):
        """sub_Materials 결과 생성"""
        unit_price = _to_float_safe(material_row.get('단가', 0), 0)
        spec = str(material_row['규격']).strip()
        
        return {
            '완전규격': spec,
            '단가': unit_price,
            '품목': material_row['품목'],
            '규격': material_row['규격']
        }

    def _compare_complete_specs(self, bom_spec, main_spec):
        """완전한 규격 비교"""
        if not bom_spec or not main_spec:
            return False
        
        bom_clean = str(bom_spec).strip()
        main_clean = str(main_spec).strip()
        
        if bom_clean == main_clean:
            return True
        
        bom_normalized = self._normalize_special_chars(bom_clean)
        main_normalized = self._normalize_special_chars(main_clean)
        
        return bom_normalized == main_normalized

    def _normalize_special_chars(self, spec):
        """특수문자 정규화"""
        normalized = spec.replace('∅', 'Ø').replace('Φ', 'Ø').replace('φ', 'Ø')
        normalized = normalized.upper()
        return normalized
            
        return material_items

    def _find_material_info_by_category(self, category, standard, data, material_name=None):
            """카테고리로 자재 정보 찾기 - 자재명 포함 강화된 검색"""
            
            # 1. main_materials에서 검색
            if 'main_materials' in data:
                main_materials = data['main_materials']
                try:
                    category_match = main_materials[
                        main_materials['품목'].astype(str).str.strip() == str(category).strip()
                    ]
                    
                    if not category_match.empty:
                        bom_standard = str(standard).strip()
                        
                        for _, material_row in category_match.iterrows():
                            main_spec = str(material_row['규격']).strip() if pd.notna(material_row['규격']) else ''
                            
                            if self._compare_specs_order_agnostic(bom_standard, main_spec):
                                st.success(f"✅ main_materials 매칭: `{main_spec}`")
                                return self._create_material_result_from_main(material_row, category)
                        
                        st.error(f"❌ `{category}`에서 `{standard}` 매칭 실패 → sub_materials 검색")
                    else:
                        st.info(f"💡 `{category}` 카테고리가 main_materials에 없음 → sub_materials 검색")
                        
                except Exception as e:
                    st.error(f"main_materials 검색 오류: {e}")
            
            # 2. sub_materials에서 강화된 검색
            if 'sub_materials' in data:
                sub_materials = data['sub_materials']
                try:
                    # 2-1. 카테고리명으로 검색
                    category_match = sub_materials[
                        sub_materials['품목'].astype(str).str.contains(str(category), na=False, case=False)
                    ]
                    
                    if not category_match.empty:
                        for _, material_row in category_match.iterrows():
                            sub_spec = str(material_row['규격']).strip() if pd.notna(material_row['규격']) else ''
                            if str(standard) in sub_spec or sub_spec in str(standard):
                                st.success(f"✅ sub_materials 카테고리 매칭: `{material_row['품목']} {sub_spec}`")
                                return self._create_material_result_from_sub(material_row)
                    
                    # 2-2. 자재명으로 검색 (새로 추가)
                    if material_name:
                        material_name_match = sub_materials[
                            sub_materials['품목'].astype(str).str.contains(str(material_name), na=False, case=False)
                        ]
                        
                        if not material_name_match.empty:
                            material_row = material_name_match.iloc[0]
                            st.success(f"✅ sub_materials 자재명 매칭: `{material_row['품목']} {material_row['규격']}`")
                            return self._create_material_result_from_sub(material_row)
                    
                    # 2-3. 규격으로 직접 검색
                    standard_match = sub_materials[
                        sub_materials['규격'].astype(str).str.contains(str(standard), na=False, case=False)
                    ]
                    
                    if not standard_match.empty:
                        material_row = standard_match.iloc[0]
                        st.info(f"✅ sub_materials 규격 매칭: `{material_row['품목']} {material_row['규격']}`")
                        return self._create_material_result_from_sub(material_row)
                        
                except Exception as e:
                    st.warning(f"sub_materials 검색 오류: {e}")
            
            # 3. 완전 매칭 실패 - 빈칸 처리
            st.error(f"❌ `{category}` - `{standard}` 어디서도 찾을 수 없음 → 빈칸 처리")
            return self._create_empty_result()
    
    def _calculate_pipe_count(self, required_length_m, pipe_standard, data):
        """파이프 길이를 고려한 실제 발주 개수 계산"""
        import math
        
        # main_Materials에서 해당 파이프의 표준 길이 찾기
        main_materials = data['main_materials']
        
        # 파이프 규격으로 검색
        pipe_match = main_materials[
            main_materials['규격'].str.contains(pipe_standard, na=False, case=False)
        ]
        
        if not pipe_match.empty:
            # 길이 정보가 있으면 사용 (보통 6m)
            try:
                standard_length = 6.0  # 기본값
                # 만약 길이 컬럼이 있다면
                if '길이' in pipe_match.columns:
                    standard_length = float(pipe_match.iloc[0]['길이'])
                elif '단위길이' in pipe_match.columns:
                    standard_length = float(pipe_match.iloc[0]['단위길이'])
                elif '파이프길이(m)' in pipe_match.columns:
                    standard_length = float(pipe_match.iloc[0]['파이프길이(m)'])
            except:
                standard_length = 6.0
        else:
            standard_length = 6.0  # 기본값: 6m
        
        # 필요한 파이프 개수 계산 (올림)
        required_pipes = math.ceil(required_length_m / standard_length)
        
        return required_pipes    

    def _get_specification_with_length_fixed(self, material_name, standard, data):
        """규격에 파이프 길이 정보 추가 - 카테고리 기반"""
        # BOM에서 해당 자재의 카테고리 확인
        bom_data = data['bom']
        material_bom = bom_data[bom_data['material_name'] == material_name]
        
        is_pipe = False
        if not material_bom.empty:
            category = str(material_bom.iloc[0]['category']).upper()
            is_pipe = 'PIPE' in category
        
        if is_pipe:
            main_materials = data['main_materials']
            
            # 해당 규격의 파이프 찾기
            pipe_match = main_materials[
                main_materials['규격'].astype(str).str.contains(str(standard), na=False, case=False)
            ]
            
            # 길이 정보 찾기
            pipe_length = 6.0  # 기본값
            
            if not pipe_match.empty:
                if '파이프길이(m)' in pipe_match.columns:
                    try:
                        length_value = pipe_match.iloc[0]['파이프길이(m)']
                        if pd.notna(length_value) and _to_float_safe(length_value, 0) > 0:
                            pipe_length = float(length_value)
                    except:
                        pipe_length = 6.0
            
            # 규격에 길이 정보 추가
            return f"{standard}×{pipe_length}m"
        
        return standard

    def create_purchase_orders_by_material(self, quotation_data, delivery_location="현장", supplier_name=""):
        """재질별로 발주서 분리 생성 - 공급업체 직접 입력 방식"""
        try:
            data = self.load_data()
            purchase_items = self.generate_purchase_items_from_quotation(quotation_data)
            
            # 재질별로 그룹화
            material_groups = self._group_by_material_type(purchase_items, data)
            
            purchase_orders = []
            
            for material_type, items in material_groups.items():
                # 공급업체명이 입력되지 않은 경우 재질명 사용
                actual_supplier_name = supplier_name if supplier_name.strip() else material_type
                
                supplier_info = {'company_name': actual_supplier_name}
                
                # 발주서 생성
                excel_buffer = self._create_single_purchase_order(
                    quotation_data, items, delivery_location, supplier_info
                )
                
                if excel_buffer:
                    purchase_orders.append({
                        'material_type': material_type,
                        'supplier': supplier_info['company_name'],
                        'excel_buffer': excel_buffer,
                        'items': items,
                        'filename': f"발주서_{supplier_info['company_name']}_{quotation_data['site_info']['site_name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    })
            
            return purchase_orders
            
        except Exception as e:
            st.error(f"재질별 발주서 생성 오류: {e}")
            return []

    def _group_by_material_type(self, purchase_items, data):
        """재질별로 발주 항목 그룹화 - 카테고리 기반"""
        material_groups = {}
        
        for item in purchase_items:
            # 카테고리에서 재질 정보 추출
            category = str(item['category']).upper()
            
            if 'HGI' in category or '아연도' in category:
                material_type = '아연도'
            elif 'STS' in category:
                material_type = 'STS'
            else:
                # 기존 로직으로 fallback
                material_type = self._find_material_type(item['material_name'], item['standard'], data)
            
            if material_type not in material_groups:
                material_groups[material_type] = []
            
            material_groups[material_type].append(item)
        
        return material_groups

    def _find_material_type(self, material_name, standard, data):
        """자재의 재질 타입 확인"""
        # main_Materials에서 재질 정보 찾기
        main_materials = data['main_materials']
        
        # 컬럼명 확인 및 매칭
        possible_item_columns = ['품목', 'Item', 'item_name', 'material_name', '자재명']
        possible_spec_columns = ['규격', 'Spec', 'specification', 'standard', '사양']
        
        item_column = None
        spec_column = None
        
        # 존재하는 컬럼 찾기
        for col in possible_item_columns:
            if col in main_materials.columns:
                item_column = col
                break
        
        for col in possible_spec_columns:
            if col in main_materials.columns:
                spec_column = col
                break
        
        if item_column and spec_column:
            material_match = main_materials[
                (main_materials[item_column].str.contains(material_name, na=False)) |
                (main_materials[spec_column].str.contains(standard, na=False))
            ]
            
            if not material_match.empty:
                material_info = material_match.iloc[0]
                # 재질 컬럼도 여러 가능성 확인
                possible_material_columns = ['재질', 'Material', 'material_type', '소재']
                
                for mat_col in possible_material_columns:
                    if mat_col in material_info:
                        material_type = material_info[mat_col]
                        if 'STS' in str(material_type).upper():
                            return 'STS'
                        elif '아연도' in str(material_type):
                            return '아연도'
        
        # 기본값
        return '아연도'

    def _create_single_purchase_order(self, quotation_data, purchase_items, delivery_location, supplier_info):
        """단일 발주서 생성 (파이프 길이 정보 포함) - 수정된 버전"""
        try:
            template_path = 'templates/발주서템플릿_v2.0_20250919.xlsx'
            if not os.path.exists(template_path):
                template_path = '발주서템플릿_v2.0_20250919.xlsx'
            
            workbook = load_workbook(template_path)
            sheet = workbook['발주서']
            
            # 기본 정보 입력
            today = datetime.now()
            sheet['F4'] = today.strftime('%Y년 %m월 %d일')
            sheet['B6'] = supplier_info['company_name']  # 수신 업체
            
            site_name = quotation_data['site_info']['site_name']
            start_row = 11
            
            data = self.load_data()
            
            for idx, purchase_item in enumerate(purchase_items):
                row = start_row + idx
                
                # 규격에 파이프 길이 정보 추가 - 수정된 함수 사용
                specification = self._get_specification_with_length_fixed(
                    purchase_item['material_name'], 
                    purchase_item['standard'], 
                    data
                )
                
                sheet[f'A{row}'] = idx + 1
                sheet[f'B{row}'] = purchase_item['material_name']
                sheet[f'C{row}'] = specification  # 길이 정보 포함된 규격
                sheet[f'D{row}'] = purchase_item['unit']
                sheet[f'E{row}'] = purchase_item['quantity']
                sheet[f'F{row}'] = delivery_location
                sheet[f'G{row}'] = site_name
                sheet[f'H{row}'] = f"모델: {purchase_item['model_reference']}"
            
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            st.error(f"발주서 생성 오류: {e}")
            return None

    def create_template_quotation(self, quotation_data):
        """템플릿 기반 견적서 생성"""
        try:
            template_path = 'templates/견적서템플릿_v2.0_20250919.xlsx'
            if not os.path.exists(template_path):
                template_path = '견적서템플릿_v2.0_20250919.xlsx'
            
            workbook = load_workbook(template_path)
            
            if quotation_data['contract_type'] == '사급':
                sheet = workbook['사급견적서']
                start_row = 13
                columns = {
                    'item': 'B', 'spec': 'C', 'unit': 'D', 'qty': 'E',
                    'price': 'F', 'supply': 'G', 'vat': 'H'
                }
            else:  # 관급
                sheet = workbook['관급견적서']
                start_row = 14
                columns = {
                    'item': 'B', 'spec': 'D', 'unit': 'E', 'qty': 'F',
                    'price': 'G', 'amount': 'H', 'id_num': 'I'
                }
            
            # 견적 항목 데이터 입력
            for idx, item in enumerate(quotation_data['items']):
                row = start_row + idx
                
                # 각 셀에 값을 직접 할당
                sheet[f"{columns['item']}{row}"] = item['model_name']
                sheet[f"{columns['spec']}{row}"] = item['specification']
                sheet[f"{columns['unit']}{row}"] = item['unit']
                sheet[f"{columns['qty']}{row}"] = item['quantity']
                sheet[f"{columns['price']}{row}"] = item['unit_price']
                
                # 관급인 경우에만 식별번호 추가
                if quotation_data['contract_type'] == '관급' and '식별번호' in item:
                    sheet[f"{columns['id_num']}{row}"] = item['식별번호']
            
            # 현장명 입력 (F3 위치)
            try:
                sheet['F3'] = quotation_data['site_info']['site_name']
            except:
                pass  # 현장명 입력 실패시 무시
            
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            st.error(f"템플릿 견적서 생성 오류: {e}")
            return None

    # 인터페이스 메소드들
    def create_material_execution_interface(self):
        """자재 및 실행내역서 생성 인터페이스 - 검색 기반으로 개선"""
        st.header("📋 자재 및 실행내역서 자동생성")
        
        # 현장 정보 입력
        col1, col2 = st.columns(2)
        with col1:
            site_name = st.text_input("현장명", value="OO아파트 조경공사")
            contract_type = st.selectbox("계약 유형", ["관급", "사급"])
        with col2:
            foundation = st.selectbox("기초 유형", ["기초형", "앙카형"])
            delivery_date = st.date_input("납품기한", datetime.now() + pd.Timedelta(days=7))
        
        st.subheader("🔘자재 항목 추가")
        
        # 동적 자재 항목
        if 'material_items' not in st.session_state:
            st.session_state.material_items = []
        
        # 모델 검색 및 선택
        with st.expander("➕ 새 항목 추가", expanded=len(st.session_state.material_items)==0):
            st.markdown("**모델 검색 및 선택**")
            
            # 검색창
            search_query = st.text_input(
                "모델 검색",
                placeholder="모델명, 식별번호, 차양, 볼라드, 자전거보관대 등 입력",
                help="예: '디자인' → 디자인형울타리 전체, 'DST' → DST 시리즈 전체",
                key="material_search"
            )
            
            # 검색 결과 표시 및 선택
            if search_query:
                data = self.load_data()
                if 'material_search_system' not in st.session_state:
                    st.session_state.material_search_system = EnhancedModelSearch(data['models'])
                
                search_system = st.session_state.material_search_system
                search_results = search_system.search_models(search_query)
                
                if not search_results.empty:
                    st.write(f"🔍 검색 결과: {len(search_results)}개 모델")

                    # 수량 입력을 먼저 위쪽에 배치
                    col1, col2 = st.columns(2)
                    with col1:
                        quantity = st.number_input("수량", min_value=1, value=50, key="bulk_quantity")
                    with col2:
                        notes = st.text_input("비고", key="bulk_notes")
                    
                    # 현재 표시할 개수 관리
                    if 'display_count' not in st.session_state:
                        st.session_state.display_count = 10
                    
                    current_count = min(st.session_state.display_count, len(search_results))
                    
                    st.markdown(f"**모델 선택 ({current_count}/{len(search_results)}개 표시):**")
                    
                    selected_models = []
                    for idx in range(current_count):
                        model = search_results.iloc[idx]
                        
                        col1, col2, col3 = st.columns([1, 3, 2])
                        
                        with col1:
                            selected = st.checkbox("", key=f"sel_{model['model_id']}_{idx}")
                        with col2:
                            st.write(f"**{model['model_name']}**")
                            st.caption(f"{model['category']} | {model['model_standard']}")
                        with col3:
                            price_info = self.search_model_price(model['model_name'])
                            if price_info is not None:
                                st.success(f"💰 {price_info['단가']:,}원/{price_info['단위']}")
                            else:
                                st.warning("단가 없음")
                        
                        if selected:
                            selected_models.append({
                                'model_name': model['model_name'],
                                'model_standard': model['model_standard']
                            })
                    
                    # 더보기 버튼
                    col1, col2 = st.columns([1, 1])
                    with col1:
                        if len(search_results) > current_count:
                            if st.button(f"더보기 ({len(search_results) - current_count}개 더)", key="show_more"):
                                st.session_state.display_count += 10
                                st.rerun()
                    
                    with col2:
                        if current_count > 10:
                            if st.button("처음으로", key="reset_display"):
                                st.session_state.display_count = 10
                                st.rerun()
                    
                    # 선택된 모델 처리
                    if selected_models:
                        st.write(f"✅ 선택된 모델: {len(selected_models)}개")
                        
                        if st.button("📋 선택된 모델 일괄 추가", key="add_bulk_items"):
                            for model in selected_models:
                                new_item = {
                                    'model_name': model['model_name'],
                                    'quantity': quantity,
                                    'notes': notes
                                }
                                st.session_state.material_items.append(new_item)
                            
                            st.success(f"✅ {len(selected_models)}개 모델이 추가되었습니다.")
                            # 표시 개수 초기화
                            st.session_state.display_count = 10
                            st.rerun()

                        
                else:
                    st.info("검색 결과가 없습니다. 다른 키워드로 시도해보세요.")
            else:
                st.info("검색어를 입력하여 모델을 찾아보세요.")
        
        # 추가된 항목들 표시
        if st.session_state.material_items:
            st.subheader("📋 자재 항목 목록")
            
            for i, item in enumerate(st.session_state.material_items):
                col1, col2, col3, col4 = st.columns([3, 1, 2, 1])
                
                with col1:
                    st.text(f"{i+1}. {item['model_name']}")
                with col2:
                    st.text(f"{item['quantity']:,}")
                with col3:
                    st.text(item['notes'])
                with col4:
                    if st.button("🗑️", key=f"delete_material_{i}"):
                        st.session_state.material_items.pop(i)
                        st.rerun()
            
            # 자재 및 실행내역서 생성
            if st.button("📊 자재 및 실행내역서 생성", type="primary", use_container_width=True):
                site_info = {
                    'site_name': site_name,
                    'foundation': foundation
                }
                
                quotation_data = self.generate_quotation(site_info, st.session_state.material_items, contract_type)
                
                if quotation_data['items']:
                    # 자재 및 실행내역서 생성
                    excel_buffer, material_items = self.create_material_execution_report(quotation_data, delivery_date)
                    
                    if excel_buffer:
                        st.success("✅ 자재 및 실행내역서 생성 완료!")
                        
                        # 결과 표시
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("자재 종류", f"{len(material_items)}개")
                        with col2:
                            total_cost = sum(_to_float_safe(item.get('unit_price', 0), 0) * _to_float_safe(item.get('quantity', 0), 0) for item in material_items)
                            st.metric("예상 자재비", f"{int(total_cost):,}원")
                        
                        # 자재 목록 표시
                        st.subheader("📦 자재 내역")
                        # 모델별로 그룹화
                        model_materials = {}
                        for item in material_items:
                            model_ref = item.get('content', '미분류')
                            if model_ref not in model_materials:
                                model_materials[model_ref] = []
                            model_materials[model_ref].append(item)

                        # 모델별로 표시
                        for model_ref, materials in model_materials.items():
                            with st.expander(f"{model_ref} ({len(materials)}개 자재)", expanded=True):
                                model_df = pd.DataFrame([
                                    {
                                        '자재명': item['material_name'],
                                        '규격': item['standard'],
                                        '수량': f"{item['quantity']:,.1f}",
                                        '단위': item['unit'],
                                        '단가': f"{int(item.get('unit_price', 0)):,}원"
                                    }
                                    for item in materials
                                ])
                                st.dataframe(model_df, use_container_width=True)
                                
                                # 모델별 소계
                                model_total = sum(_to_float_safe(item.get('unit_price', 0), 0) * _to_float_safe(item.get('quantity', 0), 0) for item in materials)
                                st.info(f"소계: {int(model_total):,}원")
                        
                        # 다운로드 버튼
                        filename = f"자재 및 실행내역서_{site_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                        
                        st.download_button(
                            label="📥 자재 및 실행내역서 다운로드",
                            data=excel_buffer.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )
                        
                        # 세션에 데이터 저장 (다음 단계용)
                        st.session_state.last_material_data = quotation_data
                    else:
                        st.error("자재 및 실행내역서 생성에 실패했습니다.")
        else:
            st.info("자재 항목을 추가해주세요.")

    def create_purchase_order_interface(self):
        """발주서 생성 인터페이스 - 카테고리별 업체 선택 방식"""
        st.header("📋 발주서 자동생성")
        
        if 'last_material_data' not in st.session_state:
            st.warning("먼저 자재발실행내역서를 생성해주세요. 자재 데이터를 기반으로 발주서가 생성됩니다.")
            return
        
        quotation_data = st.session_state.last_material_data
        
        st.info(f"현장: {quotation_data['site_info']['site_name']} | 자재 항목: {len(quotation_data['items'])}개")
        
        # 공통 설정
        col1, col2 = st.columns(2)
        with col1:
            delivery_location = st.text_input("하차지", value="공장")
        with col2:
            delivery_date = st.date_input("납품희망일", datetime.now() + pd.Timedelta(days=7))
        
        # 1단계: 발주 항목 미리보기
        st.subheader("🔍 1단계: 발주 항목 분석")
        
        if st.button("📦 발주 항목 분석하기", type="secondary", use_container_width=True):
            with st.spinner("발주 항목 분석 중..."):
                purchase_items = self.generate_purchase_items_from_quotation(quotation_data)
                
                if purchase_items:
                    # 세션에 발주 항목 저장
                    st.session_state.purchase_items = purchase_items
                    
                    # 카테고리별 그룹화
                    categories = {}
                    for item in purchase_items:
                        category = item['category']
                        if category not in categories:
                            categories[category] = []
                        categories[category].append(item)
                    
                    st.session_state.analyzed_categories = categories
                    st.success(f"✅ 총 {len(purchase_items)}개 자재, {len(categories)}개 카테고리로 분류 완료!")
                    
                    # 카테고리별 상세 정보 표시
                    for category, items in categories.items():
                        with st.expander(f"📂 {category} ({len(items)}개 항목)", expanded=True):
                            df_items = pd.DataFrame([
                                {
                                    '자재명': item['material_name'],
                                    '규격': item['standard'],
                                    '수량': f"{item['quantity']:,.1f}",
                                    '단위': item['unit'],
                                    '모델': item['model_reference']
                                }
                                for item in items
                            ])
                            st.dataframe(df_items, use_container_width=True)
                else:
                    st.warning("발주할 자재가 없습니다. BOM 데이터를 확인해주세요.")
        
        # 2단계: 카테고리별 업체 선택 및 발주서 생성
        if hasattr(st.session_state, 'analyzed_categories'):
            st.subheader("🏭 2단계: 카테고리별 공급업체 선택 및 발주")
            
            categories = st.session_state.analyzed_categories
            
            for category, items in categories.items():
                with st.container():
                    st.markdown(f"### 📂 **{category}** 카테고리")
                    st.caption(f"자재 {len(items)}개 항목")
                    
                    col1, col2, col3 = st.columns([2, 2, 2])
                    
                    with col1:
                        supplier_name = st.text_input(
                            "공급업체명",
                            key=f"supplier_{category}",
                            placeholder="공급업체를 입력하세요.",
                            help="해당 카테고리 자재를 공급받을 업체명을 입력하세요"
                        )
                    
                    with col2:
                        category_delivery_date = st.date_input(
                            "납품요청일",
                            delivery_date,
                            key=f"delivery_{category}"
                        )
                    
                    with col3:
                        st.write("")  # 간격 조정
                        st.write("")  # 간격 조정
                        
                        # 발주서 생성 버튼
                        if supplier_name.strip():
                            if st.button(
                                f"📋 {category} 발주서 생성",
                                key=f"create_order_{category}",
                                type="primary",
                                use_container_width=True
                            ):
                                self._create_category_purchase_order(
                                    category, items, supplier_name.strip(), 
                                    delivery_location, category_delivery_date, quotation_data
                                )
                        else:
                            st.button(
                                f"📋 {category} 발주서 생성",
                                key=f"create_order_{category}_disabled",
                                disabled=True,
                                use_container_width=True,
                                help="공급업체명을 입력하세요"
                            )
                    
                    st.markdown("---")

    def _create_category_purchase_order(self, category, items, supplier_name, 
                                    delivery_location, delivery_date, quotation_data):
        """카테고리별 발주서 생성"""
        try:
            with st.spinner(f"{category} → {supplier_name} 발주서 생성 중..."):
                # 기존 발주서 생성 함수 활용
                excel_buffer = self._create_single_purchase_order_by_category(
                    quotation_data, items, delivery_location, 
                    {'company_name': supplier_name}, delivery_date
                )
                
                if excel_buffer:
                    # 파일명 생성
                    filename = f"발주서_{supplier_name}_{category}_{quotation_data['site_info']['site_name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    
                    st.success(f"✅ {category} → {supplier_name} 발주서 생성 완료!")
                    
                    # 발주 요약 정보 표시
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("카테고리", category)
                    with col2:
                        st.metric("자재 종류", f"{len(items)}개")
                    with col3:
                        st.metric("공급업체", supplier_name)
                    
                    # 다운로드 버튼
                    st.download_button(
                        label=f"📥 {supplier_name} ({category}) 발주서 다운로드",
                        data=excel_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{category}_{supplier_name}_{datetime.now().strftime('%H%M%S')}",
                        type="primary",
                        use_container_width=True
                    )
                    
                    # 발주 내역 상세 표시
                    with st.expander(f"📋 {supplier_name} 발주 내역 상세", expanded=False):
                        df_order = pd.DataFrame([
                            {
                                '자재명': item['material_name'],
                                '규격': item['standard'],
                                '수량': f"{item['quantity']:,.1f}",
                                '단위': item['unit'],
                                '모델참조': item['model_reference']
                            }
                            for item in items
                        ])
                        st.dataframe(df_order, use_container_width=True)
                else:
                    st.error(f"{category} 발주서 생성에 실패했습니다.")
                    
        except Exception as e:
            st.error(f"발주서 생성 오류: {e}")

    def _create_single_purchase_order_by_category(self, quotation_data, purchase_items, 
                                                delivery_location, supplier_info, delivery_date):
        """카테고리별 단일 발주서 생성"""
        try:
            template_path = 'templates/발주서템플릿_v2.0_20250919.xlsx'
            if not os.path.exists(template_path):
                template_path = '발주서템플릿_v2.0_20250919.xlsx'
            
            workbook = load_workbook(template_path)
            sheet = workbook['발주서']
            
            # 기본 정보 입력
            today = datetime.now()
            sheet['F4'] = today.strftime('%Y년 %m월 %d일')
            sheet['B6'] = supplier_info['company_name']  # 수신 업체
            
            site_name = quotation_data['site_info']['site_name']
            start_row = 11
            
            data = self.load_data()
            
            for idx, purchase_item in enumerate(purchase_items):
                row = start_row + idx
                
                # 규격에 파이프 길이 정보 추가
                specification = self._get_specification_with_length_fixed(
                    purchase_item['material_name'], 
                    purchase_item['standard'], 
                    data
                )
                
                sheet[f'A{row}'] = idx + 1
                sheet[f'B{row}'] = purchase_item['material_name']
                sheet[f'C{row}'] = specification
                sheet[f'D{row}'] = purchase_item['unit']
                sheet[f'E{row}'] = purchase_item['quantity']
                sheet[f'F{row}'] = delivery_location
                sheet[f'G{row}'] = site_name
                sheet[f'H{row}'] = f"모델: {purchase_item['model_reference']}"
            
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            st.error(f"발주서 생성 오류: {e}")
            return None

    def create_quotation_interface(self):
        """견적서 생성 인터페이스"""
        st.header("💰 견적서 자동생성")
        
        if 'last_material_data' not in st.session_state:
            st.warning("먼저 자재 및 실행내역서를 생성해주세요. 해당 데이터를 기반으로 견적서가 생성됩니다.")
            return
        
        quotation_data = st.session_state.last_material_data
        
        st.info(f"현장: {quotation_data['site_info']['site_name']} | 견적 항목: {len(quotation_data['items'])}개")
        
        # 견적서 타입 선택
        col1, col2 = st.columns(2)
        with col1:
            contract_type = st.selectbox("계약 유형", ["관급", "사급"], key="quote_contract_type")
        with col2:
            quote_date = st.date_input("견적일자", datetime.now())
        
        if st.button("💰 견적서 생성", type="primary", use_container_width=True):
            # 계약 유형 업데이트
            quotation_data['contract_type'] = contract_type
            
            # 결과 표시
            st.success("✅ 견적서 생성 완료!")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("공급가", f"{quotation_data['total_supply_price']:,}원")
            with col2:
                st.metric("부가세(10%)", f"{quotation_data['vat_amount']:,}원")
            with col3:
                st.metric("총 금액", f"{quotation_data['total_amount']:,}원")
            
            # 상세 내역
            st.subheader("📄 견적 상세내역")
            detail_df = pd.DataFrame([
                {
                    '모델명': item['model_name'],
                    '규격': item['specification'],
                    '수량': f"{item['quantity']:,}{item['unit']}",
                    '단가': f"{item['unit_price']:,}원",
                    '금액': f"{item['supply_amount']:,}원"
                }
                for item in quotation_data['items']
            ])
            st.dataframe(detail_df, use_container_width=True)
            
            # 엑셀 다운로드 - 템플릿 기반
            excel_buffer = self.create_template_quotation(quotation_data)
            
            if excel_buffer:
                filename = f"두견적서_{quotation_data['site_info']['site_name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                
                st.download_button(
                    label="📥 템플릿 견적서 다운로드",
                    data=excel_buffer.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            else:
                st.error("견적서 생성에 실패했습니다.")


# 기존 검색 시스템 클래스들 유지
class EnhancedModelSearch:
    """고급 모델 검색 시스템"""
    
    def __init__(self, models_df):
        self.models_df = models_df
        self.search_columns = ['model_name', 'category', 'model_standard', '식별번호', 'description']
    
    def search_models(self, query, max_results=50):
        """통합 검색 함수"""
        if not query or not query.strip():
            return self.models_df.head(20)
        
        query = query.strip()
        search_results = []
        
        # 1. 숫자만 입력된 경우 - 식별번호 우선 검색
        if query.isdigit():
            results = self._search_by_identifier(query)
            search_results.extend(results)
        
        # 2. 치수 패턴 검색 (W2000, H1200 등)
        dimension_results = self._search_by_dimensions(query)
        search_results.extend(dimension_results)
        
        # 3. 각 컬럼별 부분 검색
        for column in self.search_columns:
            if column in self.models_df.columns:
                column_results = self._search_in_column(query, column)
                search_results.extend(column_results)
        
        # 4. 중복 제거 및 관련도 점수 계산
        unique_results = self._remove_duplicates_and_score(search_results, query)
        
        # 5. 관련도 순으로 정렬
        sorted_results = sorted(unique_results, key=lambda x: x['relevance_score'], reverse=True)
        
        # 6. DataFrame으로 변환
        if sorted_results:
            result_df = pd.DataFrame([item['model'] for item in sorted_results[:max_results]])
            return result_df
        else:
            return pd.DataFrame()
    
    def _search_by_identifier(self, query):
        """식별번호 검색"""
        results = []
        
        if '식별번호' in self.models_df.columns:
            mask = self.models_df['식별번호'].astype(str).str.contains(query, case=False, na=False)
            matched = self.models_df[mask]
            
            for _, row in matched.iterrows():
                results.append({
                    'model': row.to_dict(),
                    'match_type': 'identifier',
                    'match_column': '식별번호',
                    'match_value': str(row['식별번호'])
                })
        
        return results
    
    def _search_by_dimensions(self, query):
        """치수 기반 검색"""
        results = []
        
        patterns = [
            r'w(\d+)', r'width(\d+)', r'폭(\d+)',
            r'h(\d+)', r'height(\d+)', r'높이(\d+)',
            r'(\d+)w', r'(\d+)h'
        ]
        
        query_lower = query.lower()
        extracted_numbers = []
        
        for pattern in patterns:
            matches = re.findall(pattern, query_lower)
            extracted_numbers.extend(matches)
        
        if query.isdigit() and int(query) >= 1000:
            extracted_numbers.append(query)
        
        if extracted_numbers and 'model_standard' in self.models_df.columns:
            for number in extracted_numbers:
                mask = self.models_df['model_standard'].astype(str).str.contains(number, case=False, na=False)
                matched = self.models_df[mask]
                
                for _, row in matched.iterrows():
                    results.append({
                        'model': row.to_dict(),
                        'match_type': 'dimension',
                        'match_column': 'model_standard',
                        'match_value': row['model_standard']
                    })
        
        return results
    
    def _search_in_column(self, query, column):
        """특정 컬럼에서 부분 검색"""
        results = []
        
        try:
            mask = self.models_df[column].astype(str).str.contains(query, case=False, na=False)
            matched = self.models_df[mask]
            
            for _, row in matched.iterrows():
                results.append({
                    'model': row.to_dict(),
                    'match_type': 'partial',
                    'match_column': column,
                    'match_value': str(row[column])
                })
        except Exception:
            pass
        
        return results
    
    def _remove_duplicates_and_score(self, search_results, query):
        """중복 제거 및 관련도 점수 계산"""
        unique_models = {}
        
        for result in search_results:
            model = result['model']
            model_id = model.get('model_id', '') or model.get('model_name', '')
            
            if model_id not in unique_models:
                relevance_score = self._calculate_relevance_score(result, query)
                
                unique_models[model_id] = {
                    'model': model,
                    'relevance_score': relevance_score,
                    'match_info': [result]
                }
            else:
                unique_models[model_id]['match_info'].append(result)
                new_score = self._calculate_relevance_score(result, query)
                if new_score > unique_models[model_id]['relevance_score']:
                    unique_models[model_id]['relevance_score'] = new_score
        
        return list(unique_models.values())
    
    def _calculate_relevance_score(self, result, query):
        """관련도 점수 계산"""
        score = 0
        query_lower = query.lower()
        
        type_scores = {
            'identifier': 100,
            'dimension': 80,
            'partial': 50
        }
        
        score += type_scores.get(result['match_type'], 0)
        
        column_weights = {
            'model_name': 30,
            'category': 20,
            'model_standard': 25,
            '식별번호': 35,
            'description': 10
        }
        
        score += column_weights.get(result['match_column'], 0)
        
        match_value = str(result['match_value']).lower()
        similarity = SequenceMatcher(None, query_lower, match_value).ratio()
        score += similarity * 50
        
        if query_lower in match_value or match_value in query_lower:
            score += 20
        
        return score

# 검색 인터페이스 함수들
def create_enhanced_search_interface(models_df, quotation_system, bom_df):
    """고급 검색 인터페이스"""
    
    if 'dooho_search_system' not in st.session_state:
        st.session_state.dooho_search_system = EnhancedModelSearch(models_df)
    
    search_system = st.session_state.dooho_search_system
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        search_query = st.text_input(
            "통합 모델 검색",
            placeholder="모델명, 카테고리, 치수(W2000, H1200), 식별번호 등 입력",
            help="예: '디자인형', 'DAL', '2000', '24614649', 'W2000×H1200'",
            key="dooho_search"
        )
    
    with col2:
        search_button = st.button("🔍 검색", use_container_width=True, key="dooho_search_btn")
    
    if search_query or search_button:
        if search_query:
            with st.spinner("검색 중..."):
                search_results = search_system.search_models(search_query)
                
                if not search_results.empty:
                    st.success(f"검색 결과: {len(search_results)}개 모델 발견")
                    display_dooho_search_results(search_results, search_query, quotation_system, bom_df)
                else:
                    st.warning("검색 결과가 없습니다. 다른 키워드로 시도해보세요.")
                    show_dooho_search_tips()
        else:
            st.info("검색어를 입력해주세요.")
    else:
        st.subheader("전체 모델 목록 (처음 20개)")
        display_dooho_search_results(models_df.head(20), "", quotation_system, bom_df)

def display_dooho_search_results(results_df, search_query, quotation_system, bom_df):
    """검색 결과 표시"""
    
    for idx, (_, model) in enumerate(results_df.iterrows()):
        with st.expander(f"{model['model_name']} - {model['model_standard']}", expanded=False):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**모델 ID:** {model['model_id']}")
                st.write(f"**카테고리:** {model['category']}")
                st.write(f"**규격:** {model['model_standard']}")
            
            with col2:
                if pd.notna(model['식별번호']):
                    st.write(f"**식별번호:** {model['식별번호']}")
                st.write(f"**설명:** {model['description']}")
            
            price_info = quotation_system.search_model_price(model['model_name'])
            if price_info is not None:
                st.success(f"💰 단가: {price_info['단가']:,}원/{price_info['단위']}")
            else:
                st.warning("단가 정보 없음")
            
            model_bom = bom_df[bom_df['model_id'] == model['model_id']]
            if not model_bom.empty:
                st.write("**주요 자재:**")
                for _, bom_item in model_bom.head(3).iterrows():
                    st.write(f"- {bom_item['material_name']}: {bom_item['quantity']}{bom_item['unit']}")
            
            if search_query:
                highlight_dooho_matches(model, search_query)

def highlight_dooho_matches(model, search_query):
    """매칭된 부분 하이라이트"""
    
    matches = []
    query_lower = search_query.lower()
    
    search_fields = ['model_name', 'category', 'model_standard', '식별번호', 'description']
    
    for field_name in search_fields:
        if field_name in model and pd.notna(model[field_name]):
            field_value = str(model[field_name])
            if query_lower in field_value.lower():
                matches.append(f"{field_name}: {field_value}")
    
    if matches:
        st.info("🎯 매칭된 필드: " + ", ".join(matches[:2]))

def show_dooho_search_tips():
    """검색 팁 표시"""
    
    st.info("🔍 검색 가이드")
    st.markdown("""
    **검색 방법:**
    - **모델명**: `DAL`, `DHART`, `DHWS`, `DST` 등
    - **카테고리**: `디자인형` 입력시 디자인형울타리 전체 검색
    - **치수**: `2000`, `1200`, `W2000`, `H1500` 등
    - **식별번호**: `24614649`, `25320309` 등 8자리 숫자
    - **복합 검색**: `DAL 2000` (DAL 시리즈 중 2000 폭)
    """)

# 메인 애플리케이션
def main():
    st.title("🖥️ Project-Aegis 업무자동화 시스템")
    st.markdown("---")
    
    # 시스템 초기화
    if 'qs' not in st.session_state:
        st.session_state.qs = DoohoQuotationSystem()
    
    qs = st.session_state.qs
    data = qs.load_data()
    
    if not data:
        st.error("데이터베이스를 로딩할 수 없습니다.  파일을 확인해주세요.")
        return
    
    # 사이드바 - 데이터베이스 현황
    with st.sidebar:
        st.header("📊 데이터베이스 현황")
        st.metric("모델 수", len(data['models']))
        st.metric("단가 정보", len(data['pricing']))
        st.metric("BOM 항목", len(data['bom']))
        
        st.header("🏢 회사 정보")
        st.info('**회사명**\n금속구조물\n제작 설치 전문업체')
       
        # 모델 시리즈별 분포
        if len(data['models']) > 0:
            model_prefixes = {}
            for _, model in data['models'].iterrows():
                prefix = model['model_name'].split('-')[0][:4]
                model_prefixes[prefix] = model_prefixes.get(prefix, 0) + 1
            
            st.header("🗂️ 모델 시리즈")
            top_series = sorted(model_prefixes.items(), key=lambda x: x[1], reverse=True)[:5]
            for prefix, count in top_series:
                st.write(f"• {prefix}***: {count}개")
    
    # 메인 영역 - 작업 순서: 자재 및 실행내역서 → 발주서 → 견적서
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 자재 및 실행내역서", "📋 발주서 생성", "💰 견적서 생성", 
        "🔍 모델 조회", "📦 재고 현황", "📊 BOM 분석"
    ])
    
    with tab1:
        # 자재 및 실행내역서 생성 (1단계)
        qs.create_material_execution_interface()
    
    with tab2:
        # 발주서 생성 (2단계)
        qs.create_purchase_order_interface()
    
    with tab3:
        # 견적서 생성 (3단계)
        qs.create_quotation_interface()
    
    with tab4:
        st.header("🔍 모델 조회")
        
        # 고급 검색 시스템 적용
        create_enhanced_search_interface(data['models'], qs, data['bom'])

    with tab5:
        st.header("📦 재고 현황")
        
        # 재고 요약
        col1, col2, col3 = st.columns(3)
        with col1:
            total_items = len(data['inventory'])
            st.metric("총 자재 종류", f"{total_items}개")
        with col2:
            total_stock = data['inventory']['잔여재고'].sum()
            st.metric("총 잔여재고", f"{total_stock:,}EA")
        with col3:
            low_stock = len(data['inventory'][data['inventory']['잔여재고'] < 5])
            st.metric("재고 부족(5개 미만)", f"{low_stock}개", delta_color="inverse")
        
        # 전체 재고 현황 테이블
        st.subheader("📋 자재별 재고 현황")
        
        # 완전한 규격 생성 함수
        def create_full_specification(row):
            """규격, 두께, 길이를 조합하여 완전한 규격 생성"""
            spec = str(row['규격'])
            
            # 두께 정보 추가
            if pd.notna(row['두께']):
                spec += f"×{row['두께']}"
            
            # 파이프 길이 정보 추가
            if pd.notna(row['파이프길이(m)']):
                spec += f"×{row['파이프길이(m)']}m"
                
            return spec
        
        # 표시용 데이터프레임 생성
        inventory_display = data['inventory'].copy()
        inventory_display['완전규격'] = inventory_display.apply(create_full_specification, axis=1)
        
        # 필요한 컬럼만 선택
        display_columns = ['item_id', '재질', '완전규격', '잔여재고', '단위', '단가']
        
        # 컬럼이 존재하는지 확인 후 선택
        available_columns = [col for col in display_columns if col in inventory_display.columns]
        if '단가' not in inventory_display.columns:
            available_columns = [col for col in available_columns if col != '단가']
            
        final_display = inventory_display[available_columns].copy()
        
        # 컬럼명 한글화
        column_rename = {
            'item_id': '자재ID',
            '재질': '재질',
            '완전규격': '규격',
            '잔여재고': '잔여재고',
            '단위': '단위',
            '단가': '단가(원)'
        }
        
        final_display = final_display.rename(columns={k:v for k,v in column_rename.items() if k in final_display.columns})
        
        # 재고 부족 항목 강조
        def highlight_low_stock(val):
            if isinstance(val, (int, float)) and val < 5:
                return 'background-color: #ffcccc'
            return ''
        
        # 재고 순으로 정렬
        final_display = final_display.sort_values('잔여재고' if '잔여재고' in final_display.columns else final_display.columns[-2])
        
        # 스타일 적용하여 표시
        styled_df = final_display.style.applymap(
            highlight_low_stock, 
            subset=['잔여재고'] if '잔여재고' in final_display.columns else []
        )
        
        st.dataframe(styled_df, use_container_width=True)
        
        # 재고 부족 알림
        if low_stock > 0:
            st.warning(f"⚠️ {low_stock}개 자재의 재고가 5개 미만입니다. 발주를 검토해주세요.")
            
            # 재고 부족 자재 상세 표시
            low_stock_items = data['inventory'][data['inventory']['잔여재고'] < 5]
            if not low_stock_items.empty:
                with st.expander("재고 부족 자재 상세"):
                    for _, item in low_stock_items.iterrows():
                        full_spec = create_full_specification(item)
                        st.write(f"- {item['재질']} {full_spec}: {item['잔여재고']}개 남음")
    
    with tab6:
        st.header("📊 BOM 분석")
        
        # 자재 카테고리별 분포
        category_counts = data['bom']['category'].value_counts()
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📦 자재 카테고리별 분포")
            for category, count in category_counts.items():
                st.write(f"• {category}: {count}개")
        
        with col2:
            st.subheader("🔍 특정 모델 BOM 조회")
            selected_model_for_bom = st.selectbox(
                "모델 선택", 
                data['models']['model_name'].tolist(),
                key="bom_model"
            )
            
            if selected_model_for_bom:
                model_info = data['models'][data['models']['model_name'] == selected_model_for_bom].iloc[0]
                model_bom = data['bom'][data['bom']['model_id'] == model_info['model_id']]
                
                if not model_bom.empty:
                    st.write(f"**{selected_model_for_bom}** 자재 구성:")
                    for _, bom_item in model_bom.iterrows():
                        st.write(f"- {bom_item['material_name']} ({bom_item['standard']}): {bom_item['quantity']}{bom_item['unit']}")
                else:
                    st.info("해당 모델의 BOM 정보가 없습니다.")

if __name__ == "__main__":
    main()
                