import pandas as pd
import openpyxl
from openpyxl import load_workbook

def update_dooho_pricing_sheet():
    """새로운 조달단가표2.xlsx로 Pricing 시트 업데이트"""
    
    print("두호 Pricing 시트 업데이트 시작...")
    
    try:
        # 1. 새로운 조달단가표 읽기
        new_pricing_df = pd.read_excel('두호_조달단가표2.xlsx', sheet_name='Sheet1')
        print(f"새 조달단가표 로딩: {len(new_pricing_df)}개 항목")
        
        # 2. 데이터 정리 및 검증
        cleaned_pricing_data = []
        
        for idx, row in new_pricing_df.iterrows():
            품목 = str(row['품목']).strip() if pd.notna(row['품목']) else ''
            모델명 = str(row['모델명']).strip() if pd.notna(row['모델명']) else ''
            단위 = str(row['단위']).strip() if pd.notna(row['단위']) else 'm'
            단가 = row['단가'] if pd.notna(row['단가']) else 0
            식별번호 = row['식별번호'] if pd.notna(row['식별번호']) else ''
            
            # 유효한 데이터만 추가 (모델명과 단가가 있는 것)
            if 모델명 and 단가 > 0:
                cleaned_pricing_data.append({
                    'No': idx + 1,
                    '품목': 품목,
                    '모델명': 모델명,
                    '규격': '',  # 새 파일에는 규격 컬럼이 없음
                    '단위': 단위,
                    '단가': int(단가),
                    '식별번호': int(식별번호) if isinstance(식별번호, (int, float)) and 식별번호 > 0 else ''
                })
        
        print(f"데이터 정리 완료: {len(cleaned_pricing_data)}개 유효 항목")
        
        # 3. 기존 데이터베이스 파일 열기
        workbook = load_workbook('material_database_dooho.xlsx')
        
        # 4. 기존 Pricing 시트 삭제 후 재생성
        if 'Pricing' in workbook.sheetnames:
            del workbook['Pricing']
            print("기존 Pricing 시트 삭제")
        
        pricing_sheet = workbook.create_sheet('Pricing')
        
        # 5. 헤더 추가
        headers = ['No', '품목', '모델명', '규격', '단위', '단가', '식별번호']
        for col, header in enumerate(headers, 1):
            pricing_sheet.cell(row=1, column=col, value=header)
        
        # 6. 데이터 추가
        for row_idx, pricing_item in enumerate(cleaned_pricing_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = pricing_item.get(header, '')
                pricing_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 7. 파일 저장
        workbook.save('material_database_dooho.xlsx')
        
        print(f"Pricing 시트 업데이트 완료!")
        print(f"총 {len(cleaned_pricing_data)}개 모델 단가 정보 추가")
        
        # 8. Models 시트와 매칭 확인
        verify_model_pricing_match()
        
        return True
        
    except Exception as e:
        print(f"업데이트 오류: {e}")
        import traceback
        traceback.print_exc()
        return False

def verify_model_pricing_match():
    """Models와 Pricing 매칭 확인"""
    
    try:
        models_df = pd.read_excel('material_database_dooho.xlsx', sheet_name='Models ')
        pricing_df = pd.read_excel('material_database_dooho.xlsx', sheet_name='Pricing')
        
        print(f"\n=== 매칭 확인 ===")
        print(f"Models 시트: {len(models_df)}개")
        print(f"Pricing 시트: {len(pricing_df)}개")
        
        # 매칭 분석
        model_names = set(models_df['model_name'].str.strip())
        pricing_names = set(pricing_df['모델명'].str.strip())
        
        matched = model_names.intersection(pricing_names)
        models_only = model_names - pricing_names
        pricing_only = pricing_names - model_names
        
        print(f"\n매칭 결과:")
        print(f"  정확히 매칭: {len(matched)}개")
        print(f"  Models만 있음: {len(models_only)}개")
        print(f"  Pricing만 있음: {len(pricing_only)}개")
        print(f"  매칭률: {len(matched)/len(models_df)*100:.1f}%")
        
        # 매칭된 모델들 샘플 출력
        if matched:
            print(f"\n매칭된 모델 샘플 (처음 10개):")
            for i, model in enumerate(list(matched)[:10]):
                pricing_info = pricing_df[pricing_df['모델명'] == model].iloc[0]
                print(f"  {i+1}. {model}: {pricing_info['단가']:,}원/{pricing_info['단위']}")
        
        # 매칭되지 않은 Models 출력
        if models_only:
            print(f"\n단가 없는 모델들 (처음 10개):")
            for i, model in enumerate(list(models_only)[:10]):
                print(f"  {i+1}. {model}")
        
        # 매칭되지 않은 Pricing 출력  
        if pricing_only:
            print(f"\nModels에 없는 단가 정보 (처음 10개):")
            for i, model in enumerate(list(pricing_only)[:10]):
                print(f"  {i+1}. {model}")
        
        if len(matched) == len(models_df):
            print(f"\n🎉 완벽한 매칭! 모든 모델에 단가 정보가 있습니다!")
        elif len(matched) / len(models_df) >= 0.8:
            print(f"\n👍 양호한 매칭률! 대부분의 모델에 단가 정보가 있습니다.")
        else:
            print(f"\n⚠️ 매칭률이 낮습니다. 추가 작업이 필요할 수 있습니다.")
            
    except Exception as e:
        print(f"매칭 확인 오류: {e}")

def create_missing_models_from_pricing():
    """Pricing에만 있는 모델들을 Models 시트에 추가"""
    
    try:
        models_df = pd.read_excel('material_database_dooho.xlsx', sheet_name='Models ')
        pricing_df = pd.read_excel('material_database_dooho.xlsx', sheet_name='Pricing')
        
        model_names = set(models_df['model_name'].str.strip())
        pricing_names = set(pricing_df['모델명'].str.strip())
        
        pricing_only = pricing_names - model_names
        
        if not pricing_only:
            print("Pricing에만 있는 모델이 없습니다.")
            return True
        
        print(f"Pricing에만 있는 {len(pricing_only)}개 모델을 Models 시트에 추가합니다...")
        
        # 새로운 모델 데이터 생성
        new_models = []
        next_model_id = len(models_df) + 1
        
        for model_name in pricing_only:
            pricing_info = pricing_df[pricing_df['모델명'] == model_name].iloc[0]
            
            new_models.append({
                'model_id': f"DHP{str(next_model_id).zfill(3)}",
                'model_name': model_name,
                'category': '디자인형울타리',
                'model_standard': '',  # 규격 정보는 별도로 설정 필요
                '식별번호': pricing_info['식별번호'],
                'description': f"디자인형울타리 {model_name}"
            })
            next_model_id += 1
        
        # Models 시트에 추가
        extended_models_df = pd.concat([
            models_df,
            pd.DataFrame(new_models)
        ], ignore_index=True)
        
        # 데이터베이스 업데이트
        workbook = load_workbook('material_database_dooho.xlsx')
        
        if 'Models ' in workbook.sheetnames:
            del workbook['Models ']
        
        models_sheet = workbook.create_sheet('Models ')
        
        # 헤더 추가
        headers = ['model_id', 'model_name', 'category', 'model_standard', '식별번호', 'description']
        for col, header in enumerate(headers, 1):
            models_sheet.cell(row=1, column=col, value=header)
        
        # 데이터 추가
        for row_idx, (_, model_data) in enumerate(extended_models_df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                value = model_data.get(header, '')
                models_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        workbook.save('material_database_dooho.xlsx')
        
        print(f"Models 시트 확장 완료! {len(new_models)}개 모델 추가")
        return True
        
    except Exception as e:
        print(f"Models 확장 오류: {e}")
        return False

def complete_dooho_update():
    """두호 데이터베이스 완전 업데이트"""
    
    print("="*60)
    print("두호 데이터베이스 완전 업데이트")
    print("="*60)
    
    # 1단계: Pricing 시트 업데이트
    print("\n1단계: Pricing 시트 업데이트")
    if not update_dooho_pricing_sheet():
        print("Pricing 업데이트 실패")
        return False
    
    # 2단계: 매칭되지 않은 모델들 추가
    print("\n2단계: Models 시트 확장")
    if not create_missing_models_from_pricing():
        print("Models 확장 실패")
        return False
    
    # 3단계: 최종 확인
    print("\n3단계: 최종 매칭 확인")
    verify_model_pricing_match()
    
    print("\n" + "="*60)
    print("🎉 두호 데이터베이스 업데이트 완료!")
    print("이제 두호 견적서 시스템에서 모든 모델의 단가를 찾을 수 있습니다.")
    print("="*60)
    
    return True

if __name__ == "__main__":
    complete_dooho_update()
