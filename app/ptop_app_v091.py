# PTOP 통합 앱 v0.91 - 두호/국제 통합 버전
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from datetime import date, timedelta
import os
import io
import re
from difflib import SequenceMatcher
import shutil
import os
from pathlib import Path
import os

# Phase 3 auto-transfer helpers
import time
import random
try:
    from app.db_supabase_adapter import DatabaseManager  # type: ignore
except Exception:
    DatabaseManager = None  # type: ignore

try:
    import json as _json
except Exception:
    _json = None

@st.cache_resource(show_spinner=False)
def _get_phase3_db():
    if DatabaseManager is None:
        raise RuntimeError("Phase3 DatabaseManager not available")
    return DatabaseManager()

def _p3_jsonable(obj):
    try:
        from datetime import datetime, date as _date
        import numpy as _np  # type: ignore
    except Exception:
        class _NP:  # fallback
            integer = ()
            floating = ()
        _np = _NP()
        from datetime import datetime, date as _date
    if isinstance(obj, dict):
        return {k: _p3_jsonable(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_p3_jsonable(x) for x in obj]
    if _json is not None:
        # Normalize pandas/np/datetime to JSON-safe
        try:
            import pandas as _pd  # late import
            if isinstance(obj, _pd.Timestamp):
                return obj.to_pydatetime().isoformat()
        except Exception:
            pass
    if isinstance(obj, (datetime, _date)):
        return obj.isoformat()
    if hasattr(_np, 'integer') and isinstance(obj, _np.integer):
        return int(obj)
    if hasattr(_np, 'floating') and isinstance(obj, _np.floating):
        return float(obj)
    if hasattr(obj, 'item'):
        try:
            return obj.item()
        except Exception:
            return str(obj)
    return obj


# ============================================================================
# 발주서 생성 안전 장치 - KeyError 방지
# ============================================================================

def safe_get(obj, key, default=None):
    """
    안전한 딕셔너리 접근 (단일 키)

    예:
        safe_get(item, 'quantity', 0) → item['quantity'] 또는 0 반환
        safe_get(data, 'missing_key', 'N/A') → 'N/A' 반환
    """
    if not isinstance(obj, dict):
        return default
    return obj.get(key, default)


def safe_get_nested(obj, keys, default=None):
    """
    안전한 중첩 딕셔너리 접근

    예:
        safe_get_nested(quotation_data, ['site_info', 'site_name'], 'Unknown')
        → quotation_data['site_info']['site_name'] 또는 'Unknown' 반환
    """
    if not isinstance(obj, dict) or not isinstance(keys, list):
        return default

    current = obj
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current is None:
            return default

    return current


def validate_dict_keys(obj, required_keys):
    """
    딕셔너리의 필수 키 존재 여부 검증

    예:
        validate_dict_keys(item, ['quantity', 'material_name'])
        → 모든 키가 존재하고 None이 아니면 True, 아니면 False
    """
    if not isinstance(obj, dict):
        return False

    for key in required_keys:
        if key not in obj or obj[key] is None:
            return False

    return True


# ============================================================================

def _phase3_record_quotation(tenant_id: str, quotation_data: dict):
    try:
        db = _get_phase3_db()
        items = (quotation_data or {}).get('items') or []
        site = (quotation_data or {}).get('site_info') or {}
        project_id = site.get('project_id')
        try:
            total = float(quotation_data.get('total_amount')) if quotation_data.get('total_amount') is not None else None
        except Exception:
            total = None
        if total is None:
            try:
                total = float(sum(float((i.get('unit_price') or 0)) * float((i.get('quantity') or 0)) for i in items))
            except Exception:
                total = 0.0
        qid = f"Q-{int(time.time())}-{random.randint(100,999)}"
        db.add_quotation(qid, tenant_id, customer_id=None, project_id=project_id, total_amount=total)
        rows = []
        for it in items:
            name = (it.get('model_name') or it.get('material_name') or '').strip()
            spec = (it.get('specification') or it.get('standard') or '').strip()
            qty = float(it.get('quantity') or 0)
            price = float(it.get('unit_price') or 0)
            db.add_quotation_item(qid, name, spec=spec, quantity=qty, unit_price=price)
            rows.append({'품목': name, '규격': spec, '수량': qty, '단가': price, '금액': qty*price, '비고': (it.get('notes') or '')})
        # Save Excel-like snapshot for quotations
        try:
            payload = {
                'header': {
                    '프로젝트': (site.get('site_name') if isinstance(site, dict) else None),
                    '견적ID': qid,
                    '총액': total,
                },
                'items': rows,
            }
            db.add_bom_snapshot(tenant_id, 'quotation', qid, int(time.time()), _p3_jsonable(payload))
        except Exception:
            pass
    except Exception as e:
        try:
            st.session_state.setdefault('debug_messages', []).append(f"[Phase3] quotation record failed: {e}")
        except Exception:
            pass


def _phase3_record_po(tenant_id: str, category: str, supplier_name: str, items: list, project_id: str = None):
    try:
        db = _get_phase3_db()
        abbr = (category or '')[:3].upper() if category else 'GEN'
        po_id = f"PO-{int(time.time())}-{abbr}"
        db.add_purchase_order(po_id, tenant_id, vendor_id=(supplier_name or None), project_id=project_id, due_date=None, quotation_ref=None)
        rows = []
        for it in (items or []):
            name = (it.get('material_name') or '').strip()
            spec = (it.get('standard') or '').strip()
            item_name = (f"{name} {spec}").strip()
            qty = float(it.get('quantity') or 0)
            price = float(it.get('unit_price') or 0)
            db.add_po_item(po_id, item_name, material_id=None, quantity=qty, unit_price=price)
            rows.append({'품목': name, '규격': spec, '단위': (it.get('unit') or 'EA'), '수량': qty, '단가': price, '금액': qty*price, '비고': (it.get('notes') or ''), '모델참조': (it.get('model_reference') or '')})
        # Save Excel-like snapshot for POs
        try:
            payload = {
                'header': {
                    '발주ID': po_id,
                    '카테고리': category,
                    '공급업체': supplier_name,
                    '프로젝트ID': project_id,
                },
                'items': rows,
            }
            db.add_bom_snapshot(tenant_id, 'po', po_id, int(time.time()), _p3_jsonable(payload))
        except Exception:
            pass
    except Exception as e:
        try:
            st.session_state.setdefault('debug_messages', []).append(f"[Phase3] PO record failed: {e}")
        except Exception:
            pass


def _phase3_record_bom_and_execution(tenant_id: str, material_items: list, quotation_data: dict):
    try:
        db = _get_phase3_db()
        site = (quotation_data or {}).get('site_info') or {}
        linked_id = site.get('project_id') or site.get('site_name') or 'unknown'
        # Map to Excel-like columns
        rows = []
        for m in (material_items or []):
            if m.get('is_header'):
                continue
            qty = float(m.get('quantity') or 0)
            price = float(m.get('unit_price') or 0)
            rows.append({
                '품목': (m.get('material_name') or m.get('model_name') or ''),
                '규격': (m.get('standard') or m.get('specification') or ''),
                '수량': qty,
                '단위': (m.get('unit') or 'EA'),
                '단가': price,
                '금액': qty * price,
                '비고': (m.get('notes') or ''),
                '모델참조': (m.get('model_reference') or m.get('model_name') or ''),
            })
        payload_bom = {'items': rows}
        db.add_bom_snapshot(tenant_id, 'project', str(linked_id), int(time.time()), _p3_jsonable(payload_bom))
        # Also record execution snapshot with common header and empty items (editable in Demo)
        try:
            _items_q = (quotation_data or {}).get('items') or []
            _total = (quotation_data or {}).get('total_amount')
            if _total is None:
                try:
                    _total = float(sum(float(i.get('unit_price') or 0) * float(i.get('quantity') or 0) for i in _items_q))
                except Exception:
                    _total = 0.0
            payload_exec = {
                'header': { '계약금액(부가세포함)': _total },
                'items': [],
                'type': 'execution'
            }
            db.add_bom_snapshot(tenant_id, 'execution', str(linked_id), int(time.time()), _p3_jsonable(payload_exec))
        except Exception:
            pass
    except Exception as e:
        try:
            st.session_state.setdefault('debug_messages', []).append(f"[Phase3] BOM/Execution snapshot failed: {e}")
        except Exception:
            pass


# ============================================================================
# P0 생성 버튼 전환용 신규 헬퍼 함수들
# ============================================================================

def _create_quotation_and_buffer(app_instance, quotation_data, contract_type):
    """
    견적서를 생성하고 Excel 버퍼 + 세션 상태에 저장.
    반환: (success: bool, excel_buffer: BytesIO or None, error_msg: str or None)
    """
    try:
        quotation_data = dict(quotation_data)  # 사본
        quotation_data['contract_type'] = contract_type

        # Excel 생성
        excel_buffer = app_instance.create_template_quotation(quotation_data)
        if not excel_buffer:
            return False, None, "견적서 Excel 생성 실패"

        # 세션에 버퍼 저장
        st.session_state['quotation_buffer'] = excel_buffer
        st.session_state['quotation_data'] = quotation_data

        return True, excel_buffer, None

    except Exception as e:
        return False, None, f"견적서 생성 중 오류: {str(e)}"


def _save_quotation_to_db(tenant_id, quotation_data):
    """견적서를 데이터베이스에 저장"""
    try:
        _phase3_record_quotation(tenant_id, quotation_data)
        return True, None
    except Exception as e:
        return False, f"견적서 DB 저장 실패: {str(e)}"


def _create_po_and_buffer(app_instance, quotation_data, purchase_items, delivery_location, supplier_info, delivery_date, category):
    """
    발주서를 생성하고 Excel 버퍼 + 세션 상태에 저장.
    반환: (success: bool, excel_buffer: BytesIO or None, error_msg: str or None)
    """
    try:
        # Excel 생성
        excel_buffer = app_instance._create_single_purchase_order_by_category(
            quotation_data,
            purchase_items,
            delivery_location,
            supplier_info,
            delivery_date
        )
        if not excel_buffer:
            return False, None, "발주서 Excel 생성 실패"

        # 세션에 버퍼 저장 (다중 발주서용 dict)
        if 'po_buffers' not in st.session_state:
            st.session_state['po_buffers'] = {}
        po_key = f"{category}_{supplier_info.get('company_name', 'unknown')}"
        st.session_state['po_buffers'][po_key] = {
            'buffer': excel_buffer,
            'category': category,
            'supplier_name': supplier_info.get('company_name'),
            'items': purchase_items
        }

        return True, excel_buffer, None

    except Exception as e:
        return False, None, f"발주서 생성 중 오류: {str(e)}"


def _save_po_to_db(tenant_id, category, supplier_name, items, project_id=None):
    """발주서를 데이터베이스에 저장"""
    try:
        _phase3_record_po(tenant_id, category, supplier_name, items, project_id)
        return True, None
    except Exception as e:
        return False, f"발주서 DB 저장 실패: {str(e)}"


def _create_bom_and_execution_buffer(app_instance, quotation_data, material_items, delivery_date=None):
    """
    BOM과 실행내역서를 생성하고 Excel 버퍼 + 세션 상태에 저장.
    반환: (success: bool, excel_buffer: BytesIO or None, error_msg: str or None)
    """
    try:
        # Excel 생성
        excel_buffer, _ = app_instance.create_material_execution_report(quotation_data, delivery_date)
        if not excel_buffer:
            return False, None, "자재·실행내역서 Excel 생성 실패"

        # 세션에 버퍼 저장
        st.session_state['bom_execution_buffer'] = excel_buffer
        st.session_state['bom_execution_data'] = {
            'quotation_data': quotation_data,
            'material_items': material_items,
            'delivery_date': delivery_date
        }

        return True, excel_buffer, None

    except Exception as e:
        return False, None, f"자재·실행내역서 생성 중 오류: {str(e)}"


def _save_bom_and_execution_to_db(tenant_id, material_items, quotation_data):
    """BOM과 실행내역서를 데이터베이스에 저장"""
    try:
        _phase3_record_bom_and_execution(tenant_id, material_items, quotation_data)
        return True, None
    except Exception as e:
        return False, f"BOM·실행내역서 DB 저장 실패: {str(e)}"


# ============================================================================
# P0-2: 아이템 정규화 공통 함수
# ============================================================================

def normalize_item(item: dict, skip_header=True) -> dict:
    """
    다양한 출처의 아이템을 정규 컬럼으로 정규화.

    입력: {model_name, material_name, standard/spec/specification, unit, quantity, unit_price, ...}
    출력: {material_name, standard, unit, quantity, unit_price, model_reference, notes, ...}

    Args:
        item: 입력 아이템 dict
        skip_header: True면 is_header=True 아이템은 그대로 반환 (필터링 안 함)

    Returns:
        정규화된 dict
    """
    if not isinstance(item, dict):
        return {}

    # Header 아이템이면 그대로 반환
    if skip_header and item.get('is_header'):
        return item

    # 필수 필드 정규화
    normalized = {
        'material_name': (item.get('material_name') or item.get('품목') or '').strip(),
        'standard': (item.get('standard') or item.get('spec') or item.get('specification') or item.get('규격') or '').strip(),
        'unit': normalize_unit(item.get('unit') or item.get('단위') or 'EA'),
        'quantity': float(item.get('quantity') or item.get('수량') or 0),
        'unit_price': float(item.get('unit_price') or item.get('단가') or 0),
    }

    # 선택적 필드
    normalized['model_reference'] = (item.get('model_reference') or item.get('model_name') or item.get('모델참조') or '').strip()
    normalized['notes'] = (item.get('notes') or item.get('비고') or item.get('remarks') or '').strip()
    normalized['category'] = (item.get('category') or item.get('분류') or 'GENERAL').strip()

    # 추가 메타데이터 보존
    if item.get('is_header'):
        normalized['is_header'] = True
    if item.get('model_name'):
        normalized['model_name'] = item.get('model_name')
    if item.get('delivery_location'):
        normalized['delivery_location'] = item.get('delivery_location')
    if item.get('vehicle_number'):
        normalized['vehicle_number'] = item.get('vehicle_number')

    return normalized


def normalize_items_list(items: list, skip_header=True) -> list:
    """
    여러 아이템을 한 번에 정규화.

    Args:
        items: 아이템 list
        skip_header: True면 header 아이템도 보존

    Returns:
        정규화된 list
    """
    if not isinstance(items, list):
        return []
    return [normalize_item(item, skip_header) for item in items]


def get_item_required_fields(item: dict) -> dict:
    """
    아이템에서 필수 필드만 추출 (DB 저장용).

    Returns: {material_name, standard, unit, quantity, unit_price}
    """
    normalized = normalize_item(item)
    return {
        'material_name': normalized.get('material_name', ''),
        'standard': normalized.get('standard', ''),
        'unit': normalized.get('unit', 'EA'),
        'quantity': normalized.get('quantity', 0),
        'unit_price': normalized.get('unit_price', 0),
    }


# ============================================================================
# P0-5: 총 길이 → 경간 수 자동 계산 (모델별 폭 기반)
# ============================================================================

def calculate_span_count_from_total_length(total_length_m: float, model_standard: str, fallback_width_m: float = 2.0) -> int:
    """
    총 길이(m)와 모델 폭(model_standard에서 추출)으로 경간 수 자동 계산.

    Args:
        total_length_m: 총 길이 (미터, 예: 100)
        model_standard: 모델 규격 (예: "W2000", "2000", "폭2000mm")
        fallback_width_m: 폭 추출 실패 시 기본값 (기본: 2.0m)

    Returns:
        계산된 경간 수 (int)

    Examples:
        >>> calculate_span_count_from_total_length(100, "W2000")
        50  # 100m ÷ 2m = 50경간

        >>> calculate_span_count_from_total_length(100, "2500")
        40  # 100m ÷ 2.5m = 40경간
    """
    try:
        total_length = float(total_length_m)
        if total_length <= 0:
            return 1

        # parse_width_m_from_standard 함수 사용
        width_m = parse_width_m_from_standard(model_standard, fallback_width_m)
        if width_m <= 0:
            width_m = fallback_width_m

        # 경간 수 계산: 총 길이 ÷ 폭
        span_count = int(round(total_length / width_m))
        return max(1, span_count)  # 최소 1경간

    except Exception as e:
        # 계산 실패 시 1경간 반환
        print(f"[WARNING] 경간 계산 실패: {e}")
        return 1


# 테넌트 설정 - URL 파라미터 또는 쿼리 파라미터에서 가져오기
def get_tenant_from_params():
    """URL 파라미터에서 테넌트 ID 가져오기"""
    try:
        # Streamlit의 쿼리 파라미터에서 tenant_id 가져오기
        query_params = st.query_params
        if 'tenant' in query_params:
            tenant = query_params['tenant']
            if tenant in ['dooho', 'kukje', 'demo']:
                return tenant
    except:
        pass

    # 기본값 또는 세션 상태에서 가져오기
    return st.session_state.get('current_tenant', 'dooho')

def resolve_template_path(*candidate_names):
    """
    templates/ 하위 또는 루트에 있는 템플릿을 절대경로로 찾아 반환.
    우선순위: 환경변수 APP_ROOT → 스크립트 폴더 → 상위 폴더 → 현재작업폴더
    """
    roots = []
    env_root = os.getenv("APP_ROOT")
    if env_root:
        roots.append(Path(env_root))
    here = Path(__file__).resolve().parent
    roots += [here, here.parent, Path.cwd()]

    for root in roots:
        for name in candidate_names:
            for p in [root / 'templates' / name, root / name]:
                if p.exists():
                    return str(p.resolve())

    # 못 찾으면 에러를 명확히 던져서 화면에 절대경로 후보가 보이도록
    searched = []
    for root in roots:
        for name in candidate_names:
            searched.append(str((root / 'templates' / name).resolve()))
            searched.append(str((root / name).resolve()))
    raise FileNotFoundError("템플릿을 찾을 수 없습니다. 검색 경로:\n" + "\n".join(searched))

# 미리보기/편집 테이블 표준 헤더
WORKING_BOM_COLS = ["번호","품목","규격","단위","경간당수량","단가","금액","비고","모델참조"]

def init_working_bom(material_items):
    """
    material_items(list[dict]) -> 편집용 DataFrame을 만들어 세션 상태에 저장.
    """
    rows = []
    idx = 1
    for m in material_items:
        if m.get("is_header", False):
            rows.append({
                "번호": idx,
                "품목": f"[{m.get('model_name','')}]",
                "규격": "",
                "단위": "",
                "경간당수량": 0,
                "단가": 0,
                "금액": 0,
                "비고": "",
                "모델참조": m.get("model_name",""),
            })
            idx += 1
            continue

        qty = float(m.get("quantity", 0) or 0)
        price = float(m.get("unit_price", 0) or 0)
        rows.append({
            "번호": idx,
            "품목": m.get("material_name",""),
            "규격": m.get("standard",""),
            "단위": (m.get("unit","") or "").upper(),
            "경간당수량": qty,
            "단가": price,
            "금액": qty * price,
            "비고": m.get("notes",""),
            "모델참조": m.get("model_name",""),
        })
        idx += 1

    df = pd.DataFrame(rows, columns=WORKING_BOM_COLS)
    st.session_state["working_bom_df"] = df
    st.session_state["working_bom_ready"] = True
    return df

def open_bom_preview(material_items=None, *, quotation_data=None, data=None):
    """
    '자재내역서 미리보기' 진입용 헬퍼.
    """
    if material_items is None:
        material_items = []

    return init_working_bom(material_items)

# 앱 버전 정보
APP_VERSION = "0.91"

# 한 곳에서만 DB 파일명 관리
DB_FILE = "material_database.xlsx"

def get_db_path():
    return DB_FILE

UNIT_MAP = {"㎡":"M2","m²":"M2","M²":"M2","m2":"M2","ea":"EA","m":"M","kg":"KG"}
def normalize_unit(u):
    if u is None: return "EA"
    s = str(u).strip()
    return UNIT_MAP.get(s, s).upper()

# 경간(세트) 기준 전역 설정/헬퍼
DEFAULT_SPAN_WIDTH_M = 2.0

def parse_width_m_from_standard(std: str, fallback=DEFAULT_SPAN_WIDTH_M):
    """
    model_standard에서 폭 정보를 m로 변환해 추출.
    """
    if not std:
        return fallback
    s = str(std)
    m = re.search(r'[Ww\uFF37\ubc15]?\s*[-_×x]?\s*(\d{3,5})', s)
    if m:
        try:
            mm = float(m.group(1))
            if mm > 10:
                return round(mm/1000.0, 3)
            return mm
        except:
            pass
    m2 = re.search(r'(\d{3,5})', s)
    if m2:
        try:
            mm = float(m2.group(1))
            if mm > 10:
                return round(mm/1000.0, 3)
            return mm
        except:
            pass
    return fallback

PIPE_STANDARD_LENGTH_M = 6.0

def _safe_float(x, default=0.0):
    try:
        v = float(x)
        if pd.isna(v):
            return default
        return v
    except Exception:
        return default

class UnifiedQuotationSystem:
    """통합 업무자동화 시스템"""
    
    def __init__(self, tenant_id=None):
        # 테넌트 ID 설정
        self.tenant_id = tenant_id or get_tenant_from_params()
        
        # 테넌트별 설정
        self.tenant_config = {
            'dooho': {
                'name': '두호',
                'display_name': '두호'
            },
            'kukje': {
                'name': '국제',
                'display_name': '국제'
            },
            'demo': {
                'name': 'Aegis-Demo',
                'display_name': 'Aegis-Demo'
            }
        }
        
        # 세션 상태 확인 및 초기화
        current_tenant = st.session_state.get('current_tenant')
        if current_tenant != self.tenant_id:
            # 테넌트가 변경된 경우 관련 세션 상태 초기화
            keys_to_clear = ['material_items', 'model_span_plan', 'last_material_data', 
                            'quotation_items', 'selected_quote_model']
            for key in keys_to_clear:
                if key in st.session_state:
                    del st.session_state[key]
            st.session_state.current_tenant = self.tenant_id
            st.cache_data.clear()
        
        # Supabase + PtopEngine 초기화
        import sys
        from pathlib import Path

        project_root = Path(__file__).resolve().parent.parent
        if str(project_root) not in sys.path:
            sys.path.insert(0, str(project_root))

        from supabase import create_client
        import os
        from app.config_supabase import SUPABASE_URL, SUPABASE_KEY
        from utils.ptop_engine import PtopEngine

        try:
            # 환경변수에서 Supabase 설정 읽기 (demo 테넌트용 동적 설정)
            supabase_url = os.getenv('SUPABASE_URL') or SUPABASE_URL
            supabase_key = os.getenv('SUPABASE_KEY') or SUPABASE_KEY

            supabase = create_client(supabase_url, supabase_key)
            self.engine = PtopEngine(supabase, tenant_id=self.tenant_id)
            print(f"[INFO] PtopEngine 초기화 성공 (tenant: {self.tenant_id})")
        except Exception as e:
            st.error(f"Supabase 연결 실패: {e}")
            raise

        # 데이터 로드 (캐싱용)
        self.load_data()
    
    def _ukey(self, scope, *parts):
        import re
        norm = [re.sub(r'[^0-9A-Za-z]+', '_', str(p)) for p in parts if p is not None]
        return f"v091_{self.tenant_id}_{scope}_" + "_".join(norm)


    @st.cache_data
    def load_data(_self):
        """Supabase에서 데이터 로드 (PtopEngine 사용)"""
        import pandas as pd

        try:
            engine = _self.engine
            data = {}

            # Supabase에서 데이터 가져오기
            data['models'] = engine.get_all_models()
            data['pricing'] = engine.search_pricing('')
            data['main_materials'] = engine.search_main_materials('')
            data['sub_materials'] = engine.search_sub_materials('')
            data['inventory'] = engine.search_inventory('')

            # 컬럼명 변환 (Supabase → Excel 호환)
            
            # models 컬럼명 변환
            if not data['models'].empty and 'identifier_number' in data['models'].columns:
                data['models']['식별번호'] = data['models']['identifier_number']

            # pricing 컬럼명 변환
            if not data['pricing'].empty:
                pricing_map = {}
                if 'model_name' in data['pricing'].columns:
                    pricing_map['model_name'] = '모델명'
                if 'unit_price' in data['pricing'].columns:
                    pricing_map['unit_price'] = '단가'
                if 'unit' in data['pricing'].columns:
                    pricing_map['unit'] = '단위'
                if 'standard' in data['pricing'].columns:
                    pricing_map['standard'] = '규격'
                if pricing_map:
                    data['pricing'].rename(columns=pricing_map, inplace=True)

            # main_materials: NULL 값 처리 후 컬럼명 변환
            if not data['main_materials'].empty:
                if 'unit_length_m' in data['main_materials'].columns:
                    data['main_materials']['unit_length_m'] = data['main_materials']['unit_length_m'].fillna(6.0)

                main_materials_map = {}
                if 'product_name' in data['main_materials'].columns:
                    main_materials_map['product_name'] = '품목'
                if 'standard' in data['main_materials'].columns:
                    main_materials_map['standard'] = '규격'
                if 'unit_length_m' in data['main_materials'].columns:
                    main_materials_map['unit_length_m'] = '파이프길이(m)'
                if 'unit_price' in data['main_materials'].columns:
                    main_materials_map['unit_price'] = '단가'
                if main_materials_map:
                    data['main_materials'].rename(columns=main_materials_map, inplace=True)

            # sub_materials 컬럼명 변환
            if not data['sub_materials'].empty:
                sub_materials_map = {}
                if 'product_name' in data['sub_materials'].columns:
                    sub_materials_map['product_name'] = '품목'
                if 'standard' in data['sub_materials'].columns:
                    sub_materials_map['standard'] = '규격'
                if 'unit' in data['sub_materials'].columns:
                    sub_materials_map['unit'] = '단위'
                if 'unit_price' in data['sub_materials'].columns:
                    sub_materials_map['unit_price'] = '단가'
                if 'notes' in data['sub_materials'].columns:
                    sub_materials_map['notes'] = '비고'
                if 'supplier' in data['sub_materials'].columns:
                    sub_materials_map['supplier'] = '업체명'
                if sub_materials_map:
                    data['sub_materials'].rename(columns=sub_materials_map, inplace=True)

            # inventory 컬럼명 변환
            if not data['inventory'].empty:
                inventory_map = {}
                if 'item_id' in data['inventory'].columns:
                    inventory_map['item_id'] = '자재ID'
                if 'product_name' in data['inventory'].columns:
                    inventory_map['product_name'] = '재질'
                if 'standard' in data['inventory'].columns:
                    inventory_map['standard'] = '규격'
                if 'thickness' in data['inventory'].columns:
                    inventory_map['thickness'] = '두께'
                if 'unit_length_m' in data['inventory'].columns:
                    inventory_map['unit_length_m'] = '파이프길이(m)'
                if 'unit_price' in data['inventory'].columns:
                    inventory_map['unit_price'] = '단가'
                if 'current_quantity' in data['inventory'].columns:
                    inventory_map['current_quantity'] = '잔여재고'
                if 'unit' in data['inventory'].columns:
                    inventory_map['unit'] = '단위'
                if 'supplier' in data['inventory'].columns:
                    inventory_map['supplier'] = '공급업체'
                if 'notes' in data['inventory'].columns:
                    inventory_map['notes'] = '비고'
                if inventory_map:
                    data['inventory'].rename(columns=inventory_map, inplace=True)

            # BOM은 특정 모델에 대해서만 조회하므로 빈 DF로 초기화
            data['bom'] = pd.DataFrame()
            data['bom1'] = pd.DataFrame(columns=['model_id','material_name','standard','unit','quantity','category','notes'])

            return data

        except Exception as e:
            st.error(f"데이터 로드 실패: {e}")
            import traceback
            st.error(f"상세 오류:\n{traceback.format_exc()}")
            return {
                'models': pd.DataFrame(),
                'bom': pd.DataFrame(),
                'pricing': pd.DataFrame(),
                'inventory': pd.DataFrame(),
                'main_materials': pd.DataFrame(),
                'sub_materials': pd.DataFrame(),
                'bom1': pd.DataFrame()
            }

    def save_to_bom1_sheet(self, material_data):
        """BOM에 수동 자재 저장 (Supabase)"""
        try:
            if not material_data.get('model_id'):
                st.error("BOM 저장 실패: model_id가 없습니다. 먼저 모델을 선택하세요.")
                return False

            success = self.engine.add_bom_item(
                model_id=material_data['model_id'],
                material_data=material_data
            )

            if success:
                st.cache_data.clear()
                st.success(f"BOM에 '{material_data['material_name']}' 추가 완료")
                return True
            else:
                st.error("BOM 저장 실패")
                return False

        except Exception as e:
            st.error(f"BOM 저장 오류: {e}")
            return False

    def load_bom1_data(self):
        """BOM 데이터 로드 (Supabase - 더 이상 BOM1 시트 사용 안함)"""
        return pd.DataFrame(columns=['model_id', 'material_name', 'standard', 'quantity', 'unit', 'category', 'notes'])
    
    def search_model_price(self, model_name):
        """모델 단가 검색"""
        data = self.load_data()
        if not data:
            return None

        pricing_df = data.get('pricing')
        if pricing_df is None or len(pricing_df) == 0:
            if st.session_state.get("_DBG", False):
                st.warning("[DEBUG] pricing_df is empty or missing")
            return None

        if '모델명' not in pricing_df.columns:
            if st.session_state.get("_DBG", False):
                st.warning(f"[DEBUG] pricing_df columns: {list(pricing_df.columns)} — '모델명' 컬럼 없음")
            return None

        model_clean = str(model_name).strip()
        col = pricing_df['모델명'].astype(str).str.strip()
        exact_match = pricing_df[col == model_clean]

        if not exact_match.empty:
            return exact_match.iloc[0]

        if st.session_state.get("_DBG", False):
            st.warning(f"[DEBUG] price_miss(find_model_price): model={model_clean} | available_cols={list(pricing_df.columns)} | rows={len(pricing_df)}")

        return None

    def generate_quotation(self, site_info, items, contract_type="관급"):
        """견적서 생성"""
        quotation_items = []
        total_supply_price = 0
        
        for item in items:
            # 수동 입력 자재 처리
            if item.get('source') == 'MANUAL':
                unit_price = float(item.get('unit_price', 0))
                supply_amount = item['quantity'] * unit_price

                quotation_items.append({
                    'model_name': item['model_name'],
                    'specification': item.get('specification', item.get('standard', '')),
                    'unit': item.get('unit', 'EA'),
                    'quantity': item['quantity'],
                    'unit_price': unit_price,
                    'supply_amount': supply_amount,
                    'notes': item.get('notes', ''),
                    '식별번호': '',
                    'source': 'MANUAL',
                    'material_name': item['material_name']
                })

                total_supply_price += supply_amount
                continue

            # 일반 모델 처리
            price_info = self.search_model_price(item['model_name'])
            
            if price_info is None:
                st.warning(f"'{item['model_name']}' 모델의 단가를 찾을 수 없습니다.")
                continue
            
            unit_price = float(price_info['단가'])
            supply_amount = item['quantity'] * unit_price
            
            quotation_items.append({
                'model_name': item['model_name'],
                'specification': price_info['규격'],
                'unit': price_info['단위'],
                'quantity': item['quantity'],
                'unit_price': unit_price,
                'supply_amount': supply_amount,
                'notes': item.get('notes', ''),
                '식별번호': price_info.get('식별번호', '')
            })
            
            total_supply_price += supply_amount

        total_amount = total_supply_price

        return {
            'site_info': site_info,
            'contract_type': contract_type,
            'items': quotation_items,
            'total_supply_price': total_supply_price,
            'vat_amount': 0,
            'total_amount': total_amount,
            'created_date': datetime.now(),
            'company': self.tenant_config[self.tenant_id]['display_name']
        }
    
    def generate_purchase_items_from_quotation(self, quotation_data):
        """견적서 데이터를 기반으로 발주 항목 생성 (카테고리 기반)"""
        data = self.load_data()
        purchase_items = []

        # ===== 안전한 데이터 접근 시작 =====
        # quotation_data 기본 검증
        if not isinstance(quotation_data, dict):
            return purchase_items

        plan = {}
        try:
            plan = quotation_data.get('site_info', {}).get('model_span_plan', {}) or {}
        except Exception:
            plan = {}

        models_df = data.get('models', pd.DataFrame())
        model_cat_map = {}
        if not models_df.empty:
            name_col = 'model_name' if 'model_name' in models_df.columns else None
            cat_col = 'category' if 'category' in models_df.columns else None
            if name_col and cat_col:
                for _, r in models_df.iterrows():
                    model_cat_map[str(r[name_col])] = str(r[cat_col])

        # 안전한 중첩 접근: site_info.total_span_count
        total_span_count = int(safe_get_nested(quotation_data, ['site_info', 'total_span_count'], 1))

        # 안전한 items 접근
        items = safe_get(quotation_data, 'items', [])
        if not items:
            return purchase_items

        for item in items:
            # 필수 필드 검증: model_name, quantity
            if not validate_dict_keys(item, ['model_name', 'quantity']):
                continue  # 불완전한 항목은 건너뜀

            try:
                model_name = item['model_name']
                item_quantity = float(item['quantity'])

                model_info = data['models'][data['models']['model_name'] == model_name]

                if not model_info.empty:
                    model_id = model_info.iloc[0]['model_id']
                    model_bom = self.engine.get_bom(model_id)

                    # BOM 데이터 유효성 검증 (Empty DataFrame 체크)
                    if model_bom is None or model_bom.empty:
                        print(f"[WARNING] BOM not found for model: {model_name} (model_id: {model_id})")
                        continue  # 다음 item으로 진행

                    for _, bom_item in model_bom.iterrows():
                        multiplier = total_span_count
                        if model_name in plan:
                            multiplier = int(plan[model_name].get('span_count', multiplier))
                        model_cat = model_cat_map.get(model_name, '')
                        if '차양' in str(model_cat):
                            multiplier = 1

                        per_span_qty = float(bom_item['quantity'])
                        required_quantity = item_quantity * per_span_qty * multiplier

                        if 'PIPE' in str(bom_item['category']).upper():
                            required_quantity = self._calculate_pipe_count(
                                required_quantity,
                                bom_item['standard'],
                                data
                            )
                            unit = 'EA'
                        else:
                            unit = bom_item['unit']

                        existing_item = None
                        for purchase_item in purchase_items:
                            if (purchase_item['material_name'] == bom_item['material_name'] and
                                purchase_item['standard'] == bom_item['standard']):
                                existing_item = purchase_item
                                break

                        if existing_item:
                            existing_item['quantity'] += required_quantity
                        else:
                            purchase_items.append({
                                'material_name': bom_item['material_name'],
                                'standard': bom_item['standard'],
                                'unit': unit,
                                'quantity': required_quantity,
                                'category': bom_item['category'],
                                'model_reference': model_name
                            })
            except Exception as e:
                # 개별 항목 처리 실패는 로그하고 계속
                import traceback
                print(f"[ERROR] Item processing failed for model: {item.get('model_name', 'N/A')}")
                print(f"[ERROR] Exception: {e}")
                print(f"[ERROR] Traceback: {traceback.format_exc()}")
                continue

        return purchase_items

    def create_material_execution_report(self, quotation_data, delivery_date=None):
        """자재발실행내역서 자동생성"""
        try:
            template_paths = [
                '../templates/자재 및 실행내역서템플릿_v2.0_20250919.xlsx',
                'templates/자재 및 실행내역서템플릿_v2.0_20250919.xlsx',
                '자재 및 실행내역서템플릿_v2.0_20250919.xlsx',
                '../자재 및 실행내역서템플릿_v2.0_20250919.xlsx'
            ]
            
            template_path = None
            for path in template_paths:
                if os.path.exists(path):
                    template_path = path
                    break
            
            if template_path is None:
                st.error("템플릿 파일을 찾을 수 없습니다. 다음 위치 중 한 곳에 '자재 및 실행내역서템플릿_v2.0_20250919.xlsx' 파일이 있는지 확인해주세요:")
                for path in template_paths:
                    st.write(f"• {os.path.abspath(path)}")
                return None, []
            
            workbook = load_workbook(template_path)
            material_sheet = workbook['자재내역서']
            
            site_name = quotation_data['site_info']['site_name']
            material_sheet['B3'] = site_name

            plan = (quotation_data.get('site_info', {}) or {}).get('model_span_plan', {}) or {}
            total_model_length_m = sum(float(v.get('total_length_m', 0) or 0) for v in plan.values())
            material_sheet['F3'] = round(total_model_length_m, 2)
            
            if delivery_date:
                material_sheet['B5'] = delivery_date.strftime('%Y년 %m월 %d일')
            else:
                material_sheet['B5'] = (datetime.now() + pd.Timedelta(days=7)).strftime('%Y년 %m월 %d일')
            
            data = self.load_data()
            material_items = self._generate_material_items_with_pricing(quotation_data, data)
            
            start_row = 9
            
            for idx, material in enumerate(material_items):
                row = start_row + idx
                
                if material.get('is_header', False):
                    material_sheet[f'A{row}'] = idx + 1
                    material_sheet[f'B{row}'] = material['model_name']
                    material_sheet[f'C{row}'] = ''
                    material_sheet[f'D{row}'] = ''
                    material_sheet[f'E{row}'] = ''
                    material_sheet[f'F{row}'] = ''
                    material_sheet[f'G{row}'] = ''
                    material_sheet[f'H{row}'] = ''
                    material_sheet[f'I{row}'] = ''
                    material_sheet[f'J{row}'] = ''
                    material_sheet[f'K{row}'] = ''
                else:
                    material_sheet[f'A{row}'] = idx + 1
                    material_sheet[f'B{row}'] = material['material_name']

                    standard_display = material['standard']
                    if '×' in standard_display or '×' in standard_display:
                        standard_display = standard_display.split('×')[0].split('×')[0]

                    material_sheet[f'C{row}'] = standard_display
                    material_sheet[f'D{row}'] = material['unit']
                    material_sheet[f'E{row}'] = material['quantity']
                    unit_price = material.get('unit_price', 0)
                    material_sheet[f'F{row}'] = unit_price
                    material_sheet[f'G{row}'] = material['quantity'] * unit_price
                    material_sheet[f'H{row}'] = material.get('notes', '')
                    material_sheet[f'I{row}'] = '공장'
                    material_sheet[f'J{row}'] = datetime.now().strftime('%Y-%m-%d')
                    material_sheet[f'K{row}'] = '공급업체명'
            
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer, material_items
            
        except Exception as e:
            st.error(f"자재 및 실행내역서 생성 오류: {e}")
            return None, []

    def _generate_material_items_with_pricing(self, quotation_data, data):
        """BOM 데이터에 단가 정보를 결합한 자재 목록 생성"""
        import math

        plan = {}
        try:
            plan = quotation_data.get('site_info', {}).get('model_span_plan', {}) or {}
        except Exception:
            plan = {}

        import pandas as pd
        models_df = data.get('models', pd.DataFrame())
        model_cat_map = {}
        if not models_df.empty:
            name_col = 'model_name' if 'model_name' in models_df.columns else None
            cat_col = 'category' if 'category' in models_df.columns else None
            if name_col and cat_col:
                for _, r in models_df.iterrows():
                    model_cat_map[str(r[name_col])] = str(r[cat_col])

        span_width_m = _safe_float(quotation_data['site_info'].get('span_width_m'), DEFAULT_SPAN_WIDTH_M)
        total_span_count = int(quotation_data['site_info'].get('total_span_count', 1))

        material_items_by_model = {}

        for item in quotation_data['items']:
            model_name = item.get('model_name', '')

            if item.get('source') == 'MANUAL':
                continue

            model_info = data['models'][data['models']['model_name'] == model_name]
            if not model_info.empty:
                model_id = model_info.iloc[0]['model_id']
                model_bom = self.engine.get_bom(model_id)

                if model_name not in material_items_by_model:
                    material_items_by_model[model_name] = []

                for _, bom_item in model_bom.iterrows():
                    category = str(bom_item['category'])
                    material_name = bom_item['material_name']
                    bom_standard = bom_item['standard']

                    per_span_qty = _safe_float(bom_item['quantity'], 0.0)
                    unit = bom_item['unit']

                    if category == 'MANUAL':
                        unit_price = _safe_float(bom_item.get('unit_price', 0.0))
                        actual_standard = bom_standard
                    else:
                        material_info = self._find_material_info_by_category(
                            category, bom_standard, data, material_name
                        )

                        unit_price = _safe_float(material_info.get('단가', 0.0))
                        actual_standard = material_info.get('완전규격', material_info.get('규격', bom_standard))

                    if '×' in actual_standard or '×' in actual_standard:
                        actual_standard = actual_standard.split('×')[0].split('×')[0]

                    enhanced_standard = actual_standard

                    material_items_by_model[model_name].append({
                        'material_name': material_name,
                        'standard': enhanced_standard,
                        'unit': unit,
                        'quantity': per_span_qty,
                        'category': category,
                        'unit_price': unit_price,
                        'model_name': model_name,
                        'notes': ''
                    })

        final_material_items = []
        for model_name, model_materials in material_items_by_model.items():
            final_material_items.append({
                'material_name': f"=== 모델: {model_name} ===",
                'standard': '',
                'unit': '',
                'quantity': 0,
                'category': 'MODEL_HEADER',
                'unit_price': 0,
                'model_name': model_name,
                'notes': '',
                'is_header': True
            })

            for m in model_materials:
                per_span_qty = _safe_float(m.get('quantity', 0.0), 0.0)
                category_upper = str(m.get('category', '')).upper()
                unit = m.get('unit', 'EA')
                out_unit = unit
                notes = str(m.get('notes', ''))
                unit_price_safe = _safe_float(m.get('unit_price'), 0.0)

                multiplier = int(quotation_data['site_info'].get('total_span_count', 1))
                if model_name in plan:
                    multiplier = int(plan[model_name].get('span_count', multiplier))
                model_cat = model_cat_map.get(model_name, '')
                if '차양' in str(model_cat):
                    multiplier = 1

                total_qty = per_span_qty * multiplier

                if 'PIPE' in category_upper:
                    import math
                    total_length_m = total_qty
                    total_pipes = math.ceil(total_length_m / PIPE_STANDARD_LENGTH_M)
                    out_unit = 'M'
                    pipe_note = f"파이프 소모량: {PIPE_STANDARD_LENGTH_M:.0f}m×{total_pipes}본"
                    notes = f"{notes} | {pipe_note}".strip(" |")

                final_material_items.append({
                    'material_name': m['material_name'],
                    'standard': m['standard'],
                    'unit': out_unit,
                    'quantity': total_qty,
                    'category': m.get('category', ''),
                    'unit_price': unit_price_safe,
                    'model_name': model_name,
                    'notes': notes
                })

        return final_material_items

    def _find_material_info_by_category(self, category, standard, data, material_name=None):
        """카테고리로 자재 정보 찾기"""
        
        if 'main_materials' in data:
            main_materials = data['main_materials']
            try:
                if not isinstance(main_materials, pd.DataFrame) or main_materials.empty or '품목' not in main_materials.columns:
                    raise ValueError("main_materials is not a valid DataFrame or missing '품목' column")

                category_match = main_materials[
                    main_materials['품목'].astype(str).str.strip() == str(category).strip()
                ]

                if not category_match.empty:
                    bom_standard = str(standard).strip()
                    
                    for _, material_row in category_match.iterrows():
                        main_spec = str(material_row['규격']).strip() if pd.notna(material_row['규격']) else ''
                        
                        if self._compare_specs_order_agnostic(bom_standard, main_spec):
                            return self._create_material_result_from_main(material_row, category)
                    
                    st.session_state.debug_messages.append(f"🟡 [자재 매칭 주의] 카테고리 '{category}'는 찾았지만, 규격 '{standard}'와 일치하는 항목이 main_materials에 없습니다. 부자재에서 검색합니다.")
                
            except Exception as e:
                st.session_state.debug_messages.append(f"main_materials 검색 오류: {e}")
        
        if 'sub_materials' in data:
            sub_materials = data['sub_materials']
            try:
                if not isinstance(sub_materials, pd.DataFrame) or sub_materials.empty:
                    raise ValueError("sub_materials is not a valid DataFrame or is empty")

                if material_name and '품목' in sub_materials.columns:
                    material_name_match = sub_materials[
                        sub_materials['품목'].astype(str).str.contains(str(material_name), na=False, case=False)
                    ]
                    if not material_name_match.empty:
                        material_row = material_name_match.iloc[0]
                        return self._create_material_result_from_sub(material_row)

                if '규격' in sub_materials.columns:
                    # 표준 정규화 (x → *)
                    normalized_search = self._normalize_special_chars(str(standard))
                    # DB 값도 정규화하여 비교
                    standard_match = sub_materials[
                        sub_materials['규격'].astype(str).apply(
                            lambda x: normalized_search in self._normalize_special_chars(x)
                        )
                    ]
                    if not standard_match.empty:
                        material_row = standard_match.iloc[0]
                        return self._create_material_result_from_sub(material_row)

            except Exception as e:
                st.session_state.debug_messages.append(f"sub_materials 검색 오류: {e}")
        
        st.session_state.debug_messages.append(f"❌ [자재 찾기 실패] 카테고리: '{category}' / 규격: '{standard}' / 자재명: '{material_name}'을 main_materials와 sub_materials에서 찾을 수 없습니다.")
        return self._create_empty_result()
    
    def _compare_specs_order_agnostic(self, bom_spec, main_spec):
        """순서 무관 규격 비교"""
        if not bom_spec or not main_spec:
            return False
        
        if self._compare_complete_specs(bom_spec, main_spec):
            return True
        
        return self._compare_with_reversed_dimensions(bom_spec, main_spec)

    def _compare_with_reversed_dimensions(self, bom_spec, main_spec):
        """치수 순서를 바꿔서 비교"""
        import re

        # x 또는 * 모두 매칭하도록 정규표현식 수정
        bom_match = re.match(r'(\d+)[x*](\d+)[x*](.+)', bom_spec, re.IGNORECASE)
        main_match = re.match(r'(\d+)[x*](\d+)[x*](.+)', main_spec, re.IGNORECASE)
        
        if bom_match and main_match:
            bom_dim1, bom_dim2, bom_thickness = bom_match.groups()
            main_dim1, main_dim2, main_thickness = main_match.groups()
            
            if bom_thickness.strip() == main_thickness.strip():
                if ((bom_dim1 == main_dim2 and bom_dim2 == main_dim1) or
                    (bom_dim1 == main_dim1 and bom_dim2 == main_dim2)):
                    return True
        
        return False

    def _compare_complete_specs(self, bom_spec, main_spec):
        """완전한 규격 비교"""
        if not bom_spec or not main_spec:
            return False
        
        bom_clean = str(bom_spec).strip()
        main_clean = str(main_spec).strip()
        
        if bom_clean == main_clean:
            return True
        
        bom_normalized = self._normalize_special_chars(bom_clean)
        main_normalized = self._normalize_special_chars(main_clean)
        
        return bom_normalized == main_normalized

    def _normalize_special_chars(self, spec):
        """특수문자 정규화"""
        # x, X 를 * 로 정규화 (DB 표준 형식)
        normalized = spec.replace('x', '*').replace('X', '*')
        # 지름 기호 정규화
        normalized = normalized.replace('∅', 'Ø').replace('Φ', 'Ø').replace('φ', 'Ø')
        normalized = normalized.upper()
        return normalized
    
    def _create_empty_result(self):
        """빈칸 결과 생성"""
        return {
            '완전규격': '',
            '단가': '',
            '품목': '',
            '규격': ''
        }

    def _create_material_result_from_main(self, material_row, category):
        """main_Materials 결과 생성"""
        main_spec = str(material_row['규격']).strip()
        pipe_length = material_row.get('파이프길이(m)', 6.0)
        unit_price = float(material_row['단가']) if pd.notna(material_row['단가']) else 0
        
        if any(pipe_word in category.upper() for pipe_word in ['PIPE', '파이프']):
            unit_price = unit_price / pipe_length if pipe_length > 0 else unit_price     

        if any(pipe_word in category.upper() for pipe_word in ['PIPE', '파이프']):
            full_specification = f"{main_spec}×{pipe_length}m"
        else:
            full_specification = main_spec
        
        return {
            '완전규격': full_specification,
            '단가': unit_price,
            '품목': material_row['품목'],
            '규격': material_row['규격']
        }

    def _create_material_result_from_sub(self, material_row):
        """sub_Materials 결과 생성"""
        unit_price = float(material_row['단가']) if pd.notna(material_row['단가']) else 0
        spec = str(material_row['규격']).strip()
        
        return {
            '완전규격': spec,
            '단가': unit_price,
            '품목': material_row['품목'],
            '규격': material_row['규격']
        }
    
    def _calculate_pipe_count(self, required_length_m, pipe_standard, data):
        """파이프 길이를 고려한 실제 발주 개수 계산"""
        import math

        main_materials = data.get('main_materials', pd.DataFrame())

        if isinstance(main_materials, pd.DataFrame) and not main_materials.empty and '규격' in main_materials.columns:
            pipe_match = main_materials[
                main_materials['규격'].astype(str).str.contains(pipe_standard, na=False, case=False)
            ]
        else:
            pipe_match = pd.DataFrame()
        
        if not pipe_match.empty:
            try:
                standard_length = 6.0
                if '길이' in pipe_match.columns:
                    standard_length = float(pipe_match.iloc[0]['길이'])
                elif '단위길이' in pipe_match.columns:
                    standard_length = float(pipe_match.iloc[0]['단위길이'])
                elif '파이프길이(m)' in pipe_match.columns:
                    standard_length = float(pipe_match.iloc[0]['파이프길이(m)'])
            except:
                standard_length = 6.0
        else:
            standard_length = 6.0
        
        required_pipes = math.ceil(required_length_m / standard_length)
        
        return required_pipes    

    def _get_specification_with_length_fixed(self, material_name, standard, data):
        """규격에 파이프 길이 정보 추가"""
        bom_data = data.get('bom', pd.DataFrame())

        # BOM 데이터 유효성 검증
        if bom_data.empty or 'material_name' not in bom_data.columns:
            # BOM 데이터가 없거나 컬럼이 없으면 표준만 반환
            return standard

        material_bom = bom_data[bom_data['material_name'] == material_name]
        
        is_pipe = False
        if not material_bom.empty:
            category = str(material_bom.iloc[0]['category']).upper()
            is_pipe = 'PIPE' in category
        
        if is_pipe:
            main_materials = data['main_materials']
            
            pipe_match = main_materials[
                main_materials['규격'].astype(str).str.contains(str(standard), na=False, case=False)
            ]
            
            pipe_length = 6.0
            
            if not pipe_match.empty:
                if '파이프길이(m)' in pipe_match.columns:
                    try:
                        length_value = pipe_match.iloc[0]['파이프길이(m)']
                        if pd.notna(length_value) and length_value > 0:
                            pipe_length = float(length_value)
                    except:
                        pipe_length = 6.0
            
            return f"{standard}×{pipe_length}m"
        
        return standard
    
    def _render_inline_bom_editor(self, material_items):
        import pandas as pd

        rows = []
        for it in material_items:
            if it.get('is_header'):
                continue
            rows.append({
                "model_name": it.get("model_name",""),
                "material_name": it.get("material_name",""),
                "standard": it.get("standard",""),
                "unit": it.get("unit","EA"),
                "quantity": float(it.get("quantity",0)),
                "unit_price": float(it.get("unit_price",0)),
                "category": it.get("category",""),
            })

        base_df = pd.DataFrame(rows, columns=["model_name","material_name","standard","unit","quantity","unit_price","category"])

        st.caption("아래 표에서 직접 수정/추가하세요. (행 추가 버튼으로 새로운 자재를 추가할 수 있습니다)")
        edited = st.data_editor(
            base_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "model_name": st.column_config.TextColumn("모델명", help="이 행이 귀속될 모델"),
                "material_name": st.column_config.TextColumn("품목"),
                "standard": st.column_config.TextColumn("규격"),
                "unit": st.column_config.TextColumn("단위"),
                "quantity": st.column_config.NumberColumn("경간당 수량", step=0.1, min_value=0.0),
                "unit_price": st.column_config.NumberColumn("단가(원)", step=1, min_value=0.0, format="₩%.0f"),
                "category": st.column_config.TextColumn("카테고리"),
            }
        )
        edits = edited.fillna("").to_dict(orient="records")
        return edits

    def create_purchase_orders_by_material(self, quotation_data, delivery_location="현장", supplier_name=""):
        """재질별로 발주서 분리 생성"""
        try:
            data = self.load_data()
            purchase_items = self.generate_purchase_items_from_quotation(quotation_data)
            
            material_groups = self._group_by_material_type(purchase_items, data)
            
            purchase_orders = []
            
            for material_type, items in material_groups.items():
                actual_supplier_name = supplier_name if supplier_name.strip() else material_type
                
                supplier_info = {'company_name': actual_supplier_name}
                
                excel_buffer = self._create_single_purchase_order(
                    quotation_data, items, delivery_location, supplier_info
                )
                
                if excel_buffer:
                    purchase_orders.append({
                        'material_type': material_type,
                        'supplier': supplier_info['company_name'],
                        'excel_buffer': excel_buffer,
                        'items': items,
                        'filename': f"발주서_{supplier_info['company_name']}_{quotation_data['site_info']['site_name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    })
            
            return purchase_orders
            
        except Exception as e:
            st.error(f"재질별 발주서 생성 오류: {e}")
            return []

    def _group_by_material_type(self, purchase_items, data):
        """재질별로 발주 항목 그룹화"""
        material_groups = {}
        
        for item in purchase_items:
            category = str(item['category']).upper()
            
            if 'HGI' in category or '아연도' in category:
                material_type = '아연도'
            elif 'STS' in category:
                material_type = 'STS'
            else:
                material_type = self._find_material_type(item['material_name'], item['standard'], data)
            
            if material_type not in material_groups:
                material_groups[material_type] = []
            
            material_groups[material_type].append(item)
        
        return material_groups

    def _find_material_type(self, material_name, standard, data):
        """자재의 재질 타입 확인"""
        main_materials = data['main_materials']
        
        possible_item_columns = ['품목', 'Item', 'item_name', 'material_name', '자재명']
        possible_spec_columns = ['규격', 'Spec', 'specification', 'standard', '사양']
        
        item_column = None
        spec_column = None
        
        for col in possible_item_columns:
            if col in main_materials.columns:
                item_column = col
                break
        
        for col in possible_spec_columns:
            if col in main_materials.columns:
                spec_column = col
                break
        
        if item_column and spec_column:
            material_match = main_materials[
                (main_materials[item_column].str.contains(material_name, na=False)) |
                (main_materials[spec_column].str.contains(standard, na=False))
            ]
            
            if not material_match.empty:
                material_info = material_match.iloc[0]
                possible_material_columns = ['재질', 'Material', 'material_type', '소재']
                
                for mat_col in possible_material_columns:
                    if mat_col in material_info:
                        material_type = material_info[mat_col]
                        if 'STS' in str(material_type).upper():
                            return 'STS'
                        elif '아연도' in str(material_type):
                            return '아연도'
        
        return '아연도'

    def _create_single_purchase_order(self, quotation_data, purchase_items, delivery_location, supplier_info):
        """단일 발주서 생성"""
        try:
            template_path = resolve_template_path('발주서템플릿_v2.0_20250919.xlsx')
            workbook = load_workbook(template_path)
            sheet = workbook['발주서']
            
            today = datetime.now()
            sheet['F4'] = today.strftime('%Y년 %m월 %d일')
            sheet['B6'] = supplier_info['company_name']
            
            site_name = quotation_data['site_info']['site_name']
            start_row = 11
            
            data = self.load_data()
            
            for idx, purchase_item in enumerate(purchase_items):
                row = start_row + idx
                
                specification = self._get_specification_with_length_fixed(
                    purchase_item['material_name'], 
                    purchase_item['standard'], 
                    data
                )
                
                sheet[f'A{row}'] = idx + 1
                sheet[f'B{row}'] = purchase_item['material_name']
                sheet[f'C{row}'] = specification
                sheet[f'D{row}'] = purchase_item['unit']
                sheet[f'E{row}'] = purchase_item['quantity']
                sheet[f'F{row}'] = delivery_location
                sheet[f'G{row}'] = site_name
                sheet[f'H{row}'] = f"모델: {purchase_item['model_reference']}"
            
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            st.error(f"발주서 생성 오류: {e}")
            return None

    def create_template_quotation(self, quotation_data):
        """템플릿 기반 견적서 생성"""
        try:
            template_path = resolve_template_path('견적서템플릿_v2.0_20250919.xlsx')
            workbook = load_workbook(template_path)
            
            if quotation_data['contract_type'] == '사급':
                sheet = workbook['사급견적서']
                start_row = 13
                columns = {
                    'item': 'B', 'spec': 'C', 'unit': 'D', 'qty': 'E',
                    'price': 'F', 'supply': 'G', 'vat': 'H'
                }
            else:
                sheet = workbook['관급견적서']
                start_row = 14
                columns = {
                    'item': 'B', 'spec': 'D', 'unit': 'E', 'qty': 'F',
                    'price': 'G', 'amount': 'H', 'id_num': 'I'
                }
            
            plan = (quotation_data.get('site_info', {}) or {}).get('model_span_plan', {}) or {}

            data = self.load_data()
            models_df = data.get('models', pd.DataFrame())
            model_category_map = {}
            if not models_df.empty and 'model_name' in models_df.columns and 'category' in models_df.columns:
                for _, r in models_df.iterrows():
                    model_category_map[str(r['model_name'])] = str(r.get('category', ''))

            non_manual_items = [item for item in quotation_data['items'] if item.get('source') != 'MANUAL']

            for idx, item in enumerate(non_manual_items):
                row = start_row + idx

                model_name = item.get('model_name', '')
                category = model_category_map.get(model_name, '')

                qty_m = float((plan.get(model_name, {}) or {}).get('total_length_m', 0) or 0)

                sheet[f"{columns['item']}{row}"] = category if category else model_name
                spec = item.get('specification', '')
                if not spec:
                    spec = model_name
                sheet[f"{columns['spec']}{row}"] = spec
                sheet[f"{columns['unit']}{row}"] = 'm'
                sheet[f"{columns['qty']}{row}"] = round(qty_m, 2)
                sheet[f"{columns['price']}{row}"] = item.get('unit_price', 0)

                if quotation_data['contract_type'] == '관급' and 'id_num' in columns and '식별번호' in item:
                    sheet[f"{columns['id_num']}{row}"] = item['식별번호']

            try:
                sheet['F3'] = quotation_data['site_info']['site_name']
            except:
                pass
            
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            st.error(f"템플릿 견적서 생성 오류: {e}")
            return None

    def create_material_execution_interface(self):
        """자재 및 실행내역서 생성 + 인라인 BOM 편집"""
        st.subheader("📋 자재 및 실행내역서 자동생성")

        col1, col2 = st.columns(2)
        with col1:
            site_name = st.text_input("현장명", value="OO초등학교")
            contract_type = st.selectbox("계약 유형", ["관급", "사급"])
        with col2:
            foundation = st.selectbox("기초 유형", ["기초형", "앙카형"])
            delivery_date = st.date_input("납품기한", date.today() + timedelta(days=7))

        st.subheader("🔘 자재 항목 추가")
        if 'show_exec_editor' not in st.session_state:
            st.session_state.show_exec_editor = False
        if 'exec_buffer' not in st.session_state:
            st.session_state.exec_buffer = None
        if 'exec_items' not in st.session_state:
            st.session_state.exec_items = None
        if 'exec_site_info' not in st.session_state:
            st.session_state.exec_site_info = None
        if 'bom_mode' not in st.session_state:
            st.session_state.bom_mode = False
        if 'bom_input_rows' not in st.session_state:
            st.session_state.bom_input_rows = []
        if 'bom_edit_notice' not in st.session_state:
            st.session_state.bom_edit_notice = ""

        if 'material_items' not in st.session_state:
            st.session_state.material_items = []
        if 'current_selected_models' not in st.session_state:
            st.session_state.current_selected_models = []

        colm_a, colm_b = st.columns([3,2])
        with colm_a:
            st.markdown("**입력 모드 선택** · 세트당(BOM) 또는 기존 자재 입력")
        with colm_b:
            if not st.session_state.bom_mode:
                if st.button("🧩 BOM 입력 모드 열기", key=self._ukey("open_bom_mode")):
                    st.session_state.bom_mode = True
                    st.rerun()
            else:
                if st.button("❌ BOM 입력 모드 닫기", key=self._ukey("close_bom_mode")):
                    st.session_state.bom_mode = False
                    st.rerun()
        st.markdown("---")

        if not st.session_state.bom_mode:
            with st.expander("➕ 새 항목 추가", expanded=len(st.session_state.material_items)==0):
                st.markdown("**모델 검색 및 선택**")
                search_query = st.text_input(
                    "모델 검색",
                    placeholder="모델명, 식별번호, 차양, 볼라드, 자전거보관대 등 입력",
                    help="예: '디자인', 'DST', '24614649' 등",
                    key="material_search"
                )
                if search_query:
                    data = self.load_data()
                    if 'material_search_system' not in st.session_state:
                        st.session_state.material_search_system = EnhancedModelSearch(data['models'])
                    search_system = st.session_state.material_search_system
                    results = search_system.search_models(search_query)

                    if not results.empty:
                        st.write(f"🔍 검색 결과: {len(results)}개 모델")
                        if 'display_count' not in st.session_state:
                            st.session_state.display_count = 5
                        show_n = min(st.session_state.display_count, len(results))
                        st.markdown(f"**모델 선택 ({show_n}/{len(results)}개 표시):**")

                        selected_models = []
                        for idx in range(show_n):
                            row = results.iloc[idx]
                            c1, c2, c3 = st.columns([1, 3, 2])
                            with c1:
                                checked = st.checkbox("", key=self._ukey("pick", row['model_id'], idx))
                            with c2:
                                st.write(f"**{row['model_name']}**")
                                st.caption(f"{row['category']} | {row['model_standard']}")
                            with c3:
                                price = self.search_model_price(row['model_name'])
                                if price is not None:
                                    _ = st.success(f"💰 {int(price['단가']):,}원/{price['단위']}")
                                else:
                                    _ = st.warning("단가 없음")

                            if checked:
                                selected_models.append({
                                    'model_name': row['model_name'],
                                    'model_standard': row['model_standard']
                                })

                        colm1, colm2 = st.columns(2)
                        with colm1:
                            if len(results) > show_n:
                                if st.button(f"더보기 ({len(results)-show_n}개)", key="show_more"):
                                    st.session_state.display_count += 10
                                    st.rerun()
                        with colm2:
                            if show_n > 10:
                                if st.button("처음으로", key="reset_display"):
                                    st.session_state.display_count = 10
                                    st.rerun()

                        if selected_models:
                            st.write(f"✅ 선택된 모델: {len(selected_models)}개")
                            st.session_state.current_selected_models = selected_models
                            if st.button("📋 선택된 모델을 세트 목록에 추가", key="add_models_as_set"):
                                for m in selected_models:
                                    st.session_state.material_items.append({
                                        'model_name': m['model_name'],
                                        'quantity': 1.0,
                                        'notes': '세트기본',
                                        'source': 'SET_HEADER',
                                        'material_name': m['model_name'],
                                        'standard': m['model_standard'],
                                        'unit': 'SET',
                                        'category': 'MODEL'
                                    })
                                st.success(f"✅ {len(selected_models)}개 모델이 세트 목록에 추가되었습니다.")
                                st.rerun()
                        else:
                            st.session_state.current_selected_models = []

                st.markdown("---")
                st.subheader("🔎 부자재 검색 추가 (행별 경간당 수량 입력)")
                search_material = st.text_input(
                    "부자재 검색", 
                    placeholder="예: 볼트, 너트, 실리콘, M12, Ø10 등",
                    key="search_submaterial"
                )
                if search_material:
                    data = self.load_data()
                    sub_df = data['sub_materials'].copy()

                    if sub_df.empty or '품목' not in sub_df.columns:
                        st.warning("부자재 데이터가 없습니다.")
                    else:
                        mask_item = sub_df['품목'].astype(str).str.contains(search_material, case=False, na=False)
                        mask_spec = sub_df['규격'].astype(str).str.contains(search_material, case=False, na=False) if '규격' in sub_df.columns else False
                        search_results = sub_df[mask_item | mask_spec]

                        if not search_results.empty:
                            st.write(f"🔍 '{search_material}' 검색 결과: {len(search_results)}개")
                            for idx, (_, row) in enumerate(search_results.iterrows()):
                                material_name = str(row.get('품목', ''))
                                spec_display = str(row.get('규격', ''))
                                unit_display = str(row.get('단위', 'EA'))
                                unit_price = float(row.get('단가', 0) or 0)

                                col1, col2, col3, col4 = st.columns([3, 4, 3, 2])
                                with col1:
                                    st.write(f"**{material_name}**")
                                    st.caption(f"규격: {spec_display}")
                                with col2:
                                    st.caption(f"단위: {unit_display} | 단가: {int(unit_price):,}원")
                                with col3:
                                    qty = st.number_input(
                                        "경간당 수량",
                                        min_value=0.0, value=1.0, step=1.0,
                                        key=f"sub_per_span_qty_{idx}"
                                    )
                                with col4:
                                    if st.button("추가", key=f"add_sub_mat_{idx}"):
                                        selected_models = st.session_state.get('current_selected_models', [])
                                        if not selected_models:
                                            st.warning("먼저 모델을 하나 이상 선택하세요. (부자재는 모델 세트에 귀속됩니다)")
                                        else:
                                            for mdl in selected_models:
                                                st.session_state.material_items.append({
                                                    'model_name': mdl.get('model_name', ''),
                                                    'material_name': material_name,
                                                    'specification': spec_display,
                                                    'standard': spec_display,
                                                    'unit': unit_display or 'EA',
                                                    'unit_price': unit_price,
                                                    'category': 'MANUAL',
                                                    'quantity': qty,
                                                    'notes': '부자재검색추가',
                                                    'source': 'MANUAL'
                                                })
                                            st.success(f"✅ '{material_name}' (규격: {spec_display}) 이/가 선택된 모델 세트에 추가되었습니다.")
                                            st.rerun()

            if st.session_state.material_items:
                st.subheader("📋 자재 항목 목록")
                for i, item in enumerate(st.session_state.material_items):
                    row_key = self._ukey("mi_row", i, item.get('model_name',''), item.get('material_name',''), item.get('standard',''))
                    c1, c2, c3, c4 = st.columns([3, 1, 2, 1])
                    with c1:
                        st.text(f"{i+1}. {item['model_name']}")
                    with c2:
                        st.text(f"{item['quantity']:,}")
                    with c3:
                        st.text(item['notes'])
                    with c4:
                        if st.button("🗑️", key=f"mi_delete_{row_key}"):
                            st.session_state.material_items.pop(i)
                            st.rerun()

                st.markdown("---")
                st.subheader("📏 모델별 현장 길이 입력 (경간 자동계산)")

                selected_models = st.session_state.get('current_selected_models', [])
                plan = st.session_state.get('model_span_plan', {})

                for i, mdl in enumerate(selected_models):
                    mname = str(mdl.get('model_name',''))
                    mstd = str(mdl.get('model_standard',''))
                    width_m = parse_width_m_from_standard(mstd, DEFAULT_SPAN_WIDTH_M)
                    prev = plan.get(mname, {})
                    total_len_default = float(prev.get('total_length_m', 0.0))

                    c1, c2, c3, c4 = st.columns([3, 3, 3, 2])
                    with c1:
                        st.write(f"**{mname}**")
                        st.caption(f"규격: {mstd}")
                    with c2:
                        st.metric("추정 세트폭(m)", f"{width_m}")
                    with c3:
                        prev = st.session_state.get('model_span_plan', {}).get(mname, {})
                        total_length_default = float(prev.get('total_length_m', 0.0)) if prev else 0.0

                        total_len = st.number_input(
                            f"총 길이(m) — {mname}",
                            key=f"len_v091_{i}_{mname}_{mstd}_{self.tenant_id}",
                            min_value=0.0, value=total_length_default, step=0.5,
                            help="거래처가 m로 준 길이"
                        )

                    with c4:
                        import math
                        span_cnt = int(math.ceil(total_len / width_m)) if total_len > 0 else 0
                        st.metric("경간수", f"{span_cnt}")

                    plan[mname] = {
                        'width_m': float(width_m),
                        'total_length_m': float(total_len),
                        'span_count': int(span_cnt)
                    }

                st.session_state.model_span_plan = plan

                if st.button(
                    "📊 자재 및 실행내역서 생성",
                    key=self._ukey("gen_exec_report",
                                len(st.session_state.get('material_items', [])),
                                site_name, delivery_date, contract_type),
                    type="primary",
                    use_container_width=True
                ):
                    st.session_state.debug_messages = []

                    data = self.load_data()
                    models_df = data['models'].copy()
                    name_to_id = {}
                    if 'model_name' in models_df.columns and 'model_id' in models_df.columns:
                        for _, r in models_df.iterrows():
                            name_to_id[str(r['model_name']).strip()] = r['model_id']

                    for item in st.session_state.material_items:
                        if item.get('source') == 'MANUAL':
                            mname = str(item.get('model_name', '')).strip()
                            mid = name_to_id.get(mname)
                            if mid:
                                material_data = {
                                    'material_name': item.get('material_name', ''),
                                    'standard': item.get('standard', ''),
                                    'unit': item.get('unit', 'EA'),
                                    'quantity': float(item.get('quantity', 0)),
                                    'category': 'MANUAL',
                                    'notes': item.get('notes', '부자재검색추가'),
                                    'unit_price': float(item.get('unit_price', 0))
                                }
                                try:
                                    self.engine.add_bom_item(model_id=mid, material_data=material_data)
                                except:
                                    pass

                    site_info = {
                        'site_name': site_name,
                        'foundation': foundation,
                        'model_span_plan': st.session_state.get('model_span_plan', {})
                    }
                    quotation_data = self.generate_quotation(site_info, st.session_state.material_items, contract_type)

                    if quotation_data.get('items'):
                        excel_buffer, material_items = self.create_material_execution_report(quotation_data, delivery_date)
                        if excel_buffer:
                            st.session_state.exec_buffer = excel_buffer
                            st.session_state.exec_items = material_items
                            st.session_state.exec_site_info = site_info
                            st.session_state.last_material_data = quotation_data
                            st.session_state.show_exec_editor = True
                            st.rerun()
                    else:
                        st.warning("생성할 항목이 없습니다.")
                        
                if st.session_state.get('show_exec_editor') and st.session_state.get('exec_buffer') is not None:
                    _ = st.success("✅ 자재 및 실행내역서 생성 완료!")

                    material_items = st.session_state.exec_items
                    excel_buffer = st.session_state.exec_buffer

                    c_m1, c_m2 = st.columns(2)
                    with c_m1:
                        _ = st.metric("자재 종류", f"{len([i for i in material_items if not i.get('is_header')])}개")
                    with c_m2:
                        total_cost = sum(i.get('unit_price', 0) * i['quantity'] for i in material_items if not i.get('is_header'))
                        _ = st.metric("예상 자재비", f"{int(total_cost):,}원")

                    st.subheader("✏️ BOM 인라인 편집 (직접 반영)")
                    edits = self._render_inline_bom_editor(material_items)

                    filename = f"자재 및 실행내역서_{site_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

                    # 프로젝트 폴더 자동 생성 및 저장
                    def save_to_project_folder():
                        try:
                            # .env에서 프로젝트 폴더 경로 읽기
                            project_folder_path = os.getenv('PROJECT_FOLDER_PATH', 'downloads')

                            # 프로젝트명으로 폴더 생성
                            project_folder = Path(project_folder_path) / site_name
                            project_folder.mkdir(parents=True, exist_ok=True)

                            # 파일 저장
                            file_path = project_folder / filename
                            with open(file_path, 'wb') as f:
                                f.write(excel_buffer.getvalue())

                            st.success(f"✅ 파일이 저장되었습니다: {file_path}")
                        except Exception as e:
                            st.error(f"❌ 파일 저장 중 오류: {str(e)}")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            label="📥 자재 및 실행내역서 다운로드",
                            data=excel_buffer.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True,
                            key=self._ukey("download_exec_report", filename),
                            on_click=_phase3_record_bom_and_execution,
                            args=(self.tenant_id, material_items, st.session_state.get('last_material_data'))
                        )

                    with col2:
                        if st.button("💾 문서 보관", type="secondary", use_container_width=True, key=self._ukey("save_material_archive")):
                            success, msg = save_generated_document_to_archive(
                                self.engine.db,
                                self.tenant_id,
                                site_name,
                                "내역서",
                                excel_buffer,
                                username=st.session_state.get('user_id', self.tenant_id)
                            )
                            if success:
                                st.success(msg)
                            else:
                                st.error(msg)

                    # 프로젝트 폴더에 저장 버튼
                    if st.button("💾 프로젝트 폴더에 저장", key=self._ukey("save_to_folder", filename)):
                        save_to_project_folder()

                    st.markdown("---")
                    colx1, colx2 = st.columns([3,2])
                    with colx1:
                        st.info("⬆️ 위 표에서 수정/추가한 내용은 우측 버튼을 눌러야 BOM 시트에 반영됩니다.")
                    with colx2:
                        if st.button("💾 BOM 시트에 직접 반영", type="primary", key=self._ukey("apply_bom_edits", len(edits or []))):
                            ok = self._apply_bom_edits(edits)
                            if ok:
                                _ = st.success("✅ BOM 시트에 반영 완료 (앱 캐시 갱신)")
                                st.cache_data.clear()
                            else:
                                st.error("BOM 반영 실패. 로그를 확인하세요.")

                    if st.button("닫기", key=self._ukey("close_exec_editor")):
                        st.session_state.show_exec_editor = False
                        st.session_state.exec_buffer = None
                        st.session_state.exec_items = None
                        st.rerun()
        if st.session_state.bom_mode:
            st.subheader("🧩 BOM 입력 (세트당 수량)")

            data = self.load_data()
            model_names = data['models']['model_name'].astype(str).tolist()
            mcol1, mcol2 = st.columns([2,3])
            with mcol1:
                target_model = st.selectbox("대상 모델 선택", model_names, key=self._ukey("bom_target_model"))
            with mcol2:
                st.info("BOM 입력은 **세트(경간)당 수량** 기준입니다. (예: PIPE 2M/세트, 볼트 4EA/세트)")

            c1, c2, c3, c4, c5, c6 = st.columns([1.6,2,1.2,1.2,1.5,1.2])
            with c1:
                mat = st.text_input("품목", key=self._ukey("bom_in_mat"))
            with c2:
                std = st.text_input("규격", key=self._ukey("bom_in_std"))
            with c3:
                unit = st.selectbox("단위", ["EA","M","KG","M2"], key=self._ukey("bom_in_unit"))
            with c4:
                qty = st.number_input("세트당 수량", min_value=0.0, step=0.1, key=self._ukey("bom_in_qty"))
            with c5:
                price = st.number_input("단가(원)", min_value=0.0, step=1.0, key=self._ukey("bom_in_price"))
            with c6:
                cat = st.text_input("분류", value="MANUAL", key=self._ukey("bom_in_cat"))

            cc1, cc2, cc3 = st.columns([1,1,3])
            with cc1:
                if st.button("➕ 행 추가", key=self._ukey("bom_row_add", mat, std, unit, qty, price, cat)):
                    if target_model and mat and std and unit:
                        st.session_state.bom_input_rows.append({
                            "model_name": target_model,
                            "material_name": mat.strip(),
                            "standard": std.strip(),
                            "unit": unit.strip(),
                            "quantity": float(qty or 0.0),
                            "unit_price": float(price or 0.0),
                            "category": cat.strip(),
                            "notes": "BOM_INPUT_MODE"
                        })
                        st.session_state.bom_edit_notice = ""
                        st.rerun()
                    else:
                        st.session_state.bom_edit_notice = "모델/품목/규격/단위/단가를 입력해주세요."
                        st.rerun()
            with cc2:
                if st.button("🗑️ 전체 비우기", key=self._ukey("bom_rows_clear")):
                    st.session_state.bom_input_rows = []
                    st.rerun()
            with cc3:
                if st.session_state.bom_edit_notice:
                    st.warning(st.session_state.bom_edit_notice)

            if st.session_state.bom_input_rows:
                st.markdown("**현재 입력된 BOM(세트당):**")
                import pandas as pd
                prev_df = pd.DataFrame(st.session_state.bom_input_rows)
                
                st.dataframe(
                    prev_df, 
                    use_container_width=True,
                    column_config={
                        "unit_price": st.column_config.NumberColumn("단가(원)", format="₩%.0f")
                    }
                )

                st.markdown("---")
                colx1, colx2 = st.columns([3,2])
                with colx1:
                    st.info("아래 버튼을 누르면 **BOM 시트에 직접 반영**됩니다. (기존 행은 모델/품목/규격 기준으로 업데이트)")
                with colx2:
                    if st.button("💾 BOM 시트에 직접 반영", type="primary", key=self._ukey("bom_apply_now", len(prev_df))):
                        ok = self._apply_bom_edits(st.session_state.bom_input_rows)
                        if ok:
                            _ = st.success("✅ BOM 반영 완료")
                            st.session_state.bom_input_rows = []
                        else:
                            st.error("BOM 반영 실패. 로그를 확인하세요.")
            else:
                st.info("행을 추가해 BOM을 입력하세요.")

        else:
            st.info("자재 항목을 추가해주세요.")

    def _apply_bom_edits(self, edits):
        """편집 결과를 BOM 시트에 바로 반영"""
        import pandas as pd

        try:
            data = self.load_data()
            models_df = data['models'].copy()
            bom_df = data['bom'].copy()

            name_to_id = {}
            if 'model_name' in models_df.columns and 'model_id' in models_df.columns:
                for _, r in models_df.iterrows():
                    name_to_id[str(r['model_name']).strip()] = r['model_id']

            required_cols = ['model_id','material_name','standard','unit','quantity','category','notes']
            for c in required_cols:
                if c not in bom_df.columns:
                    bom_df[c] = "" if c not in ['quantity'] else 0.0

            for row in edits:
                mname = str(row.get("model_name","")).strip()
                mid = name_to_id.get(mname, None)
                if not mid:
                    st.warning(f"모델명 매핑 실패: '{mname}' (해당 행은 건너뜀)")
                    continue

                mat = str(row.get("material_name","")).strip()
                std = str(row.get("standard","")).strip()
                unit = str(row.get("unit","EA")).strip()
                qty = float(row.get("quantity", 0) or 0)
                cat = str(row.get("category","")).strip()

                if cat != "MANUAL":
                    if st.session_state.get("_DBG", False):
                        st.info(f"[DEBUG] 기존 BOM 데이터 건너뜀: {mat} (category: {cat})")
                    continue

                mask = (
                    (bom_df['model_id'] == mid) &
                    (bom_df['material_name'].astype(str).str.strip() == mat) &
                    (bom_df['standard'].astype(str).str.strip() == std)
                )
                if mask.any():
                    idx = bom_df.index[mask][0]
                    bom_df.at[idx, 'unit'] = unit
                    bom_df.at[idx, 'quantity'] = qty
                    bom_df.at[idx, 'unit_price'] = float(row.get("unit_price", 0) or 0)
                    if cat:
                        bom_df.at[idx, 'category'] = cat
                else:
                    new_row = {
                        'model_id': mid,
                        'material_name': mat,
                        'standard': std,
                        'unit': unit,
                        'quantity': qty,
                        'unit_price': float(row.get("unit_price", 0) or 0),
                        'category': cat if cat else "MANUAL",
                        'notes': "INLINE_EDIT"
                    }
                    bom_df = pd.concat([bom_df, pd.DataFrame([new_row])], ignore_index=True)

            try:
                edited_model_ids = set()
                for row in edits:
                    mname = str(row.get("model_name","")).strip()
                    mid = name_to_id.get(mname, None)
                    if mid:
                        edited_model_ids.add(mid)

                existing_items_by_model = {}

                for mid in edited_model_ids:
                    bom = self.engine.get_bom(mid)
                    if not bom.empty and 'category' in bom.columns:
                        manual_items = bom[bom['category'] == 'MANUAL']
                        existing_items_by_model[mid] = [
                            {
                                'material_name': str(row['material_name']).strip(),
                                'standard': str(row['standard']).strip()
                            }
                            for _, row in manual_items.iterrows()
                        ]
                    else:
                        existing_items_by_model[mid] = []

                edited_items_by_model = {}

                for row in edits:
                    mname = str(row.get("model_name","")).strip()
                    mid = name_to_id.get(mname, None)
                    if not mid:
                        continue

                    mat = str(row.get("material_name","")).strip()
                    std = str(row.get("standard","")).strip()
                    cat = str(row.get("category","")).strip()

                    if cat != "MANUAL":
                        continue

                    if not mat or not std:
                        continue

                    if mid not in edited_items_by_model:
                        edited_items_by_model[mid] = []

                    edited_items_by_model[mid].append({
                        'material_name': mat,
                        'standard': std,
                        'unit': str(row.get("unit","EA")).strip(),
                        'quantity': float(row.get("quantity", 0) or 0),
                        'category': cat,
                        'notes': 'INLINE_EDIT'
                    })

                deleted_count = 0
                for mid in edited_model_ids:
                    existing = existing_items_by_model.get(mid, [])
                    edited = edited_items_by_model.get(mid, [])

                    for exist_item in existing:
                        found = False
                        for edit_item in edited:
                            if (exist_item['material_name'] == edit_item['material_name'] and
                                exist_item['standard'] == edit_item['standard']):
                                found = True
                                break

                        if not found:
                            self.engine.delete_bom_item(
                                model_id=mid,
                                material_name=exist_item['material_name'],
                                standard=exist_item['standard']
                            )
                            deleted_count += 1

                added_count = 0
                for mid, items in edited_items_by_model.items():
                    existing = existing_items_by_model.get(mid, [])

                    for item in items:
                        is_new = True
                        for exist in existing:
                            if (exist['material_name'] == item['material_name'] and
                                exist['standard'] == item['standard']):
                                is_new = False
                                break

                        if is_new:
                            try:
                                self.engine.add_bom_item(model_id=mid, material_data=item)
                                added_count += 1
                            except Exception as e:
                                if 'duplicate' not in str(e).lower() and 'unique' not in str(e).lower():
                                    st.warning(f"항목 추가 실패: {item['material_name']} - {e}")

                messages = []
                if added_count > 0:
                    messages.append(f"✅ {added_count}개 항목 추가")
                if deleted_count > 0:
                    messages.append(f"🗑️ {deleted_count}개 항목 삭제")

                if messages:
                    st.success(" | ".join(messages))
                else:
                    st.info("변경 사항이 없습니다.")

                return True

            except Exception as e:
                st.error(f"BOM 저장 오류: {e}")
                import traceback
                st.error(f"상세 오류:\n{traceback.format_exc()}")
                return False

            return True

        except Exception as e:
            st.error(f"BOM 반영 오류: {e}")
            return False

    def create_purchase_order_interface(self):
        """발주서 생성 인터페이스"""
        st.header("📋 발주서 자동생성")

        if 'last_material_data' not in st.session_state:
            st.warning("먼저 자재발실행내역서를 생성해주세요. 자재 데이터를 기반으로 발주서가 생성됩니다.")
            return

        quotation_data = st.session_state.last_material_data

        # 안전한 데이터 접근
        site_name = safe_get_nested(quotation_data, ['site_info', 'site_name'], 'Unknown')
        item_count = len(safe_get(quotation_data, 'items', []))
        st.info(f"현장: {site_name} | 자재 항목: {item_count}개")
        
        col1, col2 = st.columns(2)
        with col1:
            delivery_location = st.text_input("하차지", value="공장")
        with col2:
            delivery_date = st.date_input("납품희망일", datetime.now() + pd.Timedelta(days=7))
        
        st.subheader("🔍 1단계: 발주 항목 분석")
        
        if st.button("📦 발주 항목 분석하기", type="secondary", use_container_width=True):
            with st.spinner("발주 항목 분석 중..."):
                purchase_items = self.generate_purchase_items_from_quotation(quotation_data)
                
                if purchase_items:
                    st.session_state.purchase_items = purchase_items
                    
                    categories = {}
                    for item in purchase_items:
                        category = item['category']
                        if category not in categories:
                            categories[category] = []
                        categories[category].append(item)
                    
                    st.session_state.analyzed_categories = categories
                    st.success(f"✅ 총 {len(purchase_items)}개 자재, {len(categories)}개 카테고리로 분류 완료!")
                    
                    for category, items in categories.items():
                        with st.expander(f"📂 {category} ({len(items)}개 항목)", expanded=True):
                            df_items = pd.DataFrame([
                                {
                                    '자재명': item['material_name'],
                                    '규격': item['standard'],
                                    '수량': f"{item['quantity']:,.1f}",
                                    '단위': item['unit'],
                                    '모델': item['model_reference']
                                }
                                for item in items
                            ])
                            st.dataframe(df_items, use_container_width=True)
                else:
                    st.warning("발주할 자재가 없습니다. BOM 데이터를 확인해주세요.")
        
        if hasattr(st.session_state, 'analyzed_categories'):
            st.subheader("🏭 2단계: 카테고리별 공급업체 선택 및 발주")
            
            categories = st.session_state.analyzed_categories
            
            for category, items in categories.items():
                with st.container():
                    st.markdown(f"### 📂 **{category}** 카테고리")
                    st.caption(f"자재 {len(items)}개 항목")
                    
                    col1, col2, col3 = st.columns([2, 2, 2])
                    
                    with col1:
                        supplier_name = st.text_input(
                            "공급업체명",
                            key=f"supplier_{category}",
                            placeholder="공급업체를 입력하세요.",
                            help="해당 카테고리 자재를 공급받을 업체명을 입력하세요"
                        )
                    
                    with col2:
                        category_delivery_date = st.date_input(
                            "납품요청일",
                            delivery_date,
                            key=f"delivery_{category}"
                        )
                    
                    with col3:
                        st.write("")
                        st.write("")
                        
                        if supplier_name.strip():
                            if st.button(
                                f"📋 {category} 발주서 생성",
                                key=f"create_order_{category}",
                                type="primary",
                                use_container_width=True
                            ):
                                self._create_category_purchase_order(
                                    category, items, supplier_name.strip(), 
                                    delivery_location, category_delivery_date, quotation_data
                                )
                        else:
                            st.button(
                                f"📋 {category} 발주서 생성",
                                key=f"create_order_{category}_disabled",
                                disabled=True,
                                use_container_width=True,
                                help="공급업체명을 입력하세요"
                            )
                    
                    st.markdown("---")

                    # 생성된 발주서 표시 (session state에서 읽음)
                    po_key = f"{category}_{supplier_name}"
                    if st.session_state.get(f"po_generated_{po_key}", False):
                        excel_buffer = st.session_state.get(f"po_excel_buffer_{po_key}")
                        quotation_data_stored = st.session_state.get(f"po_quotation_data_{po_key}")
                        items_stored = st.session_state.get(f"po_items_{po_key}")

                        if excel_buffer and quotation_data_stored:
                            site_name = quotation_data_stored['site_info']['site_name']
                            filename = f"발주서_{supplier_name}_{category}_{site_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

                            st.success(f"✅ {category} → {supplier_name} 발주서 생성 완료!")

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("카테고리", category)
                            with col2:
                                st.metric("자재 종류", f"{len(items_stored) if items_stored else 0}개")
                            with col3:
                                st.metric("공급업체", supplier_name)

                            col1, col2 = st.columns(2)
                            with col1:
                                st.download_button(
                                    label=f"📥 {supplier_name} ({category}) 발주서 다운로드",
                                    data=excel_buffer.getvalue(),
                                    file_name=filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"download_po_{category}_{supplier_name}",
                                    type="primary",
                                    use_container_width=True,
                                    on_click=_phase3_record_po,
                                    args=(self.tenant_id, category, supplier_name, items_stored, ((st.session_state.get('last_material_data') or {}).get('site_info') or {}).get('project_id'))
                                )

                            with col2:
                                if st.button("💾 문서 보관", type="secondary", use_container_width=True, key=f"save_po_archive_{category}_{supplier_name}"):
                                    success, msg = save_generated_document_to_archive(
                                        self.engine.db,
                                        self.tenant_id,
                                        site_name,
                                        "발주서",
                                        excel_buffer,
                                        username=st.session_state.get('user_id', self.tenant_id)
                                    )
                                    if success:
                                        st.success(msg)
                                    else:
                                        st.error(msg)

                            with st.expander(f"📋 {supplier_name} 발주 내역 상세", expanded=False):
                                if items_stored:
                                    df_order = pd.DataFrame([
                                        {
                                            '자재명': item['material_name'],
                                            '규격': item['standard'],
                                            '수량': f"{item['quantity']:,.1f}",
                                            '단위': item['unit'],
                                            '모델참조': item.get('model_reference', '')
                                        }
                                        for item in items_stored
                                    ])
                                    st.dataframe(df_order, use_container_width=True)

    def _create_category_purchase_order(self, category, items, supplier_name,
                                    delivery_location, delivery_date, quotation_data):
        """카테고리별 발주서 생성"""
        try:
            with st.spinner(f"{category} → {supplier_name} 발주서 생성 중..."):
                excel_buffer = self._create_single_purchase_order_by_category(
                    quotation_data, items, delivery_location,
                    {'company_name': supplier_name}, delivery_date
                )

                if excel_buffer:
                    # Session state에 버퍼 저장 (버튼 클릭 후에도 유지)
                    po_key = f"{category}_{supplier_name}"
                    st.session_state[f"po_excel_buffer_{po_key}"] = excel_buffer
                    st.session_state[f"po_category_{po_key}"] = category
                    st.session_state[f"po_supplier_{po_key}"] = supplier_name
                    st.session_state[f"po_quotation_data_{po_key}"] = quotation_data
                    st.session_state[f"po_items_{po_key}"] = items
                    st.session_state[f"po_generated_{po_key}"] = True

                    st.success(f"✅ {category} → {supplier_name} 발주서 생성 완료!")
                    st.rerun()

        except Exception as e:
            st.error(f"발주서 생성 중 오류: {str(e)}")

    def _create_single_purchase_order_by_category(self, quotation_data, purchase_items,
                                                delivery_location, supplier_info, delivery_date):
        """카테고리별 단일 발주서 생성"""
        try:
            template_path = resolve_template_path('발주서템플릿_v2.0_20250919.xlsx')
            workbook = load_workbook(template_path)
            sheet = workbook['발주서']

            today = datetime.now()
            sheet['F4'] = today.strftime('%Y년 %m월 %d일')

            # 안전한 supplier_info 접근
            company_name = safe_get(supplier_info, 'company_name', '미정')
            sheet['B6'] = company_name

            # 안전한 quotation_data 접근
            site_name = safe_get_nested(quotation_data, ['site_info', 'site_name'], 'Unknown')
            start_row = 11

            data = self.load_data()

            # purchase_items 검증
            if not isinstance(purchase_items, list):
                st.error("발주 항목 데이터가 유효하지 않습니다.")
                return None

            for idx, purchase_item in enumerate(purchase_items):
                # 각 항목의 필수 필드 검증
                if not validate_dict_keys(purchase_item, ['material_name', 'standard', 'unit', 'quantity']):
                    print(f"[WARNING] PO item {idx} missing required fields: {purchase_item}")
                    continue  # 불완전한 항목은 건너뜀

                try:
                    row = start_row + idx

                    specification = self._get_specification_with_length_fixed(
                        purchase_item['material_name'],
                        purchase_item['standard'],
                        data
                    )

                    sheet[f'A{row}'] = idx + 1
                    sheet[f'B{row}'] = purchase_item['material_name']
                    sheet[f'C{row}'] = specification
                    sheet[f'D{row}'] = purchase_item['unit']
                    sheet[f'E{row}'] = purchase_item['quantity']
                    sheet[f'F{row}'] = delivery_location
                    sheet[f'G{row}'] = site_name
                    sheet[f'H{row}'] = f"모델: {safe_get(purchase_item, 'model_reference', 'N/A')}"
                except Exception as e:
                    import traceback
                    print(f"[ERROR] Row {row} write failed")
                    print(f"[ERROR] purchase_item: {purchase_item}")
                    print(f"[ERROR] Exception: {e}")
                    print(f"[ERROR] Traceback: {traceback.format_exc()}")
                    continue

            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            return excel_buffer

        except Exception as e:
            st.error(f"발주서 생성 오류: {e}")
            import traceback
            print(f"[ERROR] PO generation exception: {traceback.format_exc()}")
            return None

    def create_quotation_interface(self):
        """견적서 생성 인터페이스"""
        st.header("💰 견적서 자동생성")

        if 'last_material_data' not in st.session_state:
            st.warning("먼저 자재 및 실행내역서를 생성해주세요. 해당 데이터를 기반으로 견적서가 생성됩니다.")
            return

        quotation_data = st.session_state.last_material_data

        # 안전한 데이터 접근
        site_name = safe_get_nested(quotation_data, ['site_info', 'site_name'], 'Unknown')
        item_count = len(safe_get(quotation_data, 'items', []))
        st.info(f"현장: {site_name} | 견적 항목: {item_count}개")
        
        col1, col2 = st.columns(2)
        with col1:
            contract_type = st.selectbox("계약 유형", ["관급", "사급"], key="quote_contract_type")
        with col2:
            quote_date = st.date_input("견적일자", datetime.now())
        
        if st.button("💰 견적서 생성", type="primary", use_container_width=True):
            quotation_data['contract_type'] = contract_type
            
            st.success("✅ 견적서 생성 완료!")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("공급가 (조달청 가격, 부가세 포함)", f"{quotation_data['total_supply_price']:,}원")
            with col2:
                st.metric("총 금액", f"{quotation_data['total_amount']:,}원")
            with col3:
                pass
            
            st.subheader("📄 견적 상세내역")
            detail_df = pd.DataFrame([
                {
                    '모델명': item['model_name'],
                    '규격': item['specification'],
                    '수량': f"{item['quantity']:,}{item['unit']}",
                    '단가': f"{item['unit_price']:,}원",
                    '금액': f"{item['supply_amount']:,}원"
                }
                for item in quotation_data['items']
            ])
            st.dataframe(detail_df, use_container_width=True)
            
            excel_buffer = self.create_template_quotation(quotation_data)

            if excel_buffer:
                filename = f"{self.tenant_config[self.tenant_id]['display_name']}견적서_{quotation_data['site_info']['site_name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

                # 다운로드 버튼 (기존 기능)
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="📥 템플릿 견적서 다운로드",
                        data=excel_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        on_click=_phase3_record_quotation,
                        args=(self.tenant_id, quotation_data)
                    )

                # 문서 보관 버튼 (document_archive 저장)
                with col2:
                    if st.button("💾 문서 보관", type="secondary", use_container_width=True, key="save_quotation_archive"):
                        project_name = quotation_data['site_info']['site_name']
                        success, msg = save_generated_document_to_archive(
                            self.engine.db,
                            self.tenant_id,
                            project_name,
                            "견적서",
                            excel_buffer,
                            username=st.session_state.get('user_id', self.tenant_id)
                        )
                        if success:
                            st.success(msg)
                        else:
                            st.error(msg)
            else:
                st.error("견적서 생성에 실패했습니다.")

    def create_independent_quotation_interface(self):
        """독립적인 견적서 작성 인터페이스"""
        st.header("💰 견적서 작성")

        if 'quotation_items' not in st.session_state:
            st.session_state.quotation_items = []
        if 'selected_quote_model' not in st.session_state:
            st.session_state.selected_quote_model = None

        recipient = st.text_input("수신", placeholder="예: OO건설", key="quote_recipient")

        installation_type = st.radio("설치 방식", ["기초형", "앙카형", "미정"], horizontal=True, key="installation_type")

        st.markdown("---")

        st.subheader("🔍 모델 검색 및 선택")

        search_query = st.text_input(
            "모델 검색",
            placeholder="모델명, 식별번호, 차양, 볼라드, 자전거보관대, 디자인형 등 입력",
            help="예: '디자인', 'DST', '24614649', '차양' 등",
            key="quote_model_search"
        )

        if search_query:
            data = self.load_data()
            if 'quote_search_system' not in st.session_state:
                st.session_state.quote_search_system = EnhancedModelSearch(data['models'])

            search_system = st.session_state.quote_search_system
            results = search_system.search_models(search_query)

            if not results.empty:
                st.write(f"🔍 **검색 결과: {len(results)}개 모델**")

                if 'quote_display_count' not in st.session_state:
                    st.session_state.quote_display_count = 5

                show_n = min(st.session_state.quote_display_count, len(results))
                st.caption(f"모델 선택 ({show_n}/{len(results)}개 표시):")

                for idx in range(show_n):
                    row = results.iloc[idx]

                    col1, col2, col3 = st.columns([0.5, 4.5, 2])

                    with col1:
                        is_selected = (
                            st.session_state.selected_quote_model is not None and
                            st.session_state.selected_quote_model.get('model_id') == row.get('model_id')
                        )

                        checkbox_value = st.checkbox("선택", value=is_selected, key=f"chk_quote_{row.get('model_id')}_{idx}", label_visibility="collapsed")

                        if checkbox_value != is_selected:
                            if checkbox_value:
                                price_info = self.search_model_price(row['model_name'])

                                if price_info is not None:
                                    unit_display = price_info.get('단위', 'EA')
                                    if unit_display == '㎡':
                                        unit_display = 'M2'
                                    elif unit_display in ['개', '조']:
                                        unit_display = 'EA'
                                    elif unit_display == 'm':
                                        unit_display = 'M'

                                    st.session_state.selected_quote_model = {
                                        'model_id': row.get('model_id'),
                                        'model_name': row.get('model_name'),
                                        'model_standard': row.get('model_standard'),
                                        'category': row.get('category'),
                                        'identifier_number': row.get('identifier_number', ''),
                                        'unit_price': price_info.get('단가', 0),
                                        'unit_display': unit_display
                                    }
                                    st.rerun()
                                else:
                                    st.error("가격 정보를 찾을 수 없습니다.")
                            else:
                                st.session_state.selected_quote_model = None
                                st.rerun()

                    with col2:
                        st.write(f"**{row['model_name']}**")
                        st.caption(f"{row['category']} | {row['model_standard']}")

                    with col3:
                        price = self.search_model_price(row['model_name'])
                        if price is not None:
                            unit_display = price.get('단위', 'EA')
                            if unit_display == '㎡':
                                unit_display = 'M2'
                            elif unit_display in ['개', '조']:
                                unit_display = 'EA'
                            elif unit_display == 'm':
                                unit_display = 'M'
                            st.success(f"💰 {int(price['단가']):,}원/{unit_display}")
                        else:
                            st.warning("단가 없음")

                col1, col2 = st.columns(2)
                with col1:
                    if len(results) > show_n:
                        if st.button(f"더보기 ({min(5, len(results)-show_n)}개)", key="quote_show_more"):
                            st.session_state.quote_display_count += 5
                            st.rerun()
                with col2:
                    if show_n > 5:
                        if st.button("처음으로", key="quote_reset_display"):
                            st.session_state.quote_display_count = 5
                            st.rerun()
            else:
                st.info("검색 결과가 없습니다.")

        st.markdown("---")
        st.subheader("➕ 새 항목 추가")

        if st.session_state.selected_quote_model:
            model = st.session_state.selected_quote_model
            unit_display = model.get('unit_display', 'EA')

            st.info(f"**선택:** {model.get('model_name', '')} ({model.get('model_standard', '')}) - {model.get('unit_price', 0):,}원/{unit_display}")

            quantity = st.number_input(
                f"수량 ({unit_display})",
                min_value=0.1,
                value=1.0,
                step=1.0 if unit_display == 'EA' else 0.1,
                key="quote_qty_fixed"
            )

            if st.button("➕ 견적에 추가", type="primary", use_container_width=True, key="add_to_quote_btn"):
                display_name = model.get('model_name', '')
                if installation_type == "기초형":
                    display_name = f"{display_name} (기초)"
                elif installation_type == "앙카형":
                    display_name = f"{display_name} (앙카)"

                item = {
                    'model_id': model.get('model_id', ''),
                    'model_name': display_name,
                    'category': model.get('category', ''),
                    'standard': model.get('model_standard', ''),
                    'identifier_number': model.get('identifier_number', ''),
                    'unit': unit_display,
                    'quantity': quantity,
                    'unit_price': model.get('unit_price', 0),
                    'amount': model.get('unit_price', 0) * quantity
                }
                st.session_state.quotation_items.append(item)
                st.session_state.selected_quote_model = None
                st.success(f"✅ {display_name} 추가 완료!")
                st.rerun()
        else:
            st.info("👆 위에서 모델을 검색하고 체크박스로 선택해주세요.")

        if st.session_state.quotation_items:
            st.markdown("---")
            st.subheader("📋 견적 내역")

            for idx, item in enumerate(st.session_state.quotation_items):
                col1, col2, col3, col4, col5, col6 = st.columns([3, 2, 1.5, 1, 2, 0.8])

                with col1:
                    st.text(item['model_name'])
                with col2:
                    st.text(item['standard'])
                with col3:
                    new_qty = st.number_input(
                        "수량",
                        value=item['quantity'],
                        min_value=0.0,
                        step=1.0 if item['unit'] == 'EA' else 0.1,
                        key=f"qty_{idx}",
                        label_visibility="collapsed"
                    )
                    if new_qty != item['quantity']:
                        st.session_state.quotation_items[idx]['quantity'] = new_qty
                        st.session_state.quotation_items[idx]['amount'] = item['unit_price'] * new_qty
                        st.rerun()
                with col4:
                    st.text(item['unit'])
                with col5:
                    st.text(f"{int(item['amount']):,}원")
                with col6:
                    if st.button("🗑️", key=f"del_{idx}"):
                        st.session_state.quotation_items.pop(idx)
                        st.rerun()

            total_sum = sum(item['amount'] for item in st.session_state.quotation_items)
            supply_price = round(total_sum / 1.1)
            vat = total_sum - supply_price

            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("공급가액", f"{int(supply_price):,}원")
            with col2:
                st.metric("부가세", f"{int(vat):,}원")
            with col3:
                st.metric("합계", f"{int(total_sum):,}원")

            st.markdown("---")
            col1, col2 = st.columns([1, 2])
            with col1:
                quotation_type = st.radio("견적서 타입", ["관급", "사급"], horizontal=True, key="quotation_type")
            with col2:
                if st.button("📄 견적서 생성", type="primary", use_container_width=True, key="generate_quote"):
                    excel_buffer = self.generate_quotation_excel(
                        st.session_state.quotation_items,
                        quotation_type,
                        recipient
                    )

                    if excel_buffer:
                        # Session state에 버퍼 저장 (버튼 클릭 후에도 유지)
                        st.session_state.ind_quote_excel_buffer = excel_buffer
                        st.session_state.ind_quote_recipient = recipient
                        st.session_state.ind_quote_generated = True
                    else:
                        st.error("견적서 생성에 실패했습니다.")

            # 생성된 견적서 표시 (session state에서 읽음)
            if st.session_state.get("ind_quote_generated", False):
                filename = f"견적서_{st.session_state.ind_quote_recipient}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                excel_buffer = st.session_state.ind_quote_excel_buffer

                col_dl, col_save = st.columns(2)
                with col_dl:
                    st.download_button(
                        label="📥 견적서 다운로드",
                        data=excel_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

                with col_save:
                    if st.button("💾 문서 보관", type="secondary", use_container_width=True, key="save_independent_quote_archive"):
                        success, msg = save_generated_document_to_archive(
                            self.engine.db,
                            self.tenant_id,
                            st.session_state.ind_quote_recipient,
                            "견적서",
                            excel_buffer,
                            username=st.session_state.get('user_id', self.tenant_id)
                        )
                        if success:
                            st.success(msg)
                        else:
                            st.error(msg)
        else:
            st.info("👆 위에서 모델을 검색하고 견적에 추가해주세요.")

    def generate_quotation_excel(self, items, quotation_type, recipient):
        """견적서 Excel 생성 (템플릿 기반)"""
        try:
            template_path = resolve_template_path("견적서템플릿_v2.0_20250919.xlsx")
            wb = openpyxl.load_workbook(template_path)

            if quotation_type == "관급":
                ws = wb["관급견적서"]
            else:
                ws = wb["사급견적서"]

            merged_cells_to_unmerge = []
            for merged_cell in ws.merged_cells:
                merged_cells_to_unmerge.append(merged_cell)

            for merged_cell in merged_cells_to_unmerge:
                if merged_cell.min_row >= 14:
                    ws.unmerge_cells(str(merged_cell))

            ws['A4'] = self.tenant_config[self.tenant_id]['display_name']

            ws['I4'] = date.today().strftime("%Y-%m-%d")

            ws['B6'] = recipient

            start_row = 14

            for idx, item in enumerate(items):
                row = start_row + idx

                if quotation_type == "관급":
                    ws[f'A{row}'] = idx + 1
                    ws[f'B{row}'] = item['category']
                    ws[f'C{row}'] = item['model_name']
                    ws[f'D{row}'] = item['standard']
                    ws[f'E{row}'] = item['unit']
                    ws[f'F{row}'] = item['quantity']
                    ws[f'G{row}'] = int(item['unit_price'])
                    ws[f'I{row}'] = item['identifier_number']
                else:
                    ws[f'A{row}'] = idx + 1
                    ws[f'B{row}'] = item['category']
                    ws[f'C{row}'] = item['standard']
                    ws[f'D{row}'] = item['unit']
                    ws[f'E{row}'] = item['quantity']
                    ws[f'F{row}'] = int(item['unit_price'])
                    supply = round(item['amount'] / 1.1)
                    vat = item['amount'] - supply
                    ws[f'G{row}'] = int(supply)
                    ws[f'H{row}'] = int(vat)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return buffer

        except Exception as e:
            st.error(f"견적서 생성 중 오류: {e}")
            return None


# 기존 검색 시스템 클래스들 유지
class EnhancedModelSearch:
    """고급 모델 검색 시스템"""
    
    def __init__(self, models_df):
        self.models_df = models_df
        self.search_columns = ['model_name', 'category', 'model_standard', '식별번호', 'description']

    def _ukey(self, scope, *parts):
        import re
        norm = [re.sub(r'[^0-9A-Za-z]+', '_', str(p)) for p in parts if p is not None]
        return "v091_search_" + scope + "_" + "_".join(norm)
    
    def search_models(self, query, max_results=50):
        """통합 검색 함수"""
        if not query or not query.strip():
            return self.models_df.head(20)
        
        query = query.strip()
        search_results = []
        
        if query.isdigit():
            results = self._search_by_identifier(query)
            search_results.extend(results)
        
        dimension_results = self._search_by_dimensions(query)
        search_results.extend(dimension_results)
        
        for column in self.search_columns:
            if column in self.models_df.columns:
                column_results = self._search_in_column(query, column)
                search_results.extend(column_results)
        
        unique_results = self._remove_duplicates_and_score(search_results, query)
        
        sorted_results = sorted(unique_results, key=lambda x: x['relevance_score'], reverse=True)
        
        if sorted_results:
            result_df = pd.DataFrame([item['model'] for item in sorted_results[:max_results]])
            return result_df
        else:
            return pd.DataFrame()
    
    def _search_by_identifier(self, query):
        """식별번호 검색"""
        results = []
        
        if '식별번호' in self.models_df.columns:
            mask = self.models_df['식별번호'].astype(str).str.contains(query, case=False, na=False)
            matched = self.models_df[mask]
            
            for _, row in matched.iterrows():
                results.append({
                    'model': row.to_dict(),
                    'match_type': 'identifier',
                    'match_column': '식별번호',
                    'match_value': str(row['식별번호'])
                })
        
        return results
    
    def _search_by_dimensions(self, query):
        """치수 기반 검색"""
        results = []

        patterns = [
            r'w(\d+)', r'width(\d+)', r'폭(\d+)',
            r'h(\d+)', r'height(\d+)', r'높이(\d+)',
            r'(\d+)w', r'(\d+)h'
        ]

        # 쿼리 정규화: x 를 * 로 변환
        normalized_query = self._normalize_search_string(query)
        query_lower = normalized_query.lower()
        extracted_numbers = []

        for pattern in patterns:
            matches = re.findall(pattern, query_lower)
            extracted_numbers.extend(matches)

        if query.isdigit() and int(query) >= 1000:
            extracted_numbers.append(query)

        if extracted_numbers and 'model_standard' in self.models_df.columns:
            for number in extracted_numbers:
                # DB 값도 정규화하여 비교 (* 로 통일)
                mask = self.models_df['model_standard'].astype(str).apply(
                    lambda x: number in self._normalize_search_string(x)
                )
                matched = self.models_df[mask]

                for _, row in matched.iterrows():
                    results.append({
                        'model': row.to_dict(),
                        'match_type': 'dimension',
                        'match_column': 'model_standard',
                        'match_value': row['model_standard']
                    })

        return results
    
    def _normalize_search_string(self, s: str) -> str:
        """검색 문자열 정규화: x 와 * 를 모두 * 로 통일"""
        if not isinstance(s, str):
            return str(s)
        # x, X 를 * 로 변환 (DB 표준 형식에 맞춤)
        return s.replace('x', '*').replace('X', '*').lower()

    def _search_in_column(self, query, column):
        """특정 컬럼에서 부분 검색 (정규화된 문자열 비교)"""
        results = []

        try:
            # 검색 쿼리 정규화
            normalized_query = self._normalize_search_string(query)

            # DB 값도 정규화하여 비교
            mask = self.models_df[column].astype(str).apply(
                lambda x: normalized_query in self._normalize_search_string(x)
            )
            matched = self.models_df[mask]

            for _, row in matched.iterrows():
                results.append({
                    'model': row.to_dict(),
                    'match_type': 'partial',
                    'match_column': column,
                    'match_value': str(row[column])
                })
        except Exception:
            pass

        return results
    
    def _remove_duplicates_and_score(self, search_results, query):
        """중복 제거 및 관련도 점수 계산"""
        unique_models = {}
        
        for result in search_results:
            model = result['model']
            model_id = model.get('model_id', '') or model.get('model_name', '')
            
            if model_id not in unique_models:
                relevance_score = self._calculate_relevance_score(result, query)
                
                unique_models[model_id] = {
                    'model': model,
                    'relevance_score': relevance_score,
                    'match_info': [result]
                }
            else:
                unique_models[model_id]['match_info'].append(result)
                new_score = self._calculate_relevance_score(result, query)
                if new_score > unique_models[model_id]['relevance_score']:
                    unique_models[model_id]['relevance_score'] = new_score
        
        return list(unique_models.values())
    
    def _calculate_relevance_score(self, result, query):
        """관련도 점수 계산"""
        score = 0
        query_lower = query.lower()
        
        type_scores = {
            'identifier': 100,
            'dimension': 80,
            'partial': 50
        }
        
        score += type_scores.get(result['match_type'], 0)
        
        column_weights = {
            'model_name': 30,
            'category': 20,
            'model_standard': 25,
            '식별번호': 35,
            'description': 10
        }
        
        score += column_weights.get(result['match_column'], 0)
        
        match_value = str(result['match_value']).lower()
        similarity = SequenceMatcher(None, query_lower, match_value).ratio()
        score += similarity * 50
        
        if query_lower in match_value or match_value in query_lower:
            score += 20
        
        return score


# 검색 인터페이스 함수들
def create_enhanced_search_interface(models_df, quotation_system, bom_df):
    """고급 검색 인터페이스"""
    
    if 'unified_search_system' not in st.session_state:
        st.session_state.unified_search_system = EnhancedModelSearch(models_df)
    
    search_system = st.session_state.unified_search_system
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        search_query = st.text_input(
            "통합 모델 검색",
            placeholder="모델명, 카테고리, 치수(W2000, H1200), 식별번호 등 입력",
            help="예: '디자인형', 'DAL', '2000', '24614649', 'W2000×H1200'",
            key="unified_search"
        )
    
    with col2:
        search_button = st.button("🔍 검색", use_container_width=True, key="unified_search_btn")
    
    if search_query or search_button:
        if search_query:
            with st.spinner("검색 중..."):
                search_results = search_system.search_models(search_query)
                
                if not search_results.empty:
                    st.success(f"검색 결과: {len(search_results)}개 모델 발견")
                    display_unified_search_results(search_results, search_query, quotation_system, bom_df)
                else:
                    st.warning("검색 결과가 없습니다. 다른 키워드로 시도해보세요.")
                    show_unified_search_tips()
        else:
            st.info("검색어를 입력해주세요.")
    else:
        st.subheader("전체 모델 목록 (처음 20개)")
        display_unified_search_results(models_df.head(20), "", quotation_system, bom_df)

def display_unified_search_results(results_df, search_query, quotation_system, bom_df):
    """검색 결과 표시"""
    
    for idx, (_, model) in enumerate(results_df.iterrows()):
        with st.expander(f"{model['model_name']} - {model['model_standard']}", expanded=False):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**모델 ID:** {model['model_id']}")
                st.write(f"**카테고리:** {model['category']}")
                st.write(f"**규격:** {model['model_standard']}")
            
            with col2:
                if pd.notna(model['식별번호']):
                    st.write(f"**식별번호:** {model['식별번호']}")
                st.write(f"**설명:** {model['description']}")
            
            price_info = quotation_system.search_model_price(model['model_name'])
            if price_info is not None:
                st.success(f"💰 단가: {price_info['단가']:,}원/{price_info['단위']}")
            else:
                st.warning("단가 정보 없음")
            
            model_bom = quotation_system.engine.get_bom(model['model_id'])
            if not model_bom.empty:
                st.write("**주요 자재:**")
                for _, bom_item in model_bom.head(3).iterrows():
                    st.write(f"- {bom_item['material_name']}: {bom_item['quantity']}{bom_item['unit']}")
            
            if search_query:
                highlight_unified_matches(model, search_query)

def highlight_unified_matches(model, search_query):
    """매칭된 부분 하이라이트"""
    
    matches = []
    query_lower = search_query.lower()
    
    search_fields = ['model_name', 'category', 'model_standard', '식별번호', 'description']
    
    for field_name in search_fields:
        if field_name in model and pd.notna(model[field_name]):
            field_value = str(model[field_name])
            if query_lower in field_value.lower():
                matches.append(f"{field_name}: {field_value}")
    
    if matches:
        st.info("🎯 매칭된 필드: " + ", ".join(matches[:2]))

def show_unified_search_tips():
    """검색 팁 표시"""
    
    st.info("🔍 검색 가이드")
    st.markdown("""
    **검색 방법:**
    - **모델명**: `DAL`, `DHART`, `DHWS`, `DST` 등
    - **카테고리**: `디자인형` 입력시 디자인형울타리 전체 검색
    - **치수**: `2000`, `1200`, `W2000`, `H1500` 등
    - **식별번호**: `24614649`, `25320309` 등 8자리 숫자
    - **복합 검색**: `DAL 2000` (DAL 시리즈 중 2000 폭)
    """)


# ============================================================================
# 문서 관리 기능 (검색, 다운로드, 업로드)
# ============================================================================

def parse_search_input(search_input: str) -> tuple[str, str]:
    """
    사용자 입력을 파싱하여 프로젝트명과 문서타입 분리

    입력: "샘플 견적서"
    출력: ("샘플", "quotation")

    지원 문서타입:
    - "견적서" → "quotation"
    - "발주서" → "po"
    - "내역서" → "bom"
    """
    search_lower = search_input.lower().strip()

    # 문서타입 매핑
    doc_type_map = {
        "견적서": "quotation",
        "발주서": "po",
        "내역서": "bom",
    }

    # 마지막 단어가 문서타입인지 확인
    for korean_type, english_type in doc_type_map.items():
        if search_lower.endswith(korean_type):
            project_name = search_input[:-(len(korean_type))].strip()
            return project_name, english_type

    # 문서타입을 찾지 못한 경우 전체를 프로젝트명으로 처리
    return search_input.strip(), ""


def search_documents(db, tenant_id: str, project_name: str, document_type: str) -> list:
    """
    DB에서 문서 검색

    Args:
        db: DatabaseManager 인스턴스
        tenant_id: 테넌트 ID
        project_name: 프로젝트명 (부분 검색)
        document_type: 문서타입 (quotation/po/bom)

    Returns:
        검색 결과 (딕셔너리 리스트)
    """
    try:
        query = db.table('document_archive').select('*').eq('tenant_id', tenant_id)

        # 프로젝트명 검색 (부분 일치)
        if project_name:
            query = query.ilike('project_name', f'%{project_name}%')

        # 문서타입 필터 (정확 일치)
        if document_type:
            query = query.eq('document_type', document_type)

        # 생성 날짜 역순 정렬
        query = query.order('created_at', desc=True)

        result = query.execute()
        return result.data or []
    except Exception as e:
        st.error(f"검색 중 오류 발생: {e}")
        return []


def generate_document_filename(db, project_name: str, document_type_korean: str) -> str:
    """
    DB 규칙에 맞는 파일명 생성

    규칙: {현장명}_{문서타입}_{날짜}_v{버전}.xlsx
    예: 샘플초등학교_견적서_251022_v01.xlsx

    Args:
        db: Supabase 클라이언트
        project_name: 현장명 (예: "샘플초등학교")
        document_type_korean: 문서타입 한글 (예: "견적서", "발주서", "내역서")

    Returns:
        생성된 파일명
    """
    from datetime import datetime

    date_str = datetime.now().strftime('%y%m%d')

    # 문서타입 영문 변환
    doc_type_map = {
        "견적서": "quotation",
        "발주서": "po",
        "내역서": "bom",
    }
    document_type_eng = doc_type_map.get(document_type_korean, "")

    try:
        # DB에서 같은 날짜의 같은 프로젝트/문서타입 파일 개수 확인
        base_filename = f"{project_name}_{document_type_korean}_{date_str}"
        results = db.table('document_archive').select('filename').ilike(
            'filename', f'{base_filename}%'
        ).execute()

        # 기존 파일 개수 + 1 = 다음 버전
        version = len(results.data) + 1 if results.data else 1
    except:
        version = 1

    return f"{project_name}_{document_type_korean}_{date_str}_v{version:02d}.xlsx"


def validate_filename(filename: str) -> tuple[bool, str, dict]:
    """
    파일명 규칙 검증: {현장명}_{문서타입}_{날짜}_v{버전}.xlsx

    예: 샘플초등학교_견적서_251022_v01.xlsx

    Returns:
        (유효여부, 에러메시지, 파싱된_데이터)
    """
    import re

    if not filename.endswith('.xlsx'):
        return False, "Excel 파일만 업로드 가능합니다 (.xlsx)", {}

    # 확장자 제거
    name_without_ext = filename[:-5]

    # 패턴: {현장명}_{문서타입}_{날짜}_v{버전}
    # 문서타입: 견적서, 발주서, 내역서
    pattern = r'^(.+?)_(견적서|발주서|내역서)_(\d{6})_v(\d+)$'
    match = re.match(pattern, name_without_ext)

    if not match:
        return False, (
            "파일명 규칙을 맞춰주세요.\n"
            "형식: {현장명}_{문서타입}_{날짜}_v{버전}.xlsx\n"
            "예: 샘플초등학교_견적서_251022_v01.xlsx"
        ), {}

    project_name, doc_type_korean, date_str, version = match.groups()

    # 한글 문서타입을 영문으로 변환
    doc_type_map = {
        "견적서": "quotation",
        "발주서": "po",
        "내역서": "bom",
    }
    doc_type_eng = doc_type_map.get(doc_type_korean, "")

    return True, "", {
        "project_name": project_name,
        "document_type": doc_type_eng,
        "doc_type_korean": doc_type_korean,
        "date_str": date_str,
        "version": int(version),
    }


def upload_document_to_archive(
    db,
    storage_manager,
    tenant_id: str,
    username: str,
    file_bytes: bytes,
    filename: str,
    parsed_data: dict
):
    """
    파일을 Storage + DB에 저장

    Returns:
        (성공여부, 메시지)
    """
    try:
        # Storage에 업로드 (sanitized 경로 반환)
        success, result = storage_manager.upload_file(
            tenant_id=tenant_id,
            document_type=parsed_data['document_type'],
            document_id=parsed_data['project_name'],
            file_bytes=file_bytes,
            filename=filename
        )

        if not success:
            return False, f"Storage 업로드 실패: {result}"

        # result는 storage_manager가 반환한 sanitized 경로
        storage_path = result

        # DB에 메타데이터 저장
        import uuid
        from datetime import datetime

        archive_data = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "project_id": parsed_data['project_name'],
            "project_name": parsed_data['project_name'],
            "document_type": parsed_data['document_type'],
            "storage_path": storage_path,
            "filename": filename,
            "file_size": len(file_bytes),
            "created_at": datetime.utcnow().isoformat(),
            "created_by": username,
            "updated_at": datetime.utcnow().isoformat(),
        }

        response = db.table('document_archive').insert(archive_data).execute()

        if not response.data:
            # Storage에 업로드된 파일 삭제
            storage_manager.delete_file(storage_path)
            return False, "DB 저장 실패"

        return True, f"✅ {filename} 업로드 완료"
    except Exception as e:
        return False, f"업로드 중 오류: {str(e)}"


def delete_document_from_archive(
    db,
    storage_manager,
    document_id: str,
    storage_path: str
):
    """
    문서를 Storage + DB에서 삭제

    Returns:
        (성공여부, 메시지)
    """
    try:
        # Storage에서 삭제
        if not storage_manager.delete_file(storage_path):
            return False, "Storage 삭제 실패"

        # DB에서 삭제
        response = db.table('document_archive').delete().eq('id', document_id).execute()

        if response.data:
            return True, "✅ 문서가 삭제되었습니다"
        else:
            return False, "DB 삭제 실패"
    except Exception as e:
        return False, f"삭제 중 오류: {str(e)}"


def save_generated_document_to_archive(
    db,
    tenant_id: str,
    project_name: str,
    document_type_korean: str,
    file_buffer,
    username: str = None
) -> tuple[bool, str]:
    """
    생성된 문서를 document_archive에 저장

    Args:
        db: Supabase 클라이언트 (qs.engine.db)
        tenant_id: 테넌트 ID
        project_name: 현장명
        document_type_korean: 문서타입 한글 ("견적서", "발주서", "내역서")
        file_buffer: BytesIO 파일 버퍼
        username: 생성자명 (기본값: tenant_id)

    Returns:
        (성공여부, 메시지)
    """
    try:
        from app.storage_manager import get_storage_manager
        import uuid
        from datetime import datetime

        storage_manager = get_storage_manager()

        # 파일명 생성 (DB 규칙, 버전 자동 증가)
        filename = generate_document_filename(db, project_name, document_type_korean)

        # 파일 바이트 추출
        if hasattr(file_buffer, 'getvalue'):
            file_bytes = file_buffer.getvalue()
        else:
            file_bytes = file_buffer

        # 문서타입 영문 변환
        doc_type_map = {
            "견적서": "quotation",
            "발주서": "po",
            "내역서": "bom",
        }
        document_type_eng = doc_type_map.get(document_type_korean, "")

        if not document_type_eng:
            return False, f"지원하지 않는 문서타입: {document_type_korean}"

        # Storage에 업로드 (sanitized 경로 반환)
        success, result = storage_manager.upload_file(
            tenant_id=tenant_id,
            document_type=document_type_eng,
            document_id=project_name,
            file_bytes=file_bytes,
            filename=filename
        )

        if not success:
            return False, f"Storage 업로드 실패: {result}"

        # result는 storage_manager가 반환한 sanitized 경로
        storage_path = result

        # DB에 메타데이터 저장
        archive_data = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "project_id": project_name,
            "project_name": project_name,
            "document_type": document_type_eng,
            "storage_path": storage_path,
            "filename": filename,
            "file_size": len(file_bytes),
            "created_at": datetime.utcnow().isoformat(),
            "created_by": username or tenant_id,
            "updated_at": datetime.utcnow().isoformat(),
        }

        response = db.table('document_archive').insert(archive_data).execute()

        if not response.data:
            storage_manager.delete_file(storage_path)
            error_msg = getattr(response, 'error', 'Unknown error')
            return False, f"DB 저장 실패: {error_msg}"

        return True, f"✅ {filename} 저장 완료"

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return False, f"문서 저장 중 오류: {str(e)}\n{error_details}"


# 메인 애플리케이션
def main(mode="pilot"):
    # Initialize session state for debug messages
    if 'debug_messages' not in st.session_state:
        st.session_state.debug_messages = []

    # 테넌트 ID 가져오기
    tenant_id = get_tenant_from_params()
    
    # 테넌트 정보 표시
    tenant_config = {
        'dooho': {'name': '두호', 'display_name': '두호'},
        'kukje': {'name': '국제', 'display_name': '국제'}
    }
    
    tenant_info = tenant_config.get(tenant_id, tenant_config['dooho'])

    st.markdown("---")
    
    # 업체 변경 UI (사이드바) - 파일럿 모드일 때만 표시
    if mode == "pilot":
        with st.sidebar:
            st.subheader("🏢 업체 변경")

            # tenant_id는 스크립트 시작 부분에서 URL 파라미터로 계산된 값
            current_tenant_name = tenant_config[tenant_id]['display_name']
            st.info(f"**현재 업체**: {current_tenant_name}")

            col1, col2 = st.columns(2)
            with col1:
                if st.button("두호", use_container_width=True, disabled=(tenant_id == 'dooho')):
                    st.query_params['tenant'] = 'dooho'
                    st.rerun()
            with col2:
                if st.button("국제", use_container_width=True, disabled=(tenant_id == 'kukje')):
                    st.query_params['tenant'] = 'kukje'
                    st.rerun()

    # --- DEBUG MESSAGE DISPLAY ---
    if st.session_state.get('debug_messages'):
        st.subheader("🐞 디버그 메시지")
        with st.expander("메시지 보기", expanded=True):
            for msg in st.session_state.debug_messages:
                st.warning(msg)
            if st.button("디버그 메시지 지우기"):
                st.session_state.debug_messages = []
                st.rerun()
    
    # 시스템 초기화
    if 'qs' not in st.session_state or st.session_state.get('current_tenant') != tenant_id:
        st.session_state.qs = UnifiedQuotationSystem(tenant_id)
    
    qs = st.session_state.qs
    data = qs.load_data()
    
    if not data:
        st.error("데이터베이스를 로딩할 수 없습니다. 파일을 확인해주세요.")
        return
    
    # 사이드바 - 데이터베이스 현황
    with st.sidebar:
        st.header("📊 데이터베이스 현황")
        st.metric("모델 수", len(data['models']))
        st.metric("단가 정보", len(data['pricing']))
        st.metric("BOM 항목", len(data['bom']))
        
        st.header("🏢 회사 정보")
        st.info(f'**회사명**\n{tenant_info["display_name"]}\n금속구조물\n제작 설치 전문업체')
       
        if len(data['models']) > 0:
            model_prefixes = {}
            for _, model in data['models'].iterrows():
                prefix = model['model_name'].split('-')[0][:4]
                model_prefixes[prefix] = model_prefixes.get(prefix, 0) + 1
            
            st.header("🗂️ 모델 시리즈")
            top_series = sorted(model_prefixes.items(), key=lambda x: x[1], reverse=True)[:5]
            for prefix, count in top_series:
                st.write(f"• {prefix}***: {count}개")
    
    # 메인 영역 - 작업 순서: 자재 및 실행내역서 → 발주서 → 견적서
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "💰 견적서 작성", "📊 자재 및 실행내역서", "📋 발주서 생성",
        "🔍 모델 조회", "📦 재고 현황", "📊 BOM 분석"
    ])

    with tab1:
        qs.create_independent_quotation_interface()

    with tab2:
        qs.create_material_execution_interface()

    with tab3:
        qs.create_purchase_order_interface()
    
    with tab4:
        st.header("🔍 모델 조회")
        
        create_enhanced_search_interface(data['models'], qs, data['bom'])

    with tab5:
        st.header("📦 재고 현황")

        if data['inventory'].empty:
            st.info("재고 데이터가 없습니다.")
        elif '잔여재고' not in data['inventory'].columns:
            st.warning("재고 데이터 컬럼 구조가 올바르지 않습니다.")
            st.write("현재 컬럼:", list(data['inventory'].columns))
        else:
            col1, col2, col3 = st.columns(3)
            with col1:
                total_items = len(data['inventory'])
                st.metric("총 자재 종류", f"{total_items}개")
            with col2:
                total_stock = data['inventory']['잔여재고'].sum()
                st.metric("총 잔여재고", f"{total_stock:,}EA")
            with col3:
                low_stock = len(data['inventory'][data['inventory']['잔여재고'] < 5])
                st.metric("재고 부족(5개 미만)", f"{low_stock}개", delta_color="inverse")

        st.subheader("📋 자재별 재고 현황")

        if not data['inventory'].empty and '잔여재고' in data['inventory'].columns:
            def create_full_specification(row):
                spec = str(row.get('규격', ''))
                
                if '두께' in row and pd.notna(row['두께']):
                    spec += f"×{row['두께']}"
                
                if '파이프길이(m)' in row and pd.notna(row['파이프길이(m)']):
                    spec += f"×{row['파이프길이(m)']}m"
                    
                return spec
            
            inventory_display = data['inventory'].copy()
            inventory_display['완전규격'] = inventory_display.apply(create_full_specification, axis=1)

            display_columns = ['item_id', '재질', '완전규격', '잔여재고', '단위', '단가']

            available_columns = [col for col in display_columns if col in inventory_display.columns]
            if '단가' not in inventory_display.columns:
                available_columns = [col for col in available_columns if col != '단가']

            final_display = inventory_display[available_columns].copy()

            column_rename = {
                'item_id': '자재ID',
                '재질': '재질',
                '완전규격': '규격',
                '잔여재고': '잔여재고',
                '단위': '단위',
                '단가': '단가(원)'
            }

            final_display = final_display.rename(columns={k:v for k,v in column_rename.items() if k in final_display.columns})

            def highlight_low_stock(val):
                if isinstance(val, (int, float)) and val < 5:
                    return 'background-color: #ffcccc'
                return ''

            final_display = final_display.sort_values('잔여재고' if '잔여재고' in final_display.columns else final_display.columns[-2])

            styled_df = final_display.style.applymap(
                highlight_low_stock,
                subset=['잔여재고'] if '잔여재고' in final_display.columns else []
            )

            st.dataframe(styled_df, use_container_width=True)

            if '잔여재고' in data['inventory'].columns:
                low_stock_count = len(data['inventory'][data['inventory']['잔여재고'] < 5])
                if low_stock_count > 0:
                    st.warning(f"⚠️ {low_stock_count}개 자재의 재고가 5개 미만입니다. 발주를 검토해주세요.")

                low_stock_items = data['inventory'][data['inventory']['잔여재고'] < 5]
                if not low_stock_items.empty:
                    with st.expander("재고 부족 자재 상세"):
                        for _, item in low_stock_items.iterrows():
                            full_spec = create_full_specification(item)
                            st.write(f"- {item.get('재질', 'N/A')} {full_spec}: {item['잔여재고']}개 남음")
    
    with tab6:
        st.header("📊 BOM 분석")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📦 자재 카테고리별 분포")
            st.info("이 기능은 현재 비활성화되어 있습니다.")
        
        with col2:
            st.subheader("🔍 특정 모델 BOM 조회")
            if 'models' in data and not data['models'].empty and 'model_name' in data['models'].columns:
                model_list = data['models']['model_name'].tolist()
                selected_model_for_bom = st.selectbox(
                    "모델 선택", 
                    model_list,
                    key="bom_model"
                )
                
                if selected_model_for_bom:
                    model_info_df = data['models'][data['models']['model_name'] == selected_model_for_bom]
                    if not model_info_df.empty:
                        model_info = model_info_df.iloc[0]
                        model_bom = qs.engine.get_bom(model_info['model_id'])
                        
                        if not model_bom.empty:
                            st.write(f"**{selected_model_for_bom}** 자재 구성:")
                            display_bom = model_bom.rename(columns={
                                'material_name': '자재명',
                                'standard': '규격',
                                'quantity': '수량',
                                'unit': '단위'
                            })
                            st.dataframe(display_bom[['자재명', '규격', '수량', '단위']])
                        else:
                            st.info("해당 모델의 BOM 정보가 없습니다.")
                    else:
                        st.warning("선택된 모델 정보를 찾을 수 없습니다.")
            else:
                st.warning("모델 데이터가 로드되지 않았습니다.")

if __name__ == "__main__":
    main()
