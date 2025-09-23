from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() #새로운 시트 기존 이름으로 생성
ws.title = "Mysheet" #시트 이름이 변경
ws.sheet_properties.tabColor = "3399ff" #RGB형태로 

#Sheet, MySheet, YoirSheet
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index 에 생성

new_ws = wb["NewSheet"] #딕셔너리 형태로 sheet에 접근이 가능

print(wb.sheetnames)

wb.save("sample.xlsx")